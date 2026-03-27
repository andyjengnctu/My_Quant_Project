import os
import sys
import copy
import time
import importlib.util
import pandas as pd

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from core.v16_params_io import load_params_from_json, build_params_from_mapping, params_to_json_dict
from core.v16_core import run_v16_backtest, calc_reference_candidate_qty, calc_entry_price, can_execute_half_take_profit, resize_candidate_plan_to_capital, build_cash_capped_entry_plan, evaluate_history_candidate_metrics
from core.v16_buy_sort import calc_buy_sort_value
from core.v16_config import BUY_SORT_METHOD, V16StrategyParams
from core.v16_portfolio_engine import (
    prep_stock_data_and_trades,
    pack_prepared_stock_data,
    get_fast_dates,
    run_portfolio_timeline,
    find_sim_start_idx,
)
from core.v16_data_utils import sanitize_ohlcv_dataframe, get_required_min_rows, discover_unique_csv_map, discover_unique_csv_inputs
from core.v16_log_utils import format_exception_summary
from contextlib import redirect_stdout, redirect_stderr
import tempfile
import io


OUTPUT_DIR = os.path.join(PROJECT_ROOT, "outputs")
DATA_DIR = os.path.join(PROJECT_ROOT, "tw_stock_data_vip")
PARAMS_FILE = os.path.join(PROJECT_ROOT, "models", "v16_best_params.json")

FLOAT_TOL = 1e-6
VALIDATE_PROGRESS_EVERY = 25
MAX_CONSOLE_FAIL_PREVIEW = 20


VALIDATION_RECOVERABLE_EXCEPTIONS = (
    AssertionError,
    ArithmeticError,
    AttributeError,
    ImportError,
    LookupError,
    NameError,
    OSError,
    RuntimeError,
    TypeError,
    ValueError,
    pd.errors.EmptyDataError,
    pd.errors.ParserError,
)

MODULE_LOAD_RECOVERABLE_EXCEPTIONS = VALIDATION_RECOVERABLE_EXCEPTIONS + (
    SyntaxError,
)


CSV_PATH_CACHE = None
CSV_DUPLICATE_ISSUES = None
CSV_PATH_CACHE_DATA_DIR = None


def get_data_dir_csv_map():
    global CSV_PATH_CACHE, CSV_DUPLICATE_ISSUES, CSV_PATH_CACHE_DATA_DIR

    resolved_data_dir = os.path.abspath(DATA_DIR)
    if (CSV_PATH_CACHE is None) or (CSV_PATH_CACHE_DATA_DIR != resolved_data_dir):
        CSV_PATH_CACHE, CSV_DUPLICATE_ISSUES = discover_unique_csv_map(resolved_data_dir)
        CSV_PATH_CACHE_DATA_DIR = resolved_data_dir
    return CSV_PATH_CACHE


# # (AI註: consistency test 與 portfolio engine 共用相同模擬起點，避免完整年度/年化交易次數驗證口徑漂移)
def calc_validation_sim_years(sorted_dates, start_year):
    if not sorted_dates:
        return 0.0

    start_idx = find_sim_start_idx(sorted_dates, start_year)
    if start_idx >= len(sorted_dates):
        return 0.0

    first_dt = pd.Timestamp(sorted_dates[start_idx])
    last_dt = pd.Timestamp(sorted_dates[-1])
    span_days = (last_dt - first_dt).days + 1
    if span_days <= 0:
        return 0.0
    return span_days / 365.25


# # (AI註: 用 CAGR 驗證 annual return 指標，避免只驗證總報酬卻漏掉年化口徑)
def calc_validation_annual_return_pct(start_value, end_value, years):
    if start_value <= 0 or years <= 0:
        return 0.0
    if end_value <= 0:
        return -100.0
    return ((end_value / start_value) ** (1.0 / years) - 1.0) * 100.0


def load_params():
    return load_params_from_json(PARAMS_FILE)


def normalize_ticker_text(value):
    if pd.isna(value):
        return ""

    text = str(value).strip()

    if text[:-2].isdigit() and text.endswith(".0"):
        text = text[:-2]

    if text.isdigit() and len(text) < 4:
        text = text.zfill(4)

    return text

def is_insufficient_data_error(exc):
    return isinstance(exc, ValueError) and ("有效資料不足" in str(exc))

def make_consistency_params(base_params):
    params = copy.deepcopy(base_params)
    params.min_history_trades = 0
    params.min_history_ev = -1e9
    params.min_history_win_rate = 0.0
    return params


def discover_available_tickers():
    if not os.path.isdir(DATA_DIR):
        return []

    return sorted(get_data_dir_csv_map().keys())


def resolve_csv_path(ticker):
    csv_map = get_data_dir_csv_map()
    if ticker in csv_map:
        return csv_map[ticker]

    candidates = [
        os.path.join(DATA_DIR, f"TV_Data_Full_{ticker}.csv"),
        os.path.join(DATA_DIR, f"{ticker}.csv"),
        os.path.join(PROJECT_ROOT, f"TV_Data_Full_{ticker}.csv"),
        os.path.join(PROJECT_ROOT, f"{ticker}.csv"),
    ]
    raise FileNotFoundError(f"找不到 {ticker} 的 CSV。已檢查: {candidates}")


def load_clean_df(ticker, params):
    file_path = resolve_csv_path(ticker)
    raw_df = pd.read_csv(file_path)
    min_rows_needed = get_required_min_rows(params)
    df, sanitize_stats = sanitize_ohlcv_dataframe(raw_df, ticker, min_rows=min_rows_needed)
    return file_path, df, sanitize_stats


def add_check(results, module_name, ticker, metric, expected, actual, tol=FLOAT_TOL, note=""):
    if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
        diff = abs(float(expected) - float(actual))
        passed = diff <= tol
    else:
        diff = None
        passed = expected == actual

    status = "PASS" if passed else "FAIL"

    results.append({
        "ticker": ticker,
        "module": module_name,
        "metric": metric,
        "expected": expected,
        "actual": actual,
        "abs_diff": diff,
        "passed": passed,
        "status": status,
        "note": note
    })

def add_skip_result(results, module_name, ticker, metric, note):
    results.append({
        "ticker": ticker,
        "module": module_name,
        "metric": metric,
        "expected": "SKIP",
        "actual": "SKIP",
        "abs_diff": None,
        "passed": True,
        "status": "SKIP",
        "note": note
    })


def add_fail_result(results, module_name, ticker, metric, expected, actual, note):
    results.append({
        "ticker": ticker,
        "module": module_name,
        "metric": metric,
        "expected": expected,
        "actual": actual,
        "abs_diff": None,
        "passed": False,
        "status": "FAIL",
        "note": note
    })


# # (AI註: validate 的 single_vs_portfolio 要驗證「原始執行邏輯一致性」，
# # (AI註: 不應混入 portfolio 的歷史績效濾網，否則會把設計上的選股層差異誤判成執行層 bug)
def build_execution_only_params(params):
    return make_consistency_params(params)


# # (AI註: 將年度報酬明細正規化，避免 Timestamp / numpy scalar 造成 validate 誤判)
def normalize_yearly_return_rows(rows):
    normalized = []
    for row in rows or []:
        normalized.append({
            "year": int(row.get("year", 0)),
            "year_return_pct": float(row.get("year_return_pct", 0.0)),
            "is_full_year": bool(row.get("is_full_year", False)),
            "start_date": str(row.get("start_date", "")),
            "end_date": str(row.get("end_date", "")),
        })
    return normalized


# # (AI註: 年度統計需同時驗欄位值與 yearly rows 內部一致性，避免只有 aggregate 對、明細卻漂移)
def extract_yearly_profile_fields(profile_stats):
    yearly_rows = normalize_yearly_return_rows(profile_stats.get("yearly_return_rows", []))
    bm_yearly_rows = normalize_yearly_return_rows(profile_stats.get("bm_yearly_return_rows", []))
    return {
        "full_year_count": int(profile_stats.get("full_year_count", 0)),
        "min_full_year_return_pct": float(profile_stats.get("min_full_year_return_pct", 0.0)),
        "yearly_return_rows": yearly_rows,
        "bm_full_year_count": int(profile_stats.get("bm_full_year_count", 0)),
        "bm_min_full_year_return_pct": float(profile_stats.get("bm_min_full_year_return_pct", 0.0)),
        "bm_yearly_return_rows": bm_yearly_rows,
    }


def calc_expected_full_year_metrics(yearly_rows):
    full_year_rows = [row for row in (yearly_rows or []) if row["is_full_year"]]
    return {
        "full_year_count": len(full_year_rows),
        "min_full_year_return_pct": float(min((row["year_return_pct"] for row in full_year_rows), default=0.0)),
    }


def summarize_portfolio_trade_output(df_trades):
    portfolio_trade_types = df_trades["Type"].fillna("") if df_trades is not None and len(df_trades) > 0 and "Type" in df_trades.columns else pd.Series(dtype="object")
    portfolio_buy_rows = int(portfolio_trade_types.str.startswith("買進").sum()) if len(portfolio_trade_types) > 0 else 0
    portfolio_full_exit_rows = int(portfolio_trade_types.isin({"全倉結算(停損)", "全倉結算(指標)", "期末強制結算", "汰弱賣出(Open, T+1再評估買進)"}).sum()) if len(portfolio_trade_types) > 0 else 0
    portfolio_half_take_profit_rows = int((portfolio_trade_types == "半倉停利").sum()) if len(portfolio_trade_types) > 0 else 0
    portfolio_missed_buy_rows = int(portfolio_trade_types.str.startswith("錯失買進").sum()) if len(portfolio_trade_types) > 0 else 0
    portfolio_missed_sell_rows = int((portfolio_trade_types == "錯失賣出").sum()) if len(portfolio_trade_types) > 0 else 0
    portfolio_period_closeout_rows = int((portfolio_trade_types == "期末強制結算").sum()) if len(portfolio_trade_types) > 0 else 0
    portfolio_completed_trades = rebuild_completed_trades_from_portfolio_trade_log(df_trades)
    return {
        "portfolio_buy_rows": portfolio_buy_rows,
        "portfolio_full_exit_rows": portfolio_full_exit_rows,
        "portfolio_half_take_profit_rows": portfolio_half_take_profit_rows,
        "portfolio_missed_buy_rows": portfolio_missed_buy_rows,
        "portfolio_missed_sell_rows": portfolio_missed_sell_rows,
        "portfolio_period_closeout_rows": portfolio_period_closeout_rows,
        "portfolio_completed_trades": portfolio_completed_trades,
    }


def build_portfolio_stats_payload(*, module_path, df_trades, total_return, mdd, trade_count, win_rate, pf_ev, pf_payoff, final_eq, avg_exp, max_exp, bm_ret, bm_mdd, total_missed, total_missed_sells, r_sq, m_win_rate, bm_r_sq, bm_m_win_rate, normal_trade_count, extended_trade_count, annual_trades, reserved_buy_fill_rate, annual_return_pct, bm_annual_return_pct, profile_stats):
    payload = {
        "module_path": module_path,
        "total_return": total_return,
        "mdd": mdd,
        "trade_count": trade_count,
        "win_rate": win_rate,
        "pf_ev": pf_ev,
        "pf_payoff": pf_payoff,
        "final_eq": final_eq,
        "avg_exp": avg_exp,
        "max_exp": max_exp,
        "bm_ret": bm_ret,
        "bm_mdd": bm_mdd,
        "total_missed": total_missed,
        "total_missed_sells": total_missed_sells,
        "r_sq": r_sq,
        "m_win_rate": m_win_rate,
        "bm_r_sq": bm_r_sq,
        "bm_m_win_rate": bm_m_win_rate,
        "normal_trade_count": normal_trade_count,
        "extended_trade_count": extended_trade_count,
        "annual_trades": annual_trades,
        "reserved_buy_fill_rate": reserved_buy_fill_rate,
        "annual_return_pct": annual_return_pct,
        "bm_annual_return_pct": bm_annual_return_pct,
        "df_trades": df_trades.copy() if isinstance(df_trades, pd.DataFrame) else pd.DataFrame(),
    }
    payload.update(summarize_portfolio_trade_output(payload["df_trades"]))
    payload.update(extract_yearly_profile_fields(profile_stats or {}))
    return payload


def make_synthetic_validation_params(base_params, *, tp_percent=None):
    params = make_consistency_params(base_params)
    params.high_len = 3
    params.atr_len = 2
    params.atr_buy_tol = 0.1
    params.atr_times_init = 1.0
    params.atr_times_trail = 1.5
    params.use_bb = False
    params.use_vol = False
    params.use_kc = False
    params.min_history_trades = 0
    params.min_history_ev = -1e9
    params.min_history_win_rate = 0.0
    if tp_percent is not None:
        params.tp_percent = tp_percent
    return params


def build_synthetic_baseline_frame(start_date, periods, base_price=100.0, volume=1000):
    dates = pd.bdate_range(start_date, periods=periods)
    df = pd.DataFrame({
        "Date": dates.strftime("%Y-%m-%d"),
        "Open": [base_price - 0.5] * periods,
        "High": [base_price + 0.5] * periods,
        "Low": [base_price - 1.0] * periods,
        "Close": [base_price] * periods,
        "Volume": [volume] * periods,
    })
    return df


def set_synthetic_bar(df, idx, *, open_price, high_price, low_price, close_price, volume=1000):
    df.loc[idx, ["Open", "High", "Low", "Close", "Volume"]] = [
        float(open_price),
        float(high_price),
        float(low_price),
        float(close_price),
        float(volume),
    ]


def write_synthetic_csv_bundle(temp_dir, frames_by_ticker):
    for ticker, frame in frames_by_ticker.items():
        frame.to_csv(os.path.join(temp_dir, f"{ticker}.csv"), index=False)


# # (AI註: scanner 必須用 production 門檻驗證；
# # (AI註: 若套用 consistency 的寬鬆門檻，會驗到實際 scanner 不會出現的候選)
def build_scanner_validation_params(base_params):
    return copy.deepcopy(base_params)


# # (AI註: 將 scanner 的 tuple 輸出正規化成 dict，
# # (AI註: 避免 validate 誤把 tuple 當 dict，導致 vip_scanner 完全沒被驗到)
def normalize_scanner_result(raw_result):
    if raw_result is None:
        return None

    if isinstance(raw_result, dict):
        return raw_result

    if not isinstance(raw_result, tuple) or len(raw_result) != 7:
        raise TypeError(
            f"scanner 回傳格式異常: type={type(raw_result).__name__}, value={raw_result}"
        )

    status, proj_cost, ev, sort_value, msg, ticker, sanitize_issue = raw_result
    return {
        "status": status,
        "proj_cost": proj_cost,
        "expected_value": ev,
        "sort_value": sort_value,
        "message": msg,
        "ticker": ticker,
        "sanitize_issue": sanitize_issue,
    }


# # (AI註: 用 strict scanner 參數重跑一次核心回測，作為 scanner 工具的真實參考口徑)
def run_scanner_reference_check(ticker, file_path, params):
    try:
        raw_df = pd.read_csv(file_path)
        min_rows_needed = get_required_min_rows(params)
        df, _sanitize_stats = sanitize_ohlcv_dataframe(raw_df, ticker, min_rows=min_rows_needed)
        return run_v16_backtest(df.copy(), params)
    except ValueError as e:
        if is_insufficient_data_error(e):
            return {"scanner_expected_status": "skip_insufficient"}
        raise


def derive_expected_scanner_status(scanner_ref_stats, params):
    if scanner_ref_stats.get("scanner_expected_status") == "skip_insufficient":
        return "skip_insufficient"

    if not scanner_ref_stats or not scanner_ref_stats["is_candidate"]:
        return None

    if scanner_ref_stats["is_setup_today"]:
        proj_qty = calc_reference_candidate_qty(
            scanner_ref_stats["buy_limit"],
            scanner_ref_stats["stop_loss"],
            params
        )
        return "buy" if proj_qty > 0 else "candidate"

    extended_candidate_today = scanner_ref_stats.get("extended_candidate_today")
    if extended_candidate_today is not None:
        limit_price = extended_candidate_today.get("limit_price")
        init_sl = extended_candidate_today.get("init_sl")
        if limit_price is None or init_sl is None:
            return "candidate"
        proj_qty = calc_reference_candidate_qty(limit_price, init_sl, params)
        return "extended" if proj_qty > 0 else "candidate"

    return "candidate"

def build_expected_scanner_payload(scanner_ref_stats, params):
    status = derive_expected_scanner_status(scanner_ref_stats, params)
    payload = {
        "status": status,
        "expected_value": None,
        "proj_cost": None,
        "sort_value": None,
    }

    if status not in ("buy", "extended"):
        return payload

    if status == "buy":
        limit_price = scanner_ref_stats["buy_limit"]
        stop_loss = scanner_ref_stats["stop_loss"]
        proj_qty = calc_reference_candidate_qty(limit_price, stop_loss, params)
    else:
        extended_candidate = scanner_ref_stats.get("extended_candidate_today")
        if extended_candidate is None:
            return payload
        limit_price = extended_candidate["limit_price"]
        init_sl = extended_candidate.get("init_sl")
        if init_sl is None:
            return payload
        proj_qty = calc_reference_candidate_qty(limit_price, init_sl, params)

    proj_cost = calc_entry_price(limit_price, proj_qty, params) * proj_qty
    sort_value = calc_buy_sort_value(
        BUY_SORT_METHOD,
        scanner_ref_stats["expected_value"],
        proj_cost,
        scanner_ref_stats["win_rate"] / 100.0,
        scanner_ref_stats["trade_count"],
    )
    payload.update({
        "expected_value": scanner_ref_stats["expected_value"],
        "proj_cost": proj_cost,
        "sort_value": sort_value,
    })
    return payload


def suppress_tool_output(func, *args, **kwargs):
    stdout_buffer = io.StringIO()
    stderr_buffer = io.StringIO()
    with redirect_stdout(stdout_buffer), redirect_stderr(stderr_buffer):
        return func(*args, **kwargs)


def resolve_source_date_column(source_df, ticker):
    if "Time" in source_df.columns:
        return "Time"
    if "Date" in source_df.columns:
        return "Date"
    raise KeyError(f"{ticker}: 找不到 Date/Time 欄位")


def run_portfolio_sim_tool_check_for_dir(data_dir, params, *, max_positions, enable_rotation, start_year, benchmark_ticker):
    module, module_path = load_module_from_candidates(
        "portfolio_sim_module",
        ["v16_portfolio_sim.py"],
        required_attrs=["run_portfolio_simulation"]
    )

    result = suppress_tool_output(
        module.run_portfolio_simulation,
        data_dir=data_dir,
        params=copy.deepcopy(params),
        max_positions=max_positions,
        enable_rotation=enable_rotation,
        start_year=start_year,
        benchmark_ticker=benchmark_ticker,
        verbose=False,
    )

    (
        _df_eq,
        _df_tr,
        tot_ret,
        mdd,
        trade_count,
        win_rate,
        pf_ev,
        pf_payoff,
        final_eq,
        avg_exp,
        max_exp,
        bm_ret,
        bm_mdd,
        total_missed,
        total_missed_sells,
        r_sq,
        m_win_rate,
        bm_r_sq,
        bm_m_win_rate,
        normal_trade_count,
        extended_trade_count,
        annual_trades,
        reserved_buy_fill_rate,
        annual_return_pct,
        bm_annual_return_pct,
        pf_profile,
    ) = result

    return build_portfolio_stats_payload(
        module_path=module_path,
        df_trades=_df_tr,
        total_return=tot_ret,
        mdd=mdd,
        trade_count=trade_count,
        win_rate=win_rate,
        pf_ev=pf_ev,
        pf_payoff=pf_payoff,
        final_eq=final_eq,
        avg_exp=avg_exp,
        max_exp=max_exp,
        bm_ret=bm_ret,
        bm_mdd=bm_mdd,
        total_missed=total_missed,
        total_missed_sells=total_missed_sells,
        r_sq=r_sq,
        m_win_rate=m_win_rate,
        bm_r_sq=bm_r_sq,
        bm_m_win_rate=bm_m_win_rate,
        normal_trade_count=normal_trade_count,
        extended_trade_count=extended_trade_count,
        annual_trades=annual_trades,
        reserved_buy_fill_rate=reserved_buy_fill_rate,
        annual_return_pct=annual_return_pct,
        bm_annual_return_pct=bm_annual_return_pct,
        profile_stats=pf_profile,
    )


# # (AI註: portfolio_sim 工具檢查改委派到通用 data_dir 版本，
# # (AI註: 後續 synthetic 多檔案例也能共用，避免 validate 自己重做一套解析邏輯)
def run_portfolio_sim_tool_check(ticker, file_path, params):
    with tempfile.TemporaryDirectory() as temp_dir:
        source_df = pd.read_csv(file_path)
        temp_csv_path = os.path.join(temp_dir, os.path.basename(file_path))
        source_df.to_csv(temp_csv_path, index=False)

        date_col = resolve_source_date_column(source_df, ticker)
        parsed_dates = pd.to_datetime(source_df[date_col], errors="coerce")
        min_year = parsed_dates.dt.year.min()
        if pd.isna(min_year):
            raise ValueError(f"{ticker}: 日期欄位 {date_col} 無任何可解析日期")

        return run_portfolio_sim_tool_check_for_dir(
            temp_dir,
            params,
            max_positions=1,
            enable_rotation=False,
            start_year=int(min_year),
            benchmark_ticker=ticker,
        )

def run_scanner_tool_check(ticker, file_path, params):
    module, module_path = load_module_from_candidates(
        "vip_scanner_module",
        ["v16_vip_scanner.py"],
        required_attrs=["process_single_stock"]
    )

    raw_result = suppress_tool_output(
        module.process_single_stock,
        file_path=file_path,
        ticker=ticker,
        params=copy.deepcopy(params)
    )
    return normalize_scanner_result(raw_result), module_path

def run_downloader_tool_check(ticker):
    module, module_path = load_module_from_candidates(
        "vip_downloader_module",
        ["vip_smart_downloader.py"],
        required_attrs=["smart_download_vip_data"]
    )

    class DummyDL:
        def __init__(self):
            self.calls = []

        def get_data(self, dataset, data_id, start_date):
            self.calls.append({
                "dataset": dataset,
                "data_id": data_id,
                "start_date": start_date,
            })
            if data_id != ticker:
                raise ValueError(f"unexpected ticker: {data_id}")
            return pd.DataFrame({
                'date': ['2024-01-03', '2024-01-02'],
                'open': [11.0, 10.0],
                'max': [12.0, 11.0],
                'min': [10.5, 9.5],
                'close': [11.5, 10.5],
                'trading_volume': [2000, 1000],
            })

    with tempfile.TemporaryDirectory() as temp_dir:
        original_save_dir = module.SAVE_DIR
        original_dl = module.dl
        original_sleep = module.time.sleep
        dummy_loader = DummyDL()
        try:
            module.SAVE_DIR = temp_dir
            module.dl = dummy_loader
            module.time.sleep = lambda *_args, **_kwargs: None
            suppress_tool_output(
                module.smart_download_vip_data,
                [ticker],
                market_last_date='2024-01-03',
                verbose=False
            )
            csv_path = os.path.join(temp_dir, f"{ticker}.csv")
            downloaded_df = pd.read_csv(csv_path, index_col=0)
        finally:
            module.SAVE_DIR = original_save_dir
            module.dl = original_dl
            module.time.sleep = original_sleep

    download_request = dummy_loader.calls[0] if dummy_loader.calls else None
    return downloaded_df, module_path, download_request, module.FINMIND_PRICE_DATASET

# # (AI註: validate 補回工具級檢查，避免只驗核心邏輯而漏掉 portfolio_sim / scanner / downloader)
MODULE_CACHE = {}

def load_module_from_candidates(cache_key, candidate_files, required_attrs):
    if cache_key in MODULE_CACHE:
        return MODULE_CACHE[cache_key]

    checked_paths = []
    rejected_paths = []

    for file_name in candidate_files:
        module_path = os.path.join(PROJECT_ROOT, file_name)
        checked_paths.append(module_path)

        if not os.path.exists(module_path):
            continue

        spec = importlib.util.spec_from_file_location(cache_key, module_path)
        if spec is None or spec.loader is None:
            rejected_paths.append(f"{module_path} -> 無法建立 spec/loader")
            continue

        module = importlib.util.module_from_spec(spec)
        try:
            spec.loader.exec_module(module)
        except MODULE_LOAD_RECOVERABLE_EXCEPTIONS as e:
            rejected_paths.append(f"{module_path} -> 載入失敗: {type(e).__name__}: {e}")
            continue

        missing_attrs = [attr for attr in required_attrs if not hasattr(module, attr)]
        if missing_attrs:
            rejected_paths.append(f"{module_path} -> 缺少必要屬性: {missing_attrs}")
            continue

        MODULE_CACHE[cache_key] = (module, module_path)
        return module, module_path

    detail_msg = "；".join(rejected_paths) if rejected_paths else "沒有任何可用候選檔"
    raise FileNotFoundError(
        f"找不到符合條件的模組。檢查路徑: {checked_paths}。原因: {detail_msg}"
    )

def run_single_backtest_check(ticker, df, params):
    stats, standalone_logs = run_v16_backtest(df.copy(), params, return_logs=True)
    return stats, standalone_logs

def run_single_ticker_portfolio_check(ticker, df, params):
    execution_params = build_execution_only_params(params)
    prep_df, standalone_logs = prep_stock_data_and_trades(df.copy(), execution_params)

    fast_data = pack_prepared_stock_data(prep_df)
    all_dfs_fast = {ticker: fast_data}
    all_standalone_logs = {ticker: standalone_logs}

    sorted_dates = sorted(get_fast_dates(fast_data))
    if not sorted_dates:
        raise ValueError(f"{ticker}: pack_prepared_stock_data 後沒有任何有效日期")

    start_year = int(pd.Timestamp(sorted_dates[0]).year)
    profile_stats = {}

    result = run_portfolio_timeline(
        all_dfs_fast=all_dfs_fast,
        all_standalone_logs=all_standalone_logs,
        sorted_dates=sorted_dates,
        start_year=start_year,
        params=execution_params,
        max_positions=1,
        enable_rotation=False,
        benchmark_ticker=ticker,
        benchmark_data=fast_data,
        is_training=True,
        profile_stats=profile_stats,
        verbose=False
    )

    expected_result_len = 23
    if len(result) != expected_result_len:
        raise ValueError(
            f"run_portfolio_timeline(is_training=True) 回傳長度異常: {len(result)}，"
            f"預期 {expected_result_len}"
        )

    (
        total_return,
        mdd,
        trade_count,
        final_eq,
        avg_exp,
        max_exp,
        bm_ret,
        bm_mdd,
        win_rate,
        pf_ev,
        pf_payoff,
        total_missed,
        total_missed_sells,
        r_sq,
        m_win_rate,
        bm_r_sq,
        bm_m_win_rate,
        normal_trade_count,
        extended_trade_count,
        annual_trades,
        reserved_buy_fill_rate,
        annual_return_pct,
        bm_annual_return_pct,
    ) = result

    payload = {
        "total_return": total_return,
        "mdd": mdd,
        "trade_count": trade_count,
        "final_eq": final_eq,
        "avg_exp": avg_exp,
        "max_exp": max_exp,
        "bm_ret": bm_ret,
        "bm_mdd": bm_mdd,
        "win_rate": win_rate,
        "pf_ev": pf_ev,
        "pf_payoff": pf_payoff,
        "total_missed": total_missed,
        "total_missed_sells": total_missed_sells,
        "r_sq": r_sq,
        "m_win_rate": m_win_rate,
        "bm_r_sq": bm_r_sq,
        "bm_m_win_rate": bm_m_win_rate,
        "normal_trade_count": normal_trade_count,
        "extended_trade_count": extended_trade_count,
        "annual_trades": annual_trades,
        "reserved_buy_fill_rate": reserved_buy_fill_rate,
        "annual_return_pct": annual_return_pct,
        "bm_annual_return_pct": bm_annual_return_pct,
        "sorted_dates": sorted_dates,
        "start_year": start_year,
        "standalone_logs": standalone_logs,
        "prep_df": prep_df,
    }
    payload.update(extract_yearly_profile_fields(profile_stats))
    return payload


def run_debug_trade_log_check(ticker, df, params):
    module, module_path = load_module_from_candidates(
        "debug_trade_log_module",
        ["tools/debug_trade_log.py", "debug_trade_log.py"],
        required_attrs=["run_debug_backtest"]
    )
    debug_df = module.run_debug_backtest(
        df.copy(),
        ticker,
        params,
        export_excel=False,
        verbose=False
    )
    return debug_df, module_path


def rebuild_completed_trades_from_event_log(df_logs, *, tool_name, date_col, action_col, pnl_col, buy_prefix, half_action, full_exit_actions, ignored_actions, missed_sell_actions):
    if df_logs is None or len(df_logs) == 0:
        return []

    required_cols = {date_col, action_col, pnl_col}
    missing_cols = [col for col in required_cols if col not in df_logs.columns]
    if missing_cols:
        raise KeyError(f"{tool_name} 缺少必要欄位: {missing_cols}")

    completed_trades = []
    active_trade = None

    for row in df_logs.itertuples(index=False):
        action = str(getattr(row, action_col))
        trade_date = pd.to_datetime(getattr(row, date_col)).strftime("%Y-%m-%d")
        realized_pnl = float(getattr(row, pnl_col))

        if action.startswith(buy_prefix):
            if active_trade is not None:
                raise ValueError(f"{tool_name} 出現連續買進，上一筆交易尚未完整結束。")
            active_trade = {
                "buy_date": trade_date,
                "exit_date": None,
                "total_pnl": 0.0,
                "half_exit_count": 0,
                "full_exit_action": None,
            }
            continue

        if action in ignored_actions:
            continue

        if action == half_action:
            if active_trade is None:
                raise ValueError(f"{tool_name} 出現{half_action}，但前面沒有對應買進。")
            active_trade["total_pnl"] += realized_pnl
            active_trade["half_exit_count"] += 1
            continue

        if action in missed_sell_actions:
            if active_trade is None:
                raise ValueError(f"{tool_name} 出現 {action}，但前面沒有對應買進。")
            continue

        if action in full_exit_actions:
            if active_trade is None:
                raise ValueError(f"{tool_name} 出現 {action}，但前面沒有對應買進。")
            active_trade["total_pnl"] = round(active_trade["total_pnl"] + realized_pnl, 2)
            active_trade["exit_date"] = trade_date
            active_trade["full_exit_action"] = action
            completed_trades.append(active_trade)
            active_trade = None
            continue

        raise ValueError(f"{tool_name} 出現未納入驗證的動作: {action}")

    if active_trade is not None:
        raise ValueError(f"{tool_name} 最後仍有未完成交易，缺少完整賣出列。")

    return completed_trades


# # (AI註: 將 debug_trade_log 的逐列事件重建為 completed trades，
# # (AI註: 才能嚴格驗證半倉停利 + 尾倉結算後的總損益口徑是否與核心一致)
def rebuild_completed_trades_from_debug_log(debug_df):
    return rebuild_completed_trades_from_event_log(
        debug_df,
        tool_name="debug_trade_log",
        date_col="日期",
        action_col="動作",
        pnl_col="單筆實質損益",
        buy_prefix="買進",
        half_action="半倉停利",
        full_exit_actions={"停損殺出", "指標賣出", "期末強制結算"},
        ignored_actions={"錯失買進(新訊號)", "錯失買進(延續候選)", "放棄進場(先達停損)", "放棄進場(延續先達停損)"},
        missed_sell_actions={"錯失賣出"},
    )


# # (AI註: portfolio_sim 的 df_trades 也必須能重建成 completed trades，
# # (AI註: 否則即使 aggregate 指標一致，逐筆明細仍可能漂移而不自知)
def rebuild_completed_trades_from_portfolio_trade_log(df_trades):
    return rebuild_completed_trades_from_event_log(
        df_trades,
        tool_name="portfolio_sim df_trades",
        date_col="Date",
        action_col="Type",
        pnl_col="單筆損益",
        buy_prefix="買進",
        half_action="半倉停利",
        full_exit_actions={"全倉結算(停損)", "全倉結算(指標)", "期末強制結算", "汰弱賣出(Open, T+1再評估買進)"},
        ignored_actions={"錯失買進(新訊號)", "錯失買進(延續候選)"},
        missed_sell_actions={"錯失賣出"},
    )


def validate_one_ticker(ticker, base_params):
    params = make_consistency_params(base_params)
    scanner_params = build_scanner_validation_params(base_params)
    file_path, df, sanitize_stats = load_clean_df(ticker, params)

    results = []
    summary = {
        "ticker": ticker,
        "file_path": file_path,
        "sanitize_dropped": sanitize_stats["dropped_row_count"],
        "sanitize_invalid": sanitize_stats["invalid_row_count"],
        "sanitize_duplicate": sanitize_stats["duplicate_date_count"],
    }

    single_stats, standalone_logs = run_single_backtest_check(ticker, df, params)
    scanner_ref_stats = run_scanner_reference_check(ticker, file_path, scanner_params)
    portfolio_stats = run_single_ticker_portfolio_check(ticker, df, params)
    portfolio_sim_stats = run_portfolio_sim_tool_check(ticker, file_path, params)
    scanner_result, scanner_module_path = run_scanner_tool_check(ticker, file_path, scanner_params)
    downloader_df, downloader_module_path, downloader_request, downloader_expected_dataset = None, None, None, None
    downloader_error = None
    try:
        downloader_df, downloader_module_path, downloader_request, downloader_expected_dataset = run_downloader_tool_check(ticker)
    except VALIDATION_RECOVERABLE_EXCEPTIONS as e:
        downloader_error = f"{type(e).__name__}: {e}"
    debug_df, debug_module_path = run_debug_trade_log_check(ticker, df, params)

    summary["portfolio_sim_module_path"] = portfolio_sim_stats["module_path"]
    summary["scanner_module_path"] = scanner_module_path
    summary["downloader_module_path"] = downloader_module_path
    summary["downloader_error"] = downloader_error
    summary["debug_module_path"] = debug_module_path
    summary["single_trade_count"] = single_stats["trade_count"]
    summary["portfolio_trade_count"] = portfolio_stats["trade_count"]
    summary["open_position_exists"] = bool(single_stats["current_position"] > 0)
    summary["has_extended_candidate_today"] = bool(scanner_ref_stats.get("extended_candidate_today") is not None)
    summary["has_missed_buy"] = bool(single_stats["missed_buys"] > 0)
    summary["portfolio_half_take_profit_rows"] = int(portfolio_sim_stats["portfolio_half_take_profit_rows"])

    add_check(results, "single_vs_portfolio", ticker, "asset_growth_vs_total_return",
              single_stats["asset_growth"], portfolio_stats["total_return"])

    add_check(results, "single_vs_portfolio", ticker, "max_drawdown_vs_mdd",
              single_stats["max_drawdown"], portfolio_stats["mdd"])

    add_check(results, "single_vs_portfolio", ticker, "missed_buys",
              single_stats["missed_buys"], portfolio_stats["total_missed"])

    add_check(results, "single_vs_portfolio", ticker, "missed_sells",
              single_stats["missed_sells"], portfolio_stats["total_missed_sells"])

    add_check(
        results,
        "single_vs_portfolio",
        ticker,
        "trade_count",
        single_stats["trade_count"],
        portfolio_stats["trade_count"],
        note="單股與投組都已將期末強制結算納入交易統計。"
    )

    add_check(
        results,
        "single_vs_portfolio",
        ticker,
        "normal_plus_extended_trade_count",
        portfolio_stats["trade_count"],
        portfolio_stats["normal_trade_count"] + portfolio_stats["extended_trade_count"],
        note="正常/延續完整交易數總和應等於總交易次數。"
    )

    sim_years = calc_validation_sim_years(portfolio_stats["sorted_dates"], portfolio_stats["start_year"])
    expected_annual_trades = (portfolio_stats["trade_count"] / sim_years) if sim_years > 0 else 0.0
    total_reserved_entries = portfolio_stats["normal_trade_count"] + portfolio_stats["extended_trade_count"]
    expected_reserved_buy_fill_rate = (
        total_reserved_entries / (total_reserved_entries + portfolio_stats["total_missed"]) * 100.0
        if (total_reserved_entries + portfolio_stats["total_missed"]) > 0 else 0.0
    )
    expected_annual_return_pct = calc_validation_annual_return_pct(
        params.initial_capital, portfolio_stats["final_eq"], sim_years
    )
    expected_bm_annual_return_pct = calc_validation_annual_return_pct(
        100.0, 100.0 * (1.0 + portfolio_stats["bm_ret"] / 100.0), sim_years
    )
    expected_exit_dates = [pd.to_datetime(log["exit_date"]).strftime("%Y-%m-%d") for log in standalone_logs]
    expected_trade_pnls = [round(float(log["pnl"]), 2) for log in standalone_logs]
    expected_realized_pnl_sum = round(sum(expected_trade_pnls), 2)
    expected_full_year_metrics = calc_expected_full_year_metrics(portfolio_stats["yearly_return_rows"])
    expected_bm_full_year_metrics = calc_expected_full_year_metrics(portfolio_stats["bm_yearly_return_rows"])

    add_check(results, "single_vs_portfolio", ticker, "annual_trades", expected_annual_trades, portfolio_stats["annual_trades"])
    add_check(results, "single_vs_portfolio", ticker, "reserved_buy_fill_rate", expected_reserved_buy_fill_rate, portfolio_stats["reserved_buy_fill_rate"])
    add_check(results, "single_vs_portfolio", ticker, "annual_return_pct", expected_annual_return_pct, portfolio_stats["annual_return_pct"])
    add_check(results, "single_vs_portfolio", ticker, "bm_annual_return_pct", expected_bm_annual_return_pct, portfolio_stats["bm_annual_return_pct"])
    add_check(results, "single_vs_portfolio", ticker, "full_year_count", expected_full_year_metrics["full_year_count"], portfolio_stats["full_year_count"])
    add_check(results, "single_vs_portfolio", ticker, "min_full_year_return_pct", expected_full_year_metrics["min_full_year_return_pct"], portfolio_stats["min_full_year_return_pct"])
    add_check(results, "single_vs_portfolio", ticker, "yearly_return_rows", portfolio_stats["yearly_return_rows"], portfolio_stats["yearly_return_rows"], note="年度報酬明細需保留完整列內容，供後續與 portfolio_sim 對照。")
    add_check(results, "single_vs_portfolio", ticker, "bm_full_year_count", expected_bm_full_year_metrics["full_year_count"], portfolio_stats["bm_full_year_count"])
    add_check(results, "single_vs_portfolio", ticker, "bm_min_full_year_return_pct", expected_bm_full_year_metrics["min_full_year_return_pct"], portfolio_stats["bm_min_full_year_return_pct"])
    add_check(results, "single_vs_portfolio", ticker, "bm_yearly_return_rows", portfolio_stats["bm_yearly_return_rows"], portfolio_stats["bm_yearly_return_rows"], note="Benchmark 年度報酬明細需保留完整列內容，供後續與 portfolio_sim 對照。")

    add_check(results, "single_vs_portfolio", ticker, "win_rate",
              single_stats["win_rate"], portfolio_stats["win_rate"])
    add_check(results, "single_vs_portfolio", ticker, "payoff_ratio",
              single_stats["payoff_ratio"], portfolio_stats["pf_payoff"])
    add_check(results, "single_vs_portfolio", ticker, "expected_value",
              single_stats["expected_value"], portfolio_stats["pf_ev"])

    add_check(results, "portfolio_sim", ticker, "total_return", portfolio_stats["total_return"], portfolio_sim_stats["total_return"])
    add_check(results, "portfolio_sim", ticker, "mdd", portfolio_stats["mdd"], portfolio_sim_stats["mdd"])
    add_check(results, "portfolio_sim", ticker, "trade_count", portfolio_stats["trade_count"], portfolio_sim_stats["trade_count"])
    add_check(results, "portfolio_sim", ticker, "win_rate", portfolio_stats["win_rate"], portfolio_sim_stats["win_rate"])
    add_check(results, "portfolio_sim", ticker, "pf_ev", portfolio_stats["pf_ev"], portfolio_sim_stats["pf_ev"])
    add_check(results, "portfolio_sim", ticker, "pf_payoff", portfolio_stats["pf_payoff"], portfolio_sim_stats["pf_payoff"])
    add_check(results, "portfolio_sim", ticker, "final_eq", portfolio_stats["final_eq"], portfolio_sim_stats["final_eq"])
    add_check(results, "portfolio_sim", ticker, "avg_exp", portfolio_stats["avg_exp"], portfolio_sim_stats["avg_exp"])
    add_check(results, "portfolio_sim", ticker, "max_exp", portfolio_stats["max_exp"], portfolio_sim_stats["max_exp"])
    add_check(results, "portfolio_sim", ticker, "bm_ret", portfolio_stats["bm_ret"], portfolio_sim_stats["bm_ret"])
    add_check(results, "portfolio_sim", ticker, "bm_mdd", portfolio_stats["bm_mdd"], portfolio_sim_stats["bm_mdd"])
    add_check(results, "portfolio_sim", ticker, "total_missed", portfolio_stats["total_missed"], portfolio_sim_stats["total_missed"])
    add_check(results, "portfolio_sim", ticker, "total_missed_sells", portfolio_stats["total_missed_sells"], portfolio_sim_stats["total_missed_sells"])
    add_check(
        results,
        "portfolio_sim",
        ticker,
        "df_trades_missed_buy_rows",
        portfolio_stats["total_missed"],
        portfolio_sim_stats["portfolio_missed_buy_rows"],
        note="portfolio df_trades 中的錯失買進列數，必須與 total_missed 完全一致。"
    )
    add_check(
        results,
        "portfolio_sim",
        ticker,
        "df_trades_missed_sell_rows",
        portfolio_stats["total_missed_sells"],
        portfolio_sim_stats["portfolio_missed_sell_rows"],
        note="portfolio df_trades 中的錯失賣出列數，必須與 total_missed_sells 完全一致。"
    )
    add_check(
        results,
        "portfolio_sim",
        ticker,
        "df_trades_buy_rows",
        len(standalone_logs),
        portfolio_sim_stats["portfolio_buy_rows"],
        note="portfolio df_trades 中的買進列數，必須與核心 completed trades 筆數一致。"
    )
    add_check(
        results,
        "portfolio_sim",
        ticker,
        "df_trades_full_exit_rows",
        len(standalone_logs),
        portfolio_sim_stats["portfolio_full_exit_rows"],
        note="portfolio df_trades 中的完整賣出列數，必須與核心 completed trades 筆數一致，包含期末強制結算。"
    )
    add_check(
        results,
        "portfolio_sim",
        ticker,
        "df_trades_period_closeout_rows",
        1 if single_stats["current_position"] > 0 else 0,
        portfolio_sim_stats["portfolio_period_closeout_rows"],
        note="若單股回測期末仍持有部位，portfolio df_trades 必須恰有一列期末強制結算；否則必須為 0。"
    )
    add_check(
        results,
        "portfolio_sim",
        ticker,
        "df_trades_completed_trade_count",
        len(standalone_logs),
        len(portfolio_sim_stats["portfolio_completed_trades"]),
        note="portfolio df_trades 必須能重建成與核心 completed trades 完全相同的筆數。"
    )
    add_check(
        results,
        "portfolio_sim",
        ticker,
        "df_trades_completed_trade_exit_dates",
        expected_exit_dates,
        [trade["exit_date"] for trade in portfolio_sim_stats["portfolio_completed_trades"]],
        note="portfolio df_trades 重建出的逐筆最終出場日期，必須與核心 completed trades 完全一致。"
    )
    add_check(
        results,
        "portfolio_sim",
        ticker,
        "df_trades_completed_trade_pnl_sequence",
        expected_trade_pnls,
        [trade["total_pnl"] for trade in portfolio_sim_stats["portfolio_completed_trades"]],
        note="portfolio df_trades 必須將半倉停利 + 尾倉賣出正確合併，逐筆總損益 sequence 與核心一致。"
    )
    add_check(
        results,
        "portfolio_sim",
        ticker,
        "df_trades_completed_trade_realized_pnl_sum",
        expected_realized_pnl_sum,
        round(sum(trade["total_pnl"] for trade in portfolio_sim_stats["portfolio_completed_trades"]), 2),
        tol=0.01,
        note="portfolio df_trades 重建後的 completed trades 總已實現損益，必須與核心一致。"
    )
    add_check(results, "portfolio_sim", ticker, "r_sq", portfolio_stats["r_sq"], portfolio_sim_stats["r_sq"])
    add_check(results, "portfolio_sim", ticker, "m_win_rate", portfolio_stats["m_win_rate"], portfolio_sim_stats["m_win_rate"])
    add_check(results, "portfolio_sim", ticker, "bm_r_sq", portfolio_stats["bm_r_sq"], portfolio_sim_stats["bm_r_sq"])
    add_check(results, "portfolio_sim", ticker, "bm_m_win_rate", portfolio_stats["bm_m_win_rate"], portfolio_sim_stats["bm_m_win_rate"])
    add_check(results, "portfolio_sim", ticker, "normal_trade_count", portfolio_stats["normal_trade_count"], portfolio_sim_stats["normal_trade_count"])
    add_check(results, "portfolio_sim", ticker, "extended_trade_count", portfolio_stats["extended_trade_count"], portfolio_sim_stats["extended_trade_count"])
    add_check(results, "portfolio_sim", ticker, "annual_trades", portfolio_stats["annual_trades"], portfolio_sim_stats["annual_trades"])
    add_check(results, "portfolio_sim", ticker, "reserved_buy_fill_rate", portfolio_stats["reserved_buy_fill_rate"], portfolio_sim_stats["reserved_buy_fill_rate"])
    add_check(results, "portfolio_sim", ticker, "annual_return_pct", portfolio_stats["annual_return_pct"], portfolio_sim_stats["annual_return_pct"])
    add_check(results, "portfolio_sim", ticker, "bm_annual_return_pct", portfolio_stats["bm_annual_return_pct"], portfolio_sim_stats["bm_annual_return_pct"])
    add_check(results, "portfolio_sim", ticker, "full_year_count", portfolio_stats["full_year_count"], portfolio_sim_stats["full_year_count"])
    add_check(results, "portfolio_sim", ticker, "min_full_year_return_pct", portfolio_stats["min_full_year_return_pct"], portfolio_sim_stats["min_full_year_return_pct"])
    add_check(results, "portfolio_sim", ticker, "yearly_return_rows", portfolio_stats["yearly_return_rows"], portfolio_sim_stats["yearly_return_rows"])
    add_check(results, "portfolio_sim", ticker, "bm_full_year_count", portfolio_stats["bm_full_year_count"], portfolio_sim_stats["bm_full_year_count"])
    add_check(results, "portfolio_sim", ticker, "bm_min_full_year_return_pct", portfolio_stats["bm_min_full_year_return_pct"], portfolio_sim_stats["bm_min_full_year_return_pct"])
    add_check(results, "portfolio_sim", ticker, "bm_yearly_return_rows", portfolio_stats["bm_yearly_return_rows"], portfolio_sim_stats["bm_yearly_return_rows"])

    expected_scanner_payload = build_expected_scanner_payload(scanner_ref_stats, scanner_params)
    expected_scanner_status = expected_scanner_payload["status"]

    if scanner_result is None:
        add_check(
            results,
            "vip_scanner",
            ticker,
            "status",
            expected_scanner_status,
            None,
            note="scanner 已實際執行；None 只在 strict production 門檻下無候選時才屬正確。"
        )
    else:
        add_check(
            results,
            "vip_scanner",
            ticker,
            "ticker",
            str(ticker),
            str(scanner_result["ticker"])
        )

        add_check(
            results,
            "vip_scanner",
            ticker,
            "status",
            expected_scanner_status,
            scanner_result["status"]
        )
        add_check(
            results,
            "vip_scanner",
            ticker,
            "expected_value",
            expected_scanner_payload["expected_value"],
            scanner_result["expected_value"]
        )
        add_check(
            results,
            "vip_scanner",
            ticker,
            "proj_cost",
            expected_scanner_payload["proj_cost"],
            scanner_result["proj_cost"]
        )
        add_check(
            results,
            "vip_scanner",
            ticker,
            "sort_value",
            expected_scanner_payload["sort_value"],
            scanner_result["sort_value"]
        )

    extended_candidate = scanner_ref_stats.get("extended_candidate_today")
    if extended_candidate is None:
        add_skip_result(results, "vip_scanner", ticker, "extended_reference_price_in_range", "今日無延續候選，無需驗 A 版區間定義。")
        add_skip_result(results, "vip_scanner", ticker, "extended_limit_price_in_range", "今日無延續候選，無需驗 A 版區間定義。")
    else:
        reference_price = float(df["Close"].iloc[-1])
        add_check(
            results,
            "vip_scanner",
            ticker,
            "extended_reference_price_in_range",
            True,
            bool(extended_candidate["init_sl"] < reference_price <= extended_candidate["orig_limit"]),
        )
        add_check(
            results,
            "vip_scanner",
            ticker,
            "extended_limit_price_in_range",
            True,
            bool(extended_candidate["init_sl"] < extended_candidate["limit_price"] <= extended_candidate["orig_limit"]),
        )

    if downloader_error is not None:
        add_fail_result(
            results,
            "vip_downloader",
            ticker,
            "tool_runtime",
            "tool loads and runs",
            downloader_error,
            note="downloader 工具失敗時，validate 應保留其他模組結果，不可整體中斷。"
        )
    else:
        expected_download_cols = ["Open", "High", "Low", "Close", "Volume"]
        actual_download_cols = list(downloader_df.columns)
        add_check(results, "vip_downloader", ticker, "columns", expected_download_cols, actual_download_cols)
        add_check(results, "vip_downloader", ticker, "row_count", 2, len(downloader_df))
        add_check(results, "vip_downloader", ticker, "dataset", downloader_expected_dataset, None if downloader_request is None else downloader_request["dataset"])
        add_check(results, "vip_downloader", ticker, "data_id", ticker, None if downloader_request is None else downloader_request["data_id"])
        add_check(results, "vip_downloader", ticker, "start_date", "1990-01-01", None if downloader_request is None else downloader_request["start_date"])
        expected_download_index = ["2024-01-02", "2024-01-03"]
        actual_download_index = [str(idx).split(" ")[0] for idx in downloader_df.index.tolist()]
        add_check(results, "vip_downloader", ticker, "date_index_sorted", expected_download_index, actual_download_index)
        add_check(results, "vip_downloader", ticker, "index_name", "Date", downloader_df.index.name)
        expected_download_rows = [
            {"Open": 10.0, "High": 11.0, "Low": 9.5, "Close": 10.5, "Volume": 1000},
            {"Open": 11.0, "High": 12.0, "Low": 10.5, "Close": 11.5, "Volume": 2000},
        ]
        actual_download_rows = downloader_df.reset_index(drop=True).to_dict("records")
        add_check(results, "vip_downloader", ticker, "ohlcv_values_after_sort", expected_download_rows, actual_download_rows)

    expected_buy_rows = len(standalone_logs)

    if debug_df is None or len(debug_df) == 0:
        if expected_buy_rows == 0:
            add_skip_result(
                results,
                "debug_trade_log",
                ticker,
                "debug_df_exists",
                "無交易紀錄時，debug 工具回傳 None 屬設計行為。"
            )
        else:
            add_fail_result(
                results,
                "debug_trade_log",
                ticker,
                "debug_df_exists",
                "非空",
                "None/Empty",
                "理應有交易明細，但 debug 工具回傳空值。"
            )
    else:
        action_series = debug_df["動作"].fillna("")
        buy_rows = int(action_series.str.startswith("買進").sum())
        exit_rows = int(action_series.isin(["停損殺出", "指標賣出", "期末強制結算"]).sum())
        half_rows = int((action_series == "半倉停利").sum())
        missed_buy_rows = int(action_series.str.startswith("錯失買進").sum())
        missed_sell_rows = int((action_series == "錯失賣出").sum())
        debug_completed_trades = rebuild_completed_trades_from_debug_log(debug_df)

        expected_exit_rows = len(standalone_logs)
        actual_trade_pnls = [trade["total_pnl"] for trade in debug_completed_trades]
        actual_exit_dates = [trade["exit_date"] for trade in debug_completed_trades]
        actual_realized_pnl_sum = round(sum(actual_trade_pnls), 2)

        add_check(
            results,
            "debug_trade_log",
            ticker,
            "buy_rows",
            expected_buy_rows,
            buy_rows,
            note="debug 已將期末強制結算列為完整賣出紀錄，買進筆數應等於 completed trades。"
        )

        add_check(
            results,
            "debug_trade_log",
            ticker,
            "full_exit_rows",
            expected_exit_rows,
            exit_rows
        )

        add_check(
            results,
            "debug_trade_log",
            ticker,
            "missed_buy_rows",
            int(single_stats["missed_buys"]),
            missed_buy_rows,
            note="debug 明細中的錯失買進筆數，必須與核心 missed_buys 完全一致。"
        )

        add_check(
            results,
            "debug_trade_log",
            ticker,
            "missed_sell_rows",
            int(single_stats["missed_sells"]),
            missed_sell_rows,
            note="debug 明細中的錯失賣出筆數，必須與核心 missed_sells 完全一致。"
        )

        add_check(
            results,
            "debug_trade_log",
            ticker,
            "completed_trade_count",
            len(standalone_logs),
            len(debug_completed_trades),
            note="debug 明細需能重建為與核心 completed trades 完全相同的筆數。"
        )

        add_check(
            results,
            "debug_trade_log",
            ticker,
            "completed_trade_exit_dates",
            expected_exit_dates,
            actual_exit_dates,
            note="每筆 completed trade 的最終出場日期必須與核心一致，包含期末強制結算。"
        )

        add_check(
            results,
            "debug_trade_log",
            ticker,
            "completed_trade_pnl_sequence",
            expected_trade_pnls,
            actual_trade_pnls,
            note="debug 需將半倉停利 + 尾倉賣出合併後，逐筆總損益與核心 completed trades 一致。"
        )

        add_check(
            results,
            "debug_trade_log",
            ticker,
            "completed_trade_realized_pnl_sum",
            expected_realized_pnl_sum,
            actual_realized_pnl_sum,
            tol=0.01,
            note="逐筆加總後的總已實現損益必須與核心 completed trades 一致。"
        )

        add_check(
            results,
            "debug_trade_log",
            ticker,
            "half_take_profit_rows",
            int(portfolio_sim_stats["portfolio_half_take_profit_rows"]),
            half_rows,
            note="debug 與 portfolio_sim 的半倉停利列數必須一致，避免半倉現金回收口徑漂移。"
        )

    return results, summary


def run_portfolio_core_check_for_dir(data_dir, params, *, max_positions, enable_rotation, start_year, benchmark_ticker):
    csv_inputs, _duplicate_issue_lines = discover_unique_csv_inputs(data_dir)
    if not csv_inputs:
        raise ValueError(f"synthetic data_dir 無任何 CSV: {data_dir}")

    all_dfs_fast = {}
    all_trade_logs = {}
    master_dates = set()

    for ticker, file_path in csv_inputs:
        raw_df = pd.read_csv(file_path)
        min_rows_needed = get_required_min_rows(params)
        df, _sanitize_stats = sanitize_ohlcv_dataframe(raw_df, ticker, min_rows=min_rows_needed)
        prep_df, standalone_logs = prep_stock_data_and_trades(df, params)
        fast_df = pack_prepared_stock_data(prep_df)
        all_dfs_fast[ticker] = fast_df
        all_trade_logs[ticker] = standalone_logs
        master_dates.update(prep_df.index)

    sorted_dates = sorted(master_dates)
    if not sorted_dates:
        raise ValueError(f"{data_dir}: synthetic prep 後沒有任何有效日期")

    benchmark_data = all_dfs_fast.get(benchmark_ticker)
    profile_stats = {}
    result = run_portfolio_timeline(
        all_dfs_fast=all_dfs_fast,
        all_standalone_logs=all_trade_logs,
        sorted_dates=sorted_dates,
        start_year=start_year,
        params=params,
        max_positions=max_positions,
        enable_rotation=enable_rotation,
        benchmark_ticker=benchmark_ticker,
        benchmark_data=benchmark_data,
        is_training=False,
        profile_stats=profile_stats,
        verbose=False,
    )

    (
        _df_eq,
        _df_tr,
        tot_ret,
        mdd,
        trade_count,
        win_rate,
        pf_ev,
        pf_payoff,
        final_eq,
        avg_exp,
        max_exp,
        bm_ret,
        bm_mdd,
        total_missed,
        total_missed_sells,
        r_sq,
        m_win_rate,
        bm_r_sq,
        bm_m_win_rate,
        normal_trade_count,
        extended_trade_count,
        annual_trades,
        reserved_buy_fill_rate,
        annual_return_pct,
        bm_annual_return_pct,
    ) = result

    return build_portfolio_stats_payload(
        module_path="core/v16_portfolio_engine.py",
        df_trades=_df_tr,
        total_return=tot_ret,
        mdd=mdd,
        trade_count=trade_count,
        win_rate=win_rate,
        pf_ev=pf_ev,
        pf_payoff=pf_payoff,
        final_eq=final_eq,
        avg_exp=avg_exp,
        max_exp=max_exp,
        bm_ret=bm_ret,
        bm_mdd=bm_mdd,
        total_missed=total_missed,
        total_missed_sells=total_missed_sells,
        r_sq=r_sq,
        m_win_rate=m_win_rate,
        bm_r_sq=bm_r_sq,
        bm_m_win_rate=bm_m_win_rate,
        normal_trade_count=normal_trade_count,
        extended_trade_count=extended_trade_count,
        annual_trades=annual_trades,
        reserved_buy_fill_rate=reserved_buy_fill_rate,
        annual_return_pct=annual_return_pct,
        bm_annual_return_pct=bm_annual_return_pct,
        profile_stats=profile_stats,
    )


def add_portfolio_stats_equality_checks(results, module_name, ticker, expected_stats, actual_stats):
    metric_names = [
        "total_return", "mdd", "trade_count", "win_rate", "pf_ev", "pf_payoff",
        "final_eq", "avg_exp", "max_exp", "bm_ret", "bm_mdd", "total_missed",
        "total_missed_sells", "r_sq", "m_win_rate", "bm_r_sq", "bm_m_win_rate",
        "normal_trade_count", "extended_trade_count", "annual_trades",
        "reserved_buy_fill_rate", "annual_return_pct", "bm_annual_return_pct",
        "full_year_count", "min_full_year_return_pct", "yearly_return_rows",
        "bm_full_year_count", "bm_min_full_year_return_pct", "bm_yearly_return_rows",
        "portfolio_buy_rows", "portfolio_full_exit_rows", "portfolio_half_take_profit_rows",
        "portfolio_missed_buy_rows", "portfolio_missed_sell_rows", "portfolio_period_closeout_rows",
    ]
    for metric in metric_names:
        add_check(results, module_name, ticker, metric, expected_stats[metric], actual_stats[metric])

    add_check(results, module_name, ticker, "portfolio_completed_trade_count", len(expected_stats["portfolio_completed_trades"]), len(actual_stats["portfolio_completed_trades"]))
    add_check(
        results,
        module_name,
        ticker,
        "portfolio_completed_trade_exit_dates",
        [trade["exit_date"] for trade in expected_stats["portfolio_completed_trades"]],
        [trade["exit_date"] for trade in actual_stats["portfolio_completed_trades"]],
    )
    add_check(
        results,
        module_name,
        ticker,
        "portfolio_completed_trade_pnl_sequence",
        [trade["total_pnl"] for trade in expected_stats["portfolio_completed_trades"]],
        [trade["total_pnl"] for trade in actual_stats["portfolio_completed_trades"]],
    )


def build_synthetic_half_tp_full_year_case(base_params):
    params = make_synthetic_validation_params(base_params, tp_percent=0.5)
    df = build_synthetic_baseline_frame("2024-01-01", 320)
    trigger_idx = 270
    set_synthetic_bar(df, trigger_idx, open_price=103.0, high_price=104.5, low_price=102.8, close_price=104.0)
    set_synthetic_bar(df, trigger_idx + 1, open_price=103.8, high_price=105.0, low_price=103.4, close_price=104.2)
    set_synthetic_bar(df, trigger_idx + 2, open_price=104.3, high_price=107.5, low_price=104.0, close_price=106.5)
    set_synthetic_bar(df, trigger_idx + 3, open_price=106.5, high_price=107.0, low_price=106.1, close_price=106.8)
    set_synthetic_bar(df, trigger_idx + 4, open_price=106.6, high_price=107.1, low_price=106.2, close_price=106.9)
    for idx in range(trigger_idx + 5, len(df)):
        base_close = 106.8 + (idx - (trigger_idx + 5)) * 0.01
        set_synthetic_bar(df, idx, open_price=base_close - 0.2, high_price=base_close + 0.3, low_price=base_close - 0.4, close_price=base_close)
    return {
        "case_id": "SYNTH_HALF_TP_FULL_YEAR",
        "params": params,
        "frames": {"9201": df},
        "benchmark_ticker": "9201",
        "max_positions": 1,
        "enable_rotation": False,
        "start_year": 2024,
        "primary_ticker": "9201",
    }


def build_synthetic_extended_miss_buy_case(base_params):
    params = make_synthetic_validation_params(base_params, tp_percent=0.0)
    df = build_synthetic_baseline_frame("2024-01-01", 56)
    set_synthetic_bar(df, 54, open_price=103.0, high_price=104.5, low_price=102.8, close_price=104.0)
    set_synthetic_bar(df, 55, open_price=103.8, high_price=104.0, low_price=103.6, close_price=103.9, volume=0)
    return {
        "case_id": "SYNTH_EXTENDED_MISS_BUY",
        "params": params,
        "frames": {"9301": df},
        "benchmark_ticker": "9301",
        "max_positions": 1,
        "enable_rotation": False,
        "start_year": 2024,
        "primary_ticker": "9301",
    }


def build_synthetic_competing_candidates_case(base_params):
    params = make_synthetic_validation_params(base_params, tp_percent=0.0)

    def build_frame():
        df = build_synthetic_baseline_frame("2024-01-01", 60)
        set_synthetic_bar(df, 55, open_price=103.0, high_price=104.5, low_price=102.8, close_price=104.0)
        set_synthetic_bar(df, 56, open_price=103.8, high_price=105.0, low_price=103.4, close_price=104.2)
        set_synthetic_bar(df, 57, open_price=104.8, high_price=105.3, low_price=104.4, close_price=105.0)
        set_synthetic_bar(df, 58, open_price=105.3, high_price=105.8, low_price=105.0, close_price=105.5)
        set_synthetic_bar(df, 59, open_price=105.8, high_price=106.3, low_price=105.6, close_price=106.0)
        return df

    return {
        "case_id": "SYNTH_COMPETING_CANDIDATES",
        "params": params,
        "frames": {"9401": build_frame(), "9402": build_frame()},
        "benchmark_ticker": "9401",
        "max_positions": 1,
        "enable_rotation": False,
        "start_year": 2024,
        "primary_ticker": "9402",
    }


def build_synthetic_same_day_sell_block_case(base_params):
    params = make_synthetic_validation_params(base_params, tp_percent=0.0)

    df_a = build_synthetic_baseline_frame("2024-01-01", 60)
    set_synthetic_bar(df_a, 55, open_price=103.0, high_price=104.5, low_price=102.8, close_price=104.0)
    set_synthetic_bar(df_a, 56, open_price=103.8, high_price=105.0, low_price=103.4, close_price=104.2)
    set_synthetic_bar(df_a, 57, open_price=102.5, high_price=103.0, low_price=100.5, close_price=101.5)
    set_synthetic_bar(df_a, 58, open_price=101.4, high_price=101.9, low_price=101.1, close_price=101.6)
    set_synthetic_bar(df_a, 59, open_price=101.5, high_price=102.0, low_price=101.2, close_price=101.7)

    df_b = build_synthetic_baseline_frame("2024-01-01", 60)
    set_synthetic_bar(df_b, 56, open_price=103.0, high_price=104.5, low_price=102.8, close_price=104.0)
    set_synthetic_bar(df_b, 57, open_price=103.8, high_price=105.0, low_price=103.4, close_price=104.2)
    set_synthetic_bar(df_b, 58, open_price=104.0, high_price=104.4, low_price=103.7, close_price=103.9)
    set_synthetic_bar(df_b, 59, open_price=103.8, high_price=104.1, low_price=103.5, close_price=103.7)

    return {
        "case_id": "SYNTH_SAME_DAY_SELL_BLOCK",
        "params": params,
        "frames": {"9501": df_a, "9502": df_b},
        "benchmark_ticker": "9501",
        "max_positions": 1,
        "enable_rotation": False,
        "start_year": 2024,
        "primary_ticker": "9501",
    }


def build_synthetic_unexecutable_half_tp_case(base_params):
    params = make_synthetic_validation_params(base_params, tp_percent=0.5)
    params.initial_capital = 130.0
    params.fixed_risk = 1.0

    df = build_synthetic_baseline_frame("2024-01-01", 320)
    trigger_idx = 270
    set_synthetic_bar(df, trigger_idx, open_price=103.0, high_price=104.5, low_price=102.8, close_price=104.0)
    set_synthetic_bar(df, trigger_idx + 1, open_price=103.8, high_price=105.0, low_price=103.4, close_price=104.2)
    set_synthetic_bar(df, trigger_idx + 2, open_price=104.3, high_price=107.5, low_price=104.0, close_price=106.5)
    set_synthetic_bar(df, trigger_idx + 3, open_price=106.5, high_price=107.0, low_price=106.1, close_price=106.8)
    set_synthetic_bar(df, trigger_idx + 4, open_price=106.6, high_price=107.1, low_price=106.2, close_price=106.9)
    for idx in range(trigger_idx + 5, len(df)):
        base_close = 106.8 + (idx - (trigger_idx + 5)) * 0.01
        set_synthetic_bar(df, idx, open_price=base_close - 0.2, high_price=base_close + 0.3, low_price=base_close - 0.4, close_price=base_close)

    return {
        "case_id": "SYNTH_UNEXECUTABLE_HALF_TP",
        "params": params,
        "frames": {"9601": df},
        "benchmark_ticker": "9601",
        "max_positions": 1,
        "enable_rotation": False,
        "start_year": 2024,
        "primary_ticker": "9601",
    }


def build_synthetic_rotation_t_plus_one_case(base_params):
    params = make_synthetic_validation_params(base_params, tp_percent=0.0)

    df_weak = build_synthetic_baseline_frame("2024-01-01", 140)
    set_synthetic_bar(df_weak, 20, open_price=103.0, high_price=104.5, low_price=102.8, close_price=104.0)
    set_synthetic_bar(df_weak, 21, open_price=103.8, high_price=105.0, low_price=103.4, close_price=104.2)
    set_synthetic_bar(df_weak, 22, open_price=102.5, high_price=103.0, low_price=100.5, close_price=101.5)
    set_synthetic_bar(df_weak, 23, open_price=101.4, high_price=101.9, low_price=101.1, close_price=101.6)

    set_synthetic_bar(df_weak, 70, open_price=103.0, high_price=104.5, low_price=102.8, close_price=104.0)
    set_synthetic_bar(df_weak, 71, open_price=103.8, high_price=105.0, low_price=103.4, close_price=104.2)
    for idx in range(72, len(df_weak)):
        set_synthetic_bar(df_weak, idx, open_price=104.2, high_price=104.5, low_price=103.9, close_price=104.2)

    df_strong = build_synthetic_baseline_frame("2024-01-01", 140)
    winning_bars = {
        20: (103.0, 104.5, 102.8, 104.0),
        21: (103.8, 105.0, 103.4, 104.2),
        22: (104.05535005017578, 105.7658272705768, 103.9476570515015, 105.31105462881263),
        23: (105.19304226445999, 105.8322019147044, 104.84077547539177, 105.58924143223415),
        24: (105.59205354572458, 106.13076304128805, 105.04437370111694, 105.78073381761503),
        25: (105.63103762243249, 106.23421634010937, 105.1143070306091, 105.4853284286143),
        26: (105.72662799887804, 107.34936535415748, 104.99919808701634, 106.7596940685349),
        27: (106.87008442768416, 107.19759311749098, 105.99576001905243, 106.53506299183323),
        28: (106.6015951758995, 106.77584050530604, 106.27455989056097, 106.71541978046501),
        29: (106.934605737128, 108.57791123422221, 106.42921391453338, 107.99794741355976),
        30: (107.70637243365817, 108.48096410976757, 107.3199586072383, 107.64917818953965),
        31: (107.75007011027877, 108.6877340308841, 107.15427120325072, 108.44843990554953),
        32: (108.29478643168181, 108.74458662542872, 106.93073528853994, 107.45129695383561),
        33: (107.491803398208, 107.9735796931008, 107.14966760580265, 107.26430786070408),
        34: (107.23308960356542, 107.84483201087816, 106.3789400006742, 106.86084768224212),
        35: (107.12054797677827, 107.05288044504564, 105.75719784311111, 106.06196222862346),
    }
    for idx, (o, h, l, c) in winning_bars.items():
        set_synthetic_bar(df_strong, idx, open_price=o, high_price=h, low_price=l, close_price=c)

    for idx in range(36, 100):
        set_synthetic_bar(df_strong, idx, open_price=100.0, high_price=100.4, low_price=99.6, close_price=100.0)

    set_synthetic_bar(df_strong, 100, open_price=103.0, high_price=104.5, low_price=102.8, close_price=104.0)
    set_synthetic_bar(df_strong, 101, open_price=103.8, high_price=104.0, low_price=103.7, close_price=103.9, volume=0)
    set_synthetic_bar(df_strong, 102, open_price=103.9, high_price=104.1, low_price=103.8, close_price=104.0)
    set_synthetic_bar(df_strong, 103, open_price=104.2, high_price=104.6, low_price=104.0, close_price=104.4)
    set_synthetic_bar(df_strong, 104, open_price=104.3, high_price=104.7, low_price=104.1, close_price=104.4)
    for idx in range(105, len(df_strong)):
        set_synthetic_bar(df_strong, idx, open_price=104.4, high_price=104.7, low_price=104.2, close_price=104.4)

    return {
        "case_id": "SYNTH_ROTATION_T_PLUS_ONE",
        "params": params,
        "frames": {"9701": df_weak, "9702": df_strong},
        "benchmark_ticker": "9701",
        "max_positions": 1,
        "enable_rotation": True,
        "start_year": 2024,
        "weak_ticker": "9701",
        "strong_ticker": "9702",
    }


def build_synthetic_param_guardrail_case(base_params):
    return {
        "case_id": "SYNTH_PARAM_GUARDRAIL",
        "base_payload": params_to_json_dict(make_consistency_params(base_params)),
    }


def validate_synthetic_half_tp_full_year_case(base_params):
    case = build_synthetic_half_tp_full_year_case(base_params)
    results = []
    summary = {"ticker": case["case_id"], "synthetic": True}

    with tempfile.TemporaryDirectory() as temp_dir:
        write_synthetic_csv_bundle(temp_dir, case["frames"])
        core_stats = run_portfolio_core_check_for_dir(
            temp_dir, case["params"], max_positions=case["max_positions"],
            enable_rotation=case["enable_rotation"], start_year=case["start_year"], benchmark_ticker=case["benchmark_ticker"]
        )
        sim_stats = run_portfolio_sim_tool_check_for_dir(
            temp_dir, case["params"], max_positions=case["max_positions"],
            enable_rotation=case["enable_rotation"], start_year=case["start_year"], benchmark_ticker=case["benchmark_ticker"]
        )
        add_portfolio_stats_equality_checks(results, "synthetic_half_tp_full_year", case["case_id"], core_stats, sim_stats)
        add_check(results, "synthetic_half_tp_full_year", case["case_id"], "expected_half_take_profit_rows", 1, sim_stats["portfolio_half_take_profit_rows"])
        add_check(results, "synthetic_half_tp_full_year", case["case_id"], "has_yearly_return_rows", True, bool(sim_stats["yearly_return_rows"]))

        primary_path = os.path.join(temp_dir, f"{case['primary_ticker']}.csv")
        debug_raw_df = pd.read_csv(primary_path)
        debug_input_df, _debug_sanitize_stats = sanitize_ohlcv_dataframe(
            debug_raw_df,
            case["primary_ticker"],
            min_rows=get_required_min_rows(case["params"]),
        )
        debug_df, _debug_module_path = run_debug_trade_log_check(case["primary_ticker"], debug_input_df, case["params"])
        half_rows = int((debug_df["動作"].fillna("") == "半倉停利").sum()) if debug_df is not None and len(debug_df) > 0 else 0
        add_check(results, "synthetic_half_tp_full_year", case["case_id"], "debug_half_take_profit_rows", 1, half_rows)

    summary["half_take_profit_rows"] = 1
    summary["full_year_count"] = True
    return results, summary


def validate_synthetic_extended_miss_buy_case(base_params):
    case = build_synthetic_extended_miss_buy_case(base_params)
    results = []
    summary = {"ticker": case["case_id"], "synthetic": True}

    with tempfile.TemporaryDirectory() as temp_dir:
        write_synthetic_csv_bundle(temp_dir, case["frames"])
        core_stats = run_portfolio_core_check_for_dir(
            temp_dir, case["params"], max_positions=case["max_positions"],
            enable_rotation=case["enable_rotation"], start_year=case["start_year"], benchmark_ticker=case["benchmark_ticker"]
        )
        sim_stats = run_portfolio_sim_tool_check_for_dir(
            temp_dir, case["params"], max_positions=case["max_positions"],
            enable_rotation=case["enable_rotation"], start_year=case["start_year"], benchmark_ticker=case["benchmark_ticker"]
        )
        add_portfolio_stats_equality_checks(results, "synthetic_extended_miss_buy", case["case_id"], core_stats, sim_stats)
        add_check(results, "synthetic_extended_miss_buy", case["case_id"], "expected_trade_count", 0, sim_stats["trade_count"])
        add_check(results, "synthetic_extended_miss_buy", case["case_id"], "expected_missed_buy_count", 1, sim_stats["total_missed"])
        add_check(results, "synthetic_extended_miss_buy", case["case_id"], "expected_df_trades_missed_buy_rows", 1, sim_stats["portfolio_missed_buy_rows"])

        primary_ticker = case["primary_ticker"]
        primary_path = os.path.join(temp_dir, f"{primary_ticker}.csv")
        scanner_ref_stats = run_scanner_reference_check(primary_ticker, primary_path, case["params"])
        scanner_result, _scanner_module_path = run_scanner_tool_check(primary_ticker, primary_path, case["params"])
        expected_payload = build_expected_scanner_payload(scanner_ref_stats, case["params"])
        add_check(results, "synthetic_extended_miss_buy", case["case_id"], "scanner_expected_status", "extended", expected_payload["status"])
        add_check(results, "synthetic_extended_miss_buy", case["case_id"], "scanner_tool_status", "extended", None if scanner_result is None else scanner_result["status"])
        add_check(results, "synthetic_extended_miss_buy", case["case_id"], "has_extended_candidate_today", True, bool(scanner_ref_stats.get("extended_candidate_today") is not None))

    summary["extended_candidate"] = True
    summary["missed_buy"] = True
    return results, summary


def validate_synthetic_competing_candidates_case(base_params):
    case = build_synthetic_competing_candidates_case(base_params)
    results = []
    summary = {"ticker": case["case_id"], "synthetic": True}

    with tempfile.TemporaryDirectory() as temp_dir:
        write_synthetic_csv_bundle(temp_dir, case["frames"])
        core_stats = run_portfolio_core_check_for_dir(
            temp_dir, case["params"], max_positions=case["max_positions"],
            enable_rotation=case["enable_rotation"], start_year=case["start_year"], benchmark_ticker=case["benchmark_ticker"]
        )
        sim_stats = run_portfolio_sim_tool_check_for_dir(
            temp_dir, case["params"], max_positions=case["max_positions"],
            enable_rotation=case["enable_rotation"], start_year=case["start_year"], benchmark_ticker=case["benchmark_ticker"]
        )
        add_portfolio_stats_equality_checks(results, "synthetic_competing_candidates", case["case_id"], core_stats, sim_stats)

        buy_df = sim_stats["df_trades"]
        buy_rows = buy_df[buy_df["Type"].fillna("").str.startswith("買進")].copy() if not buy_df.empty else pd.DataFrame()
        selected_ticker = buy_rows.iloc[0]["Ticker"] if len(buy_rows) > 0 else None
        rejected_buy_count = int(((buy_df["Ticker"] == "9401") & buy_df["Type"].fillna("").str.startswith("買進")).sum()) if not buy_df.empty else 0
        add_check(results, "synthetic_competing_candidates", case["case_id"], "selected_buy_row_count", 1, len(buy_rows))
        add_check(results, "synthetic_competing_candidates", case["case_id"], "selected_ticker_when_sort_ties", "9402", selected_ticker)
        add_check(results, "synthetic_competing_candidates", case["case_id"], "non_selected_ticker_has_no_buy_row", 0, rejected_buy_count)

    summary["selected_ticker"] = "9402"
    return results, summary


def validate_synthetic_same_day_sell_block_case(base_params):
    case = build_synthetic_same_day_sell_block_case(base_params)
    results = []
    summary = {"ticker": case["case_id"], "synthetic": True}

    with tempfile.TemporaryDirectory() as temp_dir:
        write_synthetic_csv_bundle(temp_dir, case["frames"])
        core_stats = run_portfolio_core_check_for_dir(
            temp_dir, case["params"], max_positions=case["max_positions"],
            enable_rotation=case["enable_rotation"], start_year=case["start_year"], benchmark_ticker=case["benchmark_ticker"]
        )
        sim_stats = run_portfolio_sim_tool_check_for_dir(
            temp_dir, case["params"], max_positions=case["max_positions"],
            enable_rotation=case["enable_rotation"], start_year=case["start_year"], benchmark_ticker=case["benchmark_ticker"]
        )
        add_portfolio_stats_equality_checks(results, "synthetic_same_day_sell_block", case["case_id"], core_stats, sim_stats)

        df_trades = sim_stats["df_trades"]
        sell_rows = df_trades[(df_trades["Ticker"] == "9501") & (df_trades["Type"].fillna("").isin(["全倉結算(停損)", "全倉結算(指標)"]))].copy() if not df_trades.empty else pd.DataFrame()
        sell_date = sell_rows.iloc[0]["Date"] if len(sell_rows) > 0 else None
        blocked_same_day_buy = bool(((df_trades["Date"] == sell_date) & (df_trades["Ticker"] == "9502") & df_trades["Type"].fillna("").str.startswith("買進")).any()) if sell_date is not None else None
        later_buy_rows = df_trades[(df_trades["Ticker"] == "9502") & df_trades["Type"].fillna("").str.startswith("買進")].copy() if not df_trades.empty else pd.DataFrame()
        later_buy_date = later_buy_rows.iloc[0]["Date"] if len(later_buy_rows) > 0 else None

        add_check(results, "synthetic_same_day_sell_block", case["case_id"], "has_sell_row", True, sell_date is not None)
        add_check(results, "synthetic_same_day_sell_block", case["case_id"], "same_day_sell_blocks_new_buy", False, blocked_same_day_buy)
        add_check(results, "synthetic_same_day_sell_block", case["case_id"], "later_reentry_is_next_day_or_after", True, (later_buy_date is not None and later_buy_date > sell_date) if sell_date is not None else False)

    summary["same_day_sell_block"] = True
    return results, summary


def validate_synthetic_unexecutable_half_tp_case(base_params):
    case = build_synthetic_unexecutable_half_tp_case(base_params)
    results = []
    summary = {"ticker": case["case_id"], "synthetic": True}

    with tempfile.TemporaryDirectory() as temp_dir:
        write_synthetic_csv_bundle(temp_dir, case["frames"])
        core_stats = run_portfolio_core_check_for_dir(
            temp_dir, case["params"], max_positions=case["max_positions"],
            enable_rotation=case["enable_rotation"], start_year=case["start_year"], benchmark_ticker=case["benchmark_ticker"]
        )
        sim_stats = run_portfolio_sim_tool_check_for_dir(
            temp_dir, case["params"], max_positions=case["max_positions"],
            enable_rotation=case["enable_rotation"], start_year=case["start_year"], benchmark_ticker=case["benchmark_ticker"]
        )
        add_portfolio_stats_equality_checks(results, "synthetic_unexecutable_half_tp", case["case_id"], core_stats, sim_stats)
        add_check(results, "synthetic_unexecutable_half_tp", case["case_id"], "expected_half_take_profit_rows", 0, sim_stats["portfolio_half_take_profit_rows"])

        primary_ticker = case["primary_ticker"]
        primary_path = os.path.join(temp_dir, f"{primary_ticker}.csv")
        debug_raw_df = pd.read_csv(primary_path)
        debug_input_df, _debug_sanitize_stats = sanitize_ohlcv_dataframe(
            debug_raw_df,
            primary_ticker,
            min_rows=get_required_min_rows(case["params"]),
        )
        debug_df, _debug_module_path = run_debug_trade_log_check(primary_ticker, debug_input_df, case["params"])
        half_rows = int((debug_df["動作"].fillna("") == "半倉停利").sum()) if debug_df is not None and len(debug_df) > 0 else 0
        buy_rows = debug_df[debug_df["動作"].fillna("").str.startswith("買進")].copy() if debug_df is not None and len(debug_df) > 0 else pd.DataFrame()
        half_tp_price_is_nan = bool(buy_rows["半倉停利價"].isna().all()) if len(buy_rows) > 0 else False
        add_check(results, "synthetic_unexecutable_half_tp", case["case_id"], "debug_half_take_profit_rows", 0, half_rows)
        add_check(results, "synthetic_unexecutable_half_tp", case["case_id"], "debug_buy_row_half_tp_price_is_nan", True, half_tp_price_is_nan)

    scanner_case = build_synthetic_half_tp_full_year_case(base_params)
    scanner_case["params"].initial_capital = 130.0
    scanner_case["params"].fixed_risk = 1.0
    scanner_case["frames"][scanner_case["primary_ticker"]] = scanner_case["frames"][scanner_case["primary_ticker"]].iloc[:271].copy()

    with tempfile.TemporaryDirectory() as temp_dir:
        write_synthetic_csv_bundle(temp_dir, scanner_case["frames"])
        scanner_ticker = scanner_case["primary_ticker"]
        scanner_path = os.path.join(temp_dir, f"{scanner_ticker}.csv")
        scanner_ref_stats = run_scanner_reference_check(scanner_ticker, scanner_path, scanner_case["params"])
        scanner_result, _scanner_module_path = run_scanner_tool_check(scanner_ticker, scanner_path, scanner_case["params"])
        proj_qty = calc_reference_candidate_qty(scanner_ref_stats["buy_limit"], scanner_ref_stats["stop_loss"], scanner_case["params"]) if scanner_ref_stats.get("is_setup_today") else 0
        scanner_message = "" if scanner_result is None or scanner_result.get("message") is None else str(scanner_result.get("message"))
        add_check(results, "synthetic_unexecutable_half_tp", case["case_id"], "scanner_projected_qty", 1, proj_qty)
        add_check(results, "synthetic_unexecutable_half_tp", case["case_id"], "scanner_half_tp_executable", False, can_execute_half_take_profit(proj_qty, scanner_case["params"].tp_percent))
        add_check(results, "synthetic_unexecutable_half_tp", case["case_id"], "scanner_tool_status", "buy", None if scanner_result is None else scanner_result["status"])
        add_check(results, "synthetic_unexecutable_half_tp", case["case_id"], "scanner_message_marks_unexecutable_half_tp", True, "半倉停利:股數不足" in scanner_message)

    summary["half_take_profit_rows"] = 0
    summary["projected_qty"] = 1
    return results, summary


def validate_synthetic_rotation_t_plus_one_case(base_params):
    case = build_synthetic_rotation_t_plus_one_case(base_params)
    results = []
    summary = {"ticker": case["case_id"], "synthetic": True}

    rotation_sell_type = "汰弱賣出(Open, T+1再評估買進)"

    with tempfile.TemporaryDirectory() as temp_dir:
        write_synthetic_csv_bundle(temp_dir, case["frames"])
        core_stats = run_portfolio_core_check_for_dir(
            temp_dir,
            case["params"],
            max_positions=case["max_positions"],
            enable_rotation=case["enable_rotation"],
            start_year=case["start_year"],
            benchmark_ticker=case["benchmark_ticker"],
        )
        sim_stats = run_portfolio_sim_tool_check_for_dir(
            temp_dir,
            case["params"],
            max_positions=case["max_positions"],
            enable_rotation=case["enable_rotation"],
            start_year=case["start_year"],
            benchmark_ticker=case["benchmark_ticker"],
        )
        add_portfolio_stats_equality_checks(results, "synthetic_rotation_t_plus_one", case["case_id"], core_stats, sim_stats)

        df_trades = sim_stats["df_trades"].copy()
        if df_trades.empty:
            add_fail_result(
                results,
                "synthetic_rotation_t_plus_one",
                case["case_id"],
                "df_trades_exists",
                "non-empty",
                "empty",
                "rotation synthetic case 應產生 trade history。"
            )
            return results, summary

        rotation_rows = df_trades[
            (df_trades["Ticker"] == case["weak_ticker"]) &
            (df_trades["Type"].fillna("") == rotation_sell_type)
        ].copy()
        rotation_sell_date = rotation_rows.iloc[0]["Date"] if len(rotation_rows) > 0 else None

        same_day_reentry = bool(
            (
                (df_trades["Date"] == rotation_sell_date) &
                (df_trades["Ticker"] == case["strong_ticker"]) &
                df_trades["Type"].fillna("").str.startswith("買進")
            ).any()
        ) if rotation_sell_date is not None else False

        post_rotation_df = (
            df_trades[pd.to_datetime(df_trades["Date"]) > pd.to_datetime(rotation_sell_date)].copy()
            if rotation_sell_date is not None else pd.DataFrame()
        )

        extended_miss_rows = post_rotation_df[
            (post_rotation_df["Ticker"] == case["strong_ticker"]) &
            (post_rotation_df["Type"].fillna("") == "錯失買進(延續候選)")
        ].copy()
        extended_miss_date = extended_miss_rows.iloc[0]["Date"] if len(extended_miss_rows) > 0 else None

        delayed_buy_rows = post_rotation_df[
            (post_rotation_df["Ticker"] == case["strong_ticker"]) &
            post_rotation_df["Type"].fillna("").str.startswith("買進")
        ].copy()
        delayed_buy_date = delayed_buy_rows.iloc[0]["Date"] if len(delayed_buy_rows) > 0 else None

        add_check(results, "synthetic_rotation_t_plus_one", case["case_id"], "rotation_sell_row_count", 1, len(rotation_rows))
        add_check(results, "synthetic_rotation_t_plus_one", case["case_id"], "rotation_same_day_reentry_blocked", False, same_day_reentry)
        add_check(results, "synthetic_rotation_t_plus_one", case["case_id"], "rotation_has_post_sell_extended_miss_buy", True, extended_miss_date is not None)
        add_check(
            results,
            "synthetic_rotation_t_plus_one",
            case["case_id"],
            "rotation_extended_miss_occurs_after_sell",
            True,
            (extended_miss_date is not None and rotation_sell_date is not None and extended_miss_date > rotation_sell_date)
        )
        add_check(
            results,
            "synthetic_rotation_t_plus_one",
            case["case_id"],
            "rotation_delayed_buy_occurs_after_sell",
            True,
            (delayed_buy_date is not None and rotation_sell_date is not None and delayed_buy_date > rotation_sell_date)
        )
        add_check(
            results,
            "synthetic_rotation_t_plus_one",
            case["case_id"],
            "rotation_delayed_buy_occurs_after_extended_miss",
            True,
            (delayed_buy_date is not None and extended_miss_date is not None and delayed_buy_date > extended_miss_date)
        )

    summary["rotation_sell"] = True
    summary["delayed_reentry"] = True
    return results, summary


def validate_synthetic_history_ev_threshold_case(base_params):
    params = make_synthetic_validation_params(base_params, tp_percent=0.0)
    params.min_history_trades = 1
    params.min_history_ev = 0.5
    params.min_history_win_rate = 0.5

    case_id = "SYNTH_HISTORY_EV_THRESHOLD"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    is_candidate, expected_value, win_rate, trade_count = evaluate_history_candidate_metrics(
        trade_count=1,
        win_count=1,
        total_r_sum=0.5,
        win_r_sum=0.5,
        loss_r_sum=0.0,
        params=params,
    )

    add_check(results, "synthetic_history_ev_threshold", case_id, "expected_value_equals_threshold", 0.5, expected_value)
    add_check(results, "synthetic_history_ev_threshold", case_id, "win_rate_equals_threshold", 1.0, win_rate)
    add_check(results, "synthetic_history_ev_threshold", case_id, "trade_count_preserved", 1, trade_count)
    add_check(results, "synthetic_history_ev_threshold", case_id, "candidate_allowed_when_ev_equals_threshold", True, is_candidate)

    summary["candidate_allowed"] = bool(is_candidate)
    summary["expected_value"] = expected_value
    return results, summary


def validate_synthetic_proj_cost_cash_capped_case(base_params):
    params = make_synthetic_validation_params(base_params, tp_percent=0.0)
    params.initial_capital = 1_000_000.0
    params.fixed_risk = 0.01

    case_id = "SYNTH_PROJ_COST_CASH_CAPPED_ORDER"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    sizing_equity = 1_000_000.0
    available_cash = 50_000.0
    candidate_rows = [
        {"ticker": "9801", "limit_px": 10.0, "init_sl": 9.5, "init_trail": 9.0},
        {"ticker": "9802", "limit_px": 14.0, "init_sl": 13.3, "init_trail": 12.8},
    ]

    estimated_rank_rows = []
    for cand in candidate_rows:
        est_plan = resize_candidate_plan_to_capital(
            {
                "limit_price": cand["limit_px"],
                "init_sl": cand["init_sl"],
                "init_trail": cand["init_trail"],
            },
            sizing_equity,
            params,
        )
        est_reserved_cost = calc_entry_price(est_plan["limit_price"], est_plan["qty"], params) * est_plan["qty"]
        estimated_rank_rows.append({
            "ticker": cand["ticker"],
            "reserved_cost": est_reserved_cost,
        })

    estimated_rank_rows.sort(key=lambda x: (x["reserved_cost"], x["ticker"]), reverse=True)
    stale_top_ticker = estimated_rank_rows[0]["ticker"]

    cash_capped_rank_rows = []
    for cand in candidate_rows:
        cash_capped_plan = build_cash_capped_entry_plan(
            {
                "limit_price": cand["limit_px"],
                "init_sl": cand["init_sl"],
                "init_trail": cand["init_trail"],
            },
            available_cash,
            params,
        )
        if cash_capped_plan is None:
            continue

        cash_capped_rank_rows.append({
            "ticker": cand["ticker"],
            "reserved_cost": cash_capped_plan["reserved_cost"],
            "qty": cash_capped_plan["qty"],
        })

    cash_capped_rank_rows.sort(key=lambda x: (x["reserved_cost"], x["ticker"]), reverse=True)
    cash_capped_top_ticker = cash_capped_rank_rows[0]["ticker"] if cash_capped_rank_rows else None
    cash_capped_top_reserved_cost = cash_capped_rank_rows[0]["reserved_cost"] if cash_capped_rank_rows else None
    cash_capped_top_qty = cash_capped_rank_rows[0]["qty"] if cash_capped_rank_rows else None

    add_check(
        results,
        "synthetic_proj_cost_cash_capped",
        case_id,
        "stale_proj_cost_top_ticker",
        "9802",
        stale_top_ticker,
    )
    add_check(
        results,
        "synthetic_proj_cost_cash_capped",
        case_id,
        "cash_capped_proj_cost_top_ticker",
        "9801",
        cash_capped_top_ticker,
    )
    add_check(
        results,
        "synthetic_proj_cost_cash_capped",
        case_id,
        "proj_cost_order_reversal_detected",
        True,
        stale_top_ticker != cash_capped_top_ticker,
    )
    add_check(
        results,
        "synthetic_proj_cost_cash_capped",
        case_id,
        "cash_capped_reserved_cost_within_available_cash",
        True,
        cash_capped_top_reserved_cost is not None and cash_capped_top_reserved_cost <= available_cash,
    )
    add_check(
        results,
        "synthetic_proj_cost_cash_capped",
        case_id,
        "cash_capped_qty_positive",
        True,
        cash_capped_top_qty is not None and cash_capped_top_qty > 0,
    )

    summary["stale_top_ticker"] = stale_top_ticker
    summary["cash_capped_top_ticker"] = cash_capped_top_ticker
    summary["available_cash"] = available_cash
    return results, summary


def validate_synthetic_param_guardrail_case(base_params):
    case = build_synthetic_param_guardrail_case(base_params)
    results = []
    summary = {"ticker": case["case_id"], "synthetic": True}

    valid_params = build_params_from_mapping(case["base_payload"])
    add_check(results, "synthetic_param_guardrail", case["case_id"], "valid_payload_loads", True, isinstance(valid_params, V16StrategyParams))

    invalid_cases = [
        ("tp_percent_ge_1_rejected", {**case["base_payload"], "tp_percent": 1.0}, "tp_percent"),
        ("fixed_risk_zero_rejected", {**case["base_payload"], "fixed_risk": 0.0}, "fixed_risk"),
        ("min_history_win_rate_gt_1_rejected", {**case["base_payload"], "min_history_win_rate": 1.1}, "min_history_win_rate"),
        ("vol_long_len_lt_short_rejected", {**case["base_payload"], "vol_short_len": 10, "vol_long_len": 5}, "vol_long_len"),
        ("use_bb_string_type_rejected", {**case["base_payload"], "use_bb": "abc"}, "use_bb"),
    ]

    runtime_valid_payload = {**case["base_payload"], "optimizer_max_workers": 3, "scanner_max_workers": 4}
    runtime_params = build_params_from_mapping(runtime_valid_payload)
    add_check(results, "synthetic_param_guardrail", case["case_id"], "runtime_optimizer_max_workers_loads", 3, getattr(runtime_params, "optimizer_max_workers", None))
    add_check(results, "synthetic_param_guardrail", case["case_id"], "runtime_scanner_max_workers_loads", 4, getattr(runtime_params, "scanner_max_workers", None))

    runtime_invalid_cases = [
        ("optimizer_max_workers_zero_rejected", {**case["base_payload"], "optimizer_max_workers": 0}, "optimizer_max_workers"),
        ("scanner_max_workers_zero_rejected", {**case["base_payload"], "scanner_max_workers": 0}, "scanner_max_workers"),
    ]

    for metric_name, payload, expected_field in invalid_cases + runtime_invalid_cases:
        try:
            build_params_from_mapping(payload)
            add_fail_result(
                results,
                "synthetic_param_guardrail",
                case["case_id"],
                metric_name,
                f"ValueError containing {expected_field}",
                "loaded_ok",
                "非法參數不應成功載入。"
            )
        except ValueError as e:
            add_check(results, "synthetic_param_guardrail", case["case_id"], metric_name, True, expected_field in str(e))

    for metric_name, payload, expected_field in invalid_cases:
        try:
            V16StrategyParams(**payload)
            add_fail_result(
                results,
                "synthetic_param_guardrail",
                case["case_id"],
                f"direct_dataclass_{metric_name}",
                f"ValueError containing {expected_field}",
                "loaded_ok",
                "直接建立 V16StrategyParams 也不應繞過 guardrail。"
            )
        except ValueError as e:
            add_check(
                results,
                "synthetic_param_guardrail",
                case["case_id"],
                f"direct_dataclass_{metric_name}",
                True,
                expected_field in str(e)
            )

    runtime_mutation_params = V16StrategyParams()
    try:
        runtime_mutation_params.optimizer_max_workers = 0
        add_fail_result(
            results,
            "synthetic_param_guardrail",
            case["case_id"],
            "direct_runtime_attr_guardrail",
            "ValueError containing optimizer_max_workers",
            "setattr_ok",
            "runtime worker 設定直接改欄位也不應繞過 guardrail。"
        )
    except ValueError as e:
        add_check(results, "synthetic_param_guardrail", case["case_id"], "direct_runtime_attr_guardrail", True, "optimizer_max_workers" in str(e))

    invalid_direct_setattr_cases = [
        ("direct_setattr_use_bb_string_rejected", "use_bb", "abc", "use_bb"),
        ("direct_setattr_high_len_string_rejected", "high_len", "10", "high_len"),
    ]

    for metric_name, field_name, invalid_value, expected_field in invalid_direct_setattr_cases:
        mutation_target = V16StrategyParams()
        try:
            setattr(mutation_target, field_name, invalid_value)
            add_fail_result(
                results,
                "synthetic_param_guardrail",
                case["case_id"],
                metric_name,
                f"ValueError containing {expected_field}",
                "setattr_ok",
                "直接改 dataclass 欄位不應繞過型別 guardrail。"
            )
        except ValueError as e:
            add_check(results, "synthetic_param_guardrail", case["case_id"], metric_name, True, expected_field in str(e))

    mutation_params = V16StrategyParams()
    try:
        mutation_params.tp_precent = 0.3
        add_fail_result(
            results,
            "synthetic_param_guardrail",
            case["case_id"],
            "unknown_attr_typo_rejected",
            "AttributeError containing tp_precent",
            "setattr_ok",
            "未知屬性 typo 不應靜默掛到 params 物件上。"
        )
    except AttributeError as e:
        add_check(results, "synthetic_param_guardrail", case["case_id"], "unknown_attr_typo_rejected", True, "tp_precent" in str(e))

    default_params_arg = run_v16_backtest.__defaults__[0] if run_v16_backtest.__defaults__ else None
    add_check(results, "synthetic_param_guardrail", case["case_id"], "run_v16_backtest_default_params_is_none", True, default_params_arg is None)

    summary["guardrail_cases"] = (len(invalid_cases) * 2) + len(runtime_invalid_cases) + len(invalid_direct_setattr_cases) + 4
    return results, summary


def run_synthetic_consistency_suite(base_params):
    all_results = []
    summaries = []
    validators = [
        validate_synthetic_half_tp_full_year_case,
        validate_synthetic_extended_miss_buy_case,
        validate_synthetic_competing_candidates_case,
        validate_synthetic_same_day_sell_block_case,
        validate_synthetic_unexecutable_half_tp_case,
        validate_synthetic_rotation_t_plus_one_case,
        validate_synthetic_proj_cost_cash_capped_case,
        validate_synthetic_history_ev_threshold_case,
        validate_synthetic_param_guardrail_case,
    ]

    for validator in validators:
        results, summary = validator(base_params)
        all_results.extend(results)
        summaries.append(summary)

    return all_results, summaries




def write_issue_excel_report(df_failed, df_failed_summary, df_failed_module, timestamp):
    from openpyxl.styles import numbers

    if df_failed.empty:
        return None

    report_path = os.path.join(OUTPUT_DIR, f"v16_consistency_issues_{timestamp}.xlsx")

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        df_failed.to_excel(writer, sheet_name="failed_only", index=False)
        df_failed_summary.to_excel(writer, sheet_name="failed_tickers", index=False)
        df_failed_module.to_excel(writer, sheet_name="failed_modules", index=False)

        for sheet_name in ["failed_only", "failed_tickers"]:
            ws = writer.book[sheet_name]
            header_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
            if "ticker" in header_map:
                col_idx = header_map["ticker"]
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.number_format = numbers.FORMAT_TEXT
                    cell.value = normalize_ticker_text(cell.value)

    return report_path


def print_console_summary(df_results, df_failed, df_summary, csv_path, xlsx_path, elapsed_time, real_summary_count, real_tickers):
    synthetic_summary_count = 0
    synthetic_ticker_set = set()
    if not df_summary.empty and "synthetic" in df_summary.columns:
        synthetic_mask = df_summary["synthetic"].fillna(False).astype(bool)
        synthetic_summary_count = int(synthetic_mask.sum())
        synthetic_ticker_set = {
            normalize_ticker_text(ticker)
            for ticker in df_summary.loc[synthetic_mask, "ticker"].dropna().tolist()
        }

    system_summary_count = max(len(df_summary) - real_summary_count - synthetic_summary_count, 0)

    real_ticker_set = {normalize_ticker_text(ticker) for ticker in real_tickers}
    failed_ticker_series = (
        df_failed["ticker"].fillna("").map(normalize_ticker_text)
        if not df_failed.empty else pd.Series(dtype="object")
    )
    failed_real_mask = failed_ticker_series.isin(real_ticker_set) if not df_failed.empty else pd.Series(dtype=bool)
    failed_synthetic_mask = failed_ticker_series.isin(synthetic_ticker_set) if not df_failed.empty else pd.Series(dtype=bool)
    failed_system_mask = ~(failed_real_mask | failed_synthetic_mask) if not df_failed.empty else pd.Series(dtype=bool)

    failed_real_tickers = int(failed_ticker_series[failed_real_mask].nunique()) if not df_failed.empty else 0
    failed_synthetic_cases = int(failed_ticker_series[failed_synthetic_mask].nunique()) if not df_failed.empty else 0
    failed_system_items = int(failed_ticker_series[failed_system_mask].nunique()) if not df_failed.empty else 0

    pass_count = int((df_results["status"] == "PASS").sum()) if not df_results.empty else 0
    skip_count = int((df_results["status"] == "SKIP").sum()) if not df_results.empty else 0
    fail_count = int((df_results["status"] == "FAIL").sum()) if not df_results.empty else 0

    print("\n================================================================================")
    print("一致性回歸摘要")
    print("================================================================================")
    print(f"耗時: {elapsed_time:.2f} 秒")
    print(f"成功進入 summary 的真實股票數: {real_summary_count}")
    print(f"synthetic case 數: {synthetic_summary_count}")
    print(f"system 檢查列數: {system_summary_count}")
    print(f"summary 總列數: {len(df_summary)}")
    print(f"總檢查數: {len(df_results)}")
    print(f"PASS 數: {pass_count}")
    print(f"SKIP 數: {skip_count}")
    print(f"FAIL 數: {fail_count}")
    print(f"有問題真實股票數: {failed_real_tickers}")
    print(f"有問題 synthetic case 數: {failed_synthetic_cases}")
    print(f"有問題 system 項目數: {failed_system_items}")
    print(f"完整 CSV: {csv_path}")
    print(f"問題 Excel: {xlsx_path if xlsx_path else '無，因為沒有 failed 項'}")

    if df_failed.empty:
        print("\n失敗項摘要：無")
        return

    print("\n失敗項前覽：")
    show_cols = ["ticker", "module", "metric", "expected", "actual", "note"]
    preview_df = df_failed[show_cols].head(MAX_CONSOLE_FAIL_PREVIEW).copy()
    print(preview_df.to_string(index=False))

    remain_count = len(df_failed) - len(preview_df)
    if remain_count > 0:
        print(f"\n... 尚有 {remain_count} 筆 FAIL 未顯示，請直接查看 CSV / Excel。")

    failed_real_summary = (
        df_failed.loc[failed_real_mask]
        .groupby("ticker", dropna=False)
        .agg(failed_checks=("passed", "size"))
        .reset_index()
        .sort_values(by=["failed_checks", "ticker"], ascending=[False, True])
        .head(MAX_CONSOLE_FAIL_PREVIEW)
        if not df_failed.empty else pd.DataFrame()
    )

    if not failed_real_summary.empty:
        print("\n失敗真實股票前覽：")
        print(failed_real_summary.to_string(index=False))

    failed_non_real_summary = (
        df_failed.loc[~failed_real_mask]
        .groupby("ticker", dropna=False)
        .agg(failed_checks=("passed", "size"))
        .reset_index()
        .sort_values(by=["failed_checks", "ticker"], ascending=[False, True])
        .head(MAX_CONSOLE_FAIL_PREVIEW)
        if not df_failed.empty else pd.DataFrame()
    )

    if not failed_non_real_summary.empty:
        print("\n失敗 synthetic/system 前覽：")
        print(failed_non_real_summary.to_string(index=False))


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    base_params = load_params()
    all_results = []
    summaries = []
    start_time = time.time()

    selected_tickers = []
    real_data_unavailable_reason = None
    if not os.path.isdir(DATA_DIR):
        real_data_unavailable_reason = f"找不到資料夾: {DATA_DIR}"
        print(real_data_unavailable_reason)
        print("將只執行 synthetic coverage suite，但本次驗證不可視為完整通過。")
    else:
        selected_tickers = discover_available_tickers()
        if not selected_tickers:
            real_data_unavailable_reason = f"資料夾內找不到任何 CSV: {DATA_DIR}"
            print(real_data_unavailable_reason)
            print("將只執行 synthetic coverage suite，但本次驗證不可視為完整通過。")

    ticker_pass_count = 0
    ticker_skip_count = 0
    ticker_fail_count = 0

    total_tickers = len(selected_tickers)
    if total_tickers > 0:
        print(f"開始自動掃描 {total_tickers} 檔股票...")

    for idx, ticker in enumerate(selected_tickers, start=1):
        ticker_results_before = len(all_results)

        try:
            results, summary = validate_one_ticker(ticker, base_params)
            all_results.extend(results)
            summaries.append(summary)
        except VALIDATION_RECOVERABLE_EXCEPTIONS as e:
            if is_insufficient_data_error(e):
                add_skip_result(
                    all_results,
                    "system",
                    ticker,
                    "validation_runtime",
                    f"資料不足，跳過驗證。({type(e).__name__}: {e})"
                )
                summaries.append({
                    "ticker": ticker,
                    "validation_runtime": f"SKIP: {format_exception_summary(e)}",
                })
            else:
                add_fail_result(
                    all_results,
                    "system",
                    ticker,
                    "validation_runtime",
                    "no exception",
                    format_exception_summary(e),
                    "單一 ticker 的 runtime / import side effect / 路徑權限問題，不可讓整體 validate 中斷。"
                )
                summaries.append({
                    "ticker": ticker,
                    "validation_runtime": f"FAIL: {format_exception_summary(e)}",
                })

        ticker_results = all_results[ticker_results_before:]
        ticker_statuses = {row["status"] for row in ticker_results}

        if "FAIL" in ticker_statuses:
            ticker_fail_count += 1
        elif "PASS" in ticker_statuses:
            ticker_pass_count += 1
        else:
            ticker_skip_count += 1

        print(
            f"\r進度: [{idx}/{total_tickers}] 目前: {ticker:<8} | PASS股票:{ticker_pass_count} | SKIP股票:{ticker_skip_count} | FAIL股票:{ticker_fail_count}",
            end="",
            flush=True
        )

    if total_tickers > 0:
        print(" " * 160, end="\r")
        print()

    print("開始執行 synthetic coverage suite...")
    try:
        synthetic_results, synthetic_summaries = run_synthetic_consistency_suite(base_params)
        all_results.extend(synthetic_results)
        summaries.extend(synthetic_summaries)
    except VALIDATION_RECOVERABLE_EXCEPTIONS as e:
        add_fail_result(
            all_results,
            "synthetic_suite",
            "SYNTHETIC_SUITE",
            "runtime",
            "suite runs successfully",
            format_exception_summary(e),
            "synthetic coverage suite 失敗時不可靜默略過，否則 miss buy / half TP / 多檔互動覆蓋會出現假象。"
        )
        summaries.append({
            "ticker": "SYNTHETIC_SUITE",
            "validation_runtime": f"FAIL: {format_exception_summary(e)}",
            "synthetic": True,
        })

    if real_data_unavailable_reason is not None:
        add_fail_result(
            all_results,
            "system",
            "REAL_DATA_COVERAGE",
            "real_data_scan_required",
            "至少 1 檔真實股票完成 validate",
            real_data_unavailable_reason,
            "最嚴格檢查不可只靠 synthetic coverage suite；若真實資料缺失，本次結果只能視為工具與合成案例檢查，不可視為完整通過。"
        )
        summaries.append({
            "ticker": "REAL_DATA_COVERAGE",
            "validation_runtime": f"FAIL: {real_data_unavailable_reason}",
            "synthetic": False,
        })

    df_results = pd.DataFrame(all_results)
    df_summary = pd.DataFrame(summaries)
    df_failed = df_results[df_results["status"] == "FAIL"].copy() if not df_results.empty else pd.DataFrame()

    for df_obj in [df_results, df_summary, df_failed]:
        if not df_obj.empty and "ticker" in df_obj.columns:
            df_obj["ticker"] = df_obj["ticker"].map(normalize_ticker_text)

    if not df_failed.empty:
        df_failed = df_failed.sort_values(by=["ticker", "module", "metric"]).reset_index(drop=True)

    timestamp = time.strftime("%Y%m%d_%H%M%S")
    csv_path = os.path.join(OUTPUT_DIR, f"v16_consistency_full_scan_{timestamp}.csv")
    df_results.to_csv(csv_path, index=False, encoding="utf-8-sig")

    if df_failed.empty:
        df_failed_summary = pd.DataFrame(columns=["ticker", "failed_checks"])
        df_failed_module = pd.DataFrame(columns=["module", "failed_checks"])
    else:
        df_failed_summary = (
            df_failed.groupby("ticker", dropna=False)
            .agg(failed_checks=("passed", "size"))
            .reset_index()
            .sort_values(by=["failed_checks", "ticker"], ascending=[False, True])
        )
        df_failed_module = (
            df_failed.groupby("module", dropna=False)
            .agg(failed_checks=("passed", "size"))
            .reset_index()
            .sort_values(by=["failed_checks", "module"], ascending=[False, True])
        )

    xlsx_path = write_issue_excel_report(
        df_failed=df_failed,
        df_failed_summary=df_failed_summary,
        df_failed_module=df_failed_module,
        timestamp=timestamp
    )

    elapsed_time = time.time() - start_time

    print_console_summary(
        df_results=df_results,
        df_failed=df_failed,
        df_summary=df_summary,
        csv_path=csv_path,
        xlsx_path=xlsx_path,
        elapsed_time=elapsed_time,
        real_summary_count=total_tickers,
        real_tickers=selected_tickers
    )

    return 1 if (not df_failed.empty or real_data_unavailable_reason is not None) else 0


if __name__ == "__main__":
    sys.exit(main())
