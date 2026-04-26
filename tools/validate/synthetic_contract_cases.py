import ast
import csv
import io
import importlib
import json
import os
import sys
import tempfile
import zipfile
from contextlib import redirect_stdout
from pathlib import Path
from unittest.mock import patch

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from core import log_utils as log_utils_module
from core.output_retention import RetentionRule, apply_retention_rules
from core.data_utils import get_required_min_rows, sanitize_ohlcv_dataframe
from tools.local_regression import common as local_common
from tools.local_regression import run_all as run_all_module
from tools.local_regression import run_quick_gate as run_quick_gate_module
from tools.local_regression.meta_quality_coverage import build_coverage_summary
from tools.local_regression.meta_quality_performance import build_performance_summary
from tools.local_regression.meta_quality_targets import COVERAGE_BRANCH_MIN_FLOOR, COVERAGE_LINE_MIN_FLOOR, COVERAGE_TARGETS, CRITICAL_COVERAGE_TARGETS
from tools.optimizer.profile import OptimizerProfileRecorder, PROFILE_FIELDS
from tools.local_regression.common import LOCAL_REGRESSION_RUN_DIR_ENV, write_json, write_csv, write_text
from tools.validate.reporting import write_issue_excel_report, write_local_regression_summary
from core.portfolio_fast_data import prep_stock_data_and_trades, build_trade_stats_index
from core.exact_accounting import calc_entry_total_cost
from core.price_utils import calc_reference_candidate_qty, calc_entry_price
from tools.scanner.stock_processor import build_scanner_response_from_stats
from tools.trade_analysis.charting import (
    bind_matplotlib_chart_navigation,
    build_debug_chart_payload,
    normalize_chart_payload_contract,
    compute_visible_value_ranges,
    create_debug_chart_context,
    create_matplotlib_trade_chart_figure,
    get_matplotlib_cjk_font_candidates,
    record_signal_annotation,
    set_chart_status_box,
    set_chart_summary_box,
)
from tools.trade_analysis.backtest import _record_buy_signal_annotation
from tools.trade_analysis.entry_flow import _record_entry_plan_marker
from tools.trade_analysis.exit_flow import append_debug_forced_closeout

from .checks import add_check, make_synthetic_validation_params, run_scanner_reference_check, run_scanner_reference_check_on_clean_df
from .synthetic_case_builders import build_synthetic_competing_candidates_case
from .synthetic_frame_utils import write_synthetic_csv_bundle
from . import module_loader as module_loader_module
from .module_loader import build_project_absolute_path, normalize_project_relative_path
from .tool_adapters import (
    run_debug_trade_log_check,
    run_portfolio_sim_tool_check,
    run_portfolio_sim_tool_check_for_dir,
    run_scanner_tool_check,
)


REQUIRED_VALIDATE_SUMMARY_KEYS = {
    "status",
    "dataset",
    "dataset_source",
    "data_dir",
    "csv_path",
    "xlsx_path",
    "output_dir",
    "elapsed_time_sec",
    "real_ticker_count",
    "real_data_coverage_ok",
    "total_checks",
    "pass_count",
    "skip_count",
    "fail_count",
    "peak_traced_memory_mb",
}

REQUIRED_PROFILE_SUMMARY_AVG_KEYS = {
    "objective_wall_sec",
    "prep_wall_sec",
    "portfolio_wall_sec",
    "score_calc_sec",
}

REQUIRED_PREFLIGHT_SUMMARY_KEYS = {
    "status",
    "python_executable",
    "duration_sec",
    "failed_packages",
}

REQUIRED_DATASET_PREP_SUMMARY_KEYS = {
    "status",
    "duration_sec",
    "dataset_dir",
    "source",
    "csv_count",
    "csv_total_bytes",
    "csv_members_sha256",
    "csv_content_sha256",
    "fingerprint_algorithm",
    "reused_existing",
}

REQUIRED_CHAIN_SUMMARY_KEYS = {
    "status",
    "dataset",
    "dataset_info",
    "ticker_count",
    "csv_count",
    "duplicate_issue_count",
    "skipped_ticker_count",
    "detail_count",
    "rows",
    "highlights",
    "portfolio_snapshot",
    "scanner_snapshot",
    "rerun_consistency",
    "failures",
    "peak_traced_memory_mb",
}

CHAIN_SUMMARY_CSV_FIELDS = [
    "ticker",
    "single_trade_count",
    "single_missed_buys",
    "setup_days",
    "pit_pass_days",
    "candidate_days",
    "orderable_days",
    "portfolio_trade_rows",
    "filled_count",
    "portfolio_missed_buy_count",
    "debug_row_count",
    "blocked_by",
    "sanitize_dropped",
]

def _normalize_nan_records(df):
    if df is None:
        return None

    def _normalize_value(value):
        return None if pd.isna(value) else value

    records = []
    for row in df.to_dict("records"):
        records.append({key: _normalize_value(value) for key, value in row.items()})
    return records


REQUIRED_ML_SMOKE_SUMMARY_KEYS = {
    "status",
    "dataset",
    "dataset_info",
    "db_path",
    "db_trial_count",
    "qualified_trial_count",
    "best_trial_value",
    "run_best_params_path",
    "run_best_params_required",
    "best_params_keys",
    "optimizer_profile_summary_path",
    "optimizer_profile_trial_count",
    "optimizer_profile_avg_objective_wall_sec",
    "optimizer_repro",
    "failures",
    "peak_traced_memory_mb",
}

REQUIRED_QUICK_GATE_SUMMARY_KEYS = {
    "status",
    "dataset",
    "dataset_info",
    "step_count",
    "failed_count",
    "failed_steps",
    "steps",
    "duration_sec",
    "peak_traced_memory_mb",
}

REQUIRED_META_QUALITY_SUMMARY_KEYS = {
    "status",
    "failures",
    "fail_count",
    "coverage",
    "checklist",
    "formal_entry",
    "performance",
    "results",
}

REQUIRED_META_QUALITY_FORMAL_ENTRY_KEYS = {
    "ok",
    "registry_steps",
    "registry_commands",
    "run_all_steps",
    "preflight_steps",
    "test_suite_steps",
}

REQUIRED_MASTER_SUMMARY_KEYS = {
    "overall_status",
    "dataset",
    "dataset_info",
    "timestamp",
    "git_commit",
    "scripts",
    "selected_steps",
    "failures",
    "preflight",
    "dataset_prepare",
    "payload_failures",
    "not_run_step_names",
    "bundle_mode",
    "bundle_entries",
    "failed_step_names",
    "suggested_rerun_command",
}


def _find_step_payload(steps, name):
    return next((item for item in steps if item.get("name") == name), None)



def _read_csv_rows(csv_path: Path):
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)
        return reader.fieldnames or [], rows


def validate_output_contract_case(_base_params):
    case_id = "OUTPUT_CONTRACT_SCHEMA"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    with tempfile.TemporaryDirectory(prefix="output_contract_") as temp_dir:
        temp_path = Path(temp_dir)

        # JSON contract: validate_consistency_summary.json
        run_dir = temp_path / "run_dir"
        run_dir.mkdir(parents=True, exist_ok=True)
        prev_run_dir = os.environ.get(LOCAL_REGRESSION_RUN_DIR_ENV)
        os.environ[LOCAL_REGRESSION_RUN_DIR_ENV] = str(run_dir)
        try:
            df_results = pd.DataFrame(
                [
                    {"ticker": "1101", "status": "PASS"},
                    {"ticker": "1102", "status": "FAIL"},
                    {"ticker": "1103", "status": "SKIP"},
                ]
            )
            df_failed = pd.DataFrame(
                [
                    {
                        "ticker": "1102",
                        "module": "portfolio",
                        "metric": "final_equity",
                        "passed": False,
                        "expected": 123.0,
                        "actual": 120.0,
                        "note": "mismatch",
                    }
                ]
            )
            write_local_regression_summary(
                dataset_profile_key="reduced",
                dataset_source="direct",
                data_dir="data/tw_stock_data_vip_reduced",
                csv_path="outputs/validate_consistency/consistency_full_scan.csv",
                xlsx_path="outputs/validate_consistency/consistency_issues.xlsx",
                elapsed_time=12.345,
                selected_tickers=["1101", "1102", "1103"],
                df_results=df_results,
                df_failed=df_failed,
                output_dir="outputs/validate_consistency",
                real_data_coverage_ok=True,
                peak_traced_memory_mb=12.345,
            )
        finally:
            if prev_run_dir is None:
                os.environ.pop(LOCAL_REGRESSION_RUN_DIR_ENV, None)
            else:
                os.environ[LOCAL_REGRESSION_RUN_DIR_ENV] = prev_run_dir

        validate_summary_path = run_dir / "validate_consistency_summary.json"
        add_check(results, "output_contract", case_id, "validate_summary_exists", True, validate_summary_path.exists())
        validate_payload = json.loads(validate_summary_path.read_text(encoding="utf-8"))
        add_check(results, "output_contract", case_id, "validate_summary_required_keys_present", [], sorted(REQUIRED_VALIDATE_SUMMARY_KEYS - set(validate_payload.keys())))
        add_check(results, "output_contract", case_id, "validate_summary_status", "FAIL", validate_payload.get("status"))
        add_check(results, "output_contract", case_id, "validate_summary_total_checks", 3, validate_payload.get("total_checks"))
        add_check(results, "output_contract", case_id, "validate_summary_pass_count", 1, validate_payload.get("pass_count"))
        add_check(results, "output_contract", case_id, "validate_summary_skip_count", 1, validate_payload.get("skip_count"))
        add_check(results, "output_contract", case_id, "validate_summary_fail_count", 1, validate_payload.get("fail_count"))
        add_check(results, "output_contract", case_id, "validate_summary_real_ticker_count", 3, validate_payload.get("real_ticker_count"))

        # XLSX contract: issue report sheet names and headers
        df_failed_summary = pd.DataFrame([{"ticker": "1102", "failed_checks": 1}])
        df_failed_module = pd.DataFrame([{"module": "portfolio", "failed_checks": 1}])
        xlsx_path = write_issue_excel_report(
            df_failed,
            df_failed_summary,
            df_failed_module,
            "20260401_000000",
            output_dir=str(temp_path),
            normalize_ticker=lambda value: str(value).zfill(4),
        )
        add_check(results, "output_contract", case_id, "issue_report_xlsx_exists", True, bool(xlsx_path) and Path(xlsx_path).exists())
        workbook = load_workbook(xlsx_path)
        expected_sheets = ["failed_only", "failed_tickers", "failed_modules"]
        add_check(results, "output_contract", case_id, "issue_report_sheet_names", expected_sheets, workbook.sheetnames)
        failed_only_headers = [cell.value for cell in workbook["failed_only"][1]]
        add_check(results, "output_contract", case_id, "issue_report_failed_only_headers", ["ticker", "module", "metric", "passed", "expected", "actual", "note"], failed_only_headers)
        failed_tickers_headers = [cell.value for cell in workbook["failed_tickers"][1]]
        add_check(results, "output_contract", case_id, "issue_report_failed_tickers_headers", ["ticker", "failed_checks"], failed_tickers_headers)
        failed_only_ticker = workbook["failed_only"]["A2"].value
        add_check(results, "output_contract", case_id, "issue_report_ticker_normalized_text", "1102", str(failed_only_ticker))

        # CSV / JSON contract: optimizer profiling outputs
        profiler = OptimizerProfileRecorder(str(temp_path), "20260401_000000", enabled=True, console_print=False)
        profiler.init_output_files()
        profiler.append_row(
            {
                "trial_number": 0,
                "objective_wall_sec": 1.2,
                "prep_wall_sec": 0.3,
                "portfolio_wall_sec": 0.4,
                "score_calc_sec": 0.05,
                "trade_count": 2,
                "trial_value": 0.91,
            }
        )
        with redirect_stdout(io.StringIO()):
            profiler.print_summary()

        profile_csv_path = Path(profiler.csv_path)
        profile_summary_path = Path(profiler.summary_path)
        add_check(results, "output_contract", case_id, "optimizer_profile_csv_exists", True, profile_csv_path.exists())
        add_check(results, "output_contract", case_id, "optimizer_profile_summary_exists", True, profile_summary_path.exists())
        fieldnames, profile_rows = _read_csv_rows(profile_csv_path)
        add_check(results, "output_contract", case_id, "optimizer_profile_csv_header", PROFILE_FIELDS, fieldnames)
        add_check(results, "output_contract", case_id, "optimizer_profile_csv_row_count", 1, len(profile_rows))
        profile_summary = json.loads(profile_summary_path.read_text(encoding="utf-8"))
        add_check(results, "output_contract", case_id, "optimizer_profile_summary_trial_count", 1, profile_summary.get("trial_count"))
        add_check(results, "output_contract", case_id, "optimizer_profile_summary_avg_keys_present", [], sorted(REQUIRED_PROFILE_SUMMARY_AVG_KEYS - set(profile_summary.get("avg", {}).keys())))
        add_check(results, "output_contract", case_id, "optimizer_profile_summary_avg_objective_wall_sec", 1.2, profile_summary.get("avg", {}).get("objective_wall_sec"))

    summary["json_contract_keys"] = sorted(REQUIRED_VALIDATE_SUMMARY_KEYS)
    summary["profile_fields"] = len(PROFILE_FIELDS)
    return results, summary


def validate_quick_gate_bare_except_guard_contract_case(_base_params):
    case_id = "QUICK_GATE_BARE_EXCEPT_GUARD_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    staging_root = run_quick_gate_module.PROJECT_ROOT / "outputs" / "local_regression" / "_staging"
    staging_root.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="quick_gate_bare_except_contract_", dir=str(staging_root)) as temp_dir:
        temp_path = Path(temp_dir)
        typed_except_file = temp_path / "typed_except_probe.py"
        typed_except_file.write_text(
            "def probe():\n"
            "    try:\n"
            "        return 1\n"
            "    except ValueError:\n"
            "        return 0\n",
            encoding="utf-8",
        )
        bare_except_file = temp_path / "bare_except_probe.py"
        bare_except_file.write_text(
            "def probe():\n"
            "    try:\n"
            "        return 1\n"
            "    except:\n"
            "        return 0\n",
            encoding="utf-8",
        )

        def _fake_compile(source, cfile=None, doraise=False, *args, **kwargs):
            return cfile or source

        with patch.object(run_quick_gate_module, "iter_python_files", return_value=[typed_except_file, bare_except_file]),              patch.object(run_quick_gate_module.py_compile, "compile", side_effect=_fake_compile):
            static_results = run_quick_gate_module.run_static_checks()

        bare_step = _find_step_payload(static_results, "bare_except_scan")
        bare_hits = [] if bare_step is None else list(bare_step.get("hits", []))

        add_check(results, "output_contract", case_id, "run_static_checks_declares_bare_except_scan", True, bare_step is not None)
        add_check(results, "output_contract", case_id, "bare_except_scan_fails_when_bare_except_exists", "FAIL", None if bare_step is None else bare_step.get("status"))
        add_check(results, "output_contract", case_id, "bare_except_scan_reports_bare_except_file", True, any("bare_except_probe.py" in hit for hit in bare_hits))
        add_check(results, "output_contract", case_id, "bare_except_scan_does_not_flag_typed_except", False, any("typed_except_probe.py" in hit for hit in bare_hits))

    return results, summary



def validate_quick_gate_output_path_guard_contract_case(_base_params):
    case_id = "QUICK_GATE_OUTPUT_PATH_GUARD_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    with tempfile.TemporaryDirectory(prefix="quick_gate_log_utils_contract_") as log_tmp_dir:
        project_root = Path(log_tmp_dir)
        outputs_root = project_root / "outputs"
        outputs_root.mkdir(parents=True, exist_ok=True)
        with patch.object(log_utils_module, "PROJECT_ROOT", str(project_root.resolve())), \
             patch.object(log_utils_module, "PROJECT_ROOT_PATH", project_root.resolve()), \
             patch.object(log_utils_module, "OUTPUTS_ROOT_PATH", outputs_root.resolve()):
            invalid_log_dir_cases = [
                ("resolve_log_dir_outputs_root_rejected", lambda: log_utils_module.resolve_log_dir("outputs")),
                ("build_timestamped_log_path_outputs_root_rejected", lambda: log_utils_module.build_timestamped_log_path("scanner_issues", log_dir="outputs", timestamp="20260405_123000")),
                ("write_issue_log_outputs_root_rejected", lambda: log_utils_module.write_issue_log("scanner_issues", ["probe"], log_dir="outputs", timestamp="20260405_123000")),
            ]
            for metric_name, probe in invalid_log_dir_cases:
                try:
                    probe()
                    ok = False
                    actual = "未拒絕 outputs/ 根目錄"
                except ValueError as exc:
                    ok = "outputs/ 根目錄" in str(exc)
                    actual = str(exc)
                except Exception as exc:
                    ok = False
                    actual = f"{type(exc).__name__}: {exc}"
                add_check(results, "output_contract", case_id, metric_name, True, ok)
                summary[metric_name] = actual

    staging_root = run_quick_gate_module.PROJECT_ROOT / "outputs" / "local_regression" / "_staging"
    staging_root.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="quick_gate_output_path_contract_", dir=str(staging_root)) as temp_dir:
        temp_path = Path(temp_dir)
        manifest = {**local_common.MANIFEST_DEFAULTS, "dataset": "reduced", "subprocess_timeout_sec": 1}
        pass_run_dir = temp_path / "pass_run"
        pass_run_dir.mkdir(parents=True, exist_ok=True)
        fail_run_dir = temp_path / "fail_run"
        fail_run_dir.mkdir(parents=True, exist_ok=True)

        def _run_probe(run_dir: Path, *, outputs_ok: bool):
            with patch.object(run_quick_gate_module, "resolve_run_dir", return_value=run_dir),                  patch.object(run_quick_gate_module, "load_manifest", return_value=manifest),                  patch.object(run_quick_gate_module, "ensure_reduced_dataset", return_value={"csv_count": 10}),                  patch.object(run_quick_gate_module, "run_static_checks", return_value=[]),                  patch.object(run_quick_gate_module, "check_output_path_contract", return_value=[run_quick_gate_module.summarize_result("output_path_contract::valid_category_resolves_under_outputs", True, detail="probe")]),                  patch.object(run_quick_gate_module, "check_outputs_root_layout", return_value=[run_quick_gate_module.summarize_result("outputs_root_layout::root_has_only_category_dirs", outputs_ok, detail="probe")]),                  patch.object(run_quick_gate_module, "check_dataset_profile_contract", return_value=[]),                  patch.object(run_quick_gate_module, "check_log_path_contract", return_value=[
                      run_quick_gate_module.summarize_result("log_path_contract::append_issue_log_outputs_root_file_rejected", True, detail="probe"),
                      run_quick_gate_module.summarize_result("log_path_contract::resolve_log_dir_outputs_root_rejected", True, detail="probe"),
                  ]),                  patch.object(run_quick_gate_module, "check_local_regression_contract", return_value=[]),                  patch.object(run_quick_gate_module, "check_help", return_value=[]),                  patch.object(run_quick_gate_module, "check_dataset_cli_errors", return_value=[]),                  patch.object(run_quick_gate_module, "check_generic_cli_errors", return_value=[]),                  patch.object(run_quick_gate_module, "check_error_paths", return_value=[]),                  patch.object(run_quick_gate_module, "check_dataset_runtime_error_paths", return_value=[]),                  patch("builtins.print"):
                rc = run_quick_gate_module.main(["tools/local_regression/run_quick_gate.py"])
            payload = json.loads((run_dir / "quick_gate_summary.json").read_text(encoding="utf-8"))
            return rc, payload

        pass_rc, pass_payload = _run_probe(pass_run_dir, outputs_ok=True)
        fail_rc, fail_payload = _run_probe(fail_run_dir, outputs_ok=False)

        pass_step_names = [step.get("name") for step in pass_payload.get("steps", [])]
        add_check(results, "output_contract", case_id, "quick_gate_main_wires_output_path_contract_guard", True, "output_path_contract::valid_category_resolves_under_outputs" in pass_step_names)
        add_check(results, "output_contract", case_id, "quick_gate_main_wires_outputs_root_layout_guard", True, "outputs_root_layout::root_has_only_category_dirs" in pass_step_names)
        add_check(results, "output_contract", case_id, "quick_gate_main_wires_log_path_contract_guard", True, "log_path_contract::append_issue_log_outputs_root_file_rejected" in pass_step_names)
        add_check(results, "output_contract", case_id, "quick_gate_main_wires_log_dir_outputs_root_guard", True, "log_path_contract::resolve_log_dir_outputs_root_rejected" in pass_step_names)
        add_check(results, "output_contract", case_id, "quick_gate_output_path_probe_pass_return_code", 0, pass_rc)
        add_check(results, "output_contract", case_id, "quick_gate_output_path_probe_pass_status", "PASS", pass_payload.get("status"))
        add_check(results, "output_contract", case_id, "quick_gate_output_path_probe_fail_return_code", 1, fail_rc)
        add_check(results, "output_contract", case_id, "quick_gate_output_path_probe_fail_status", "FAIL", fail_payload.get("status"))
        add_check(results, "output_contract", case_id, "quick_gate_output_path_probe_fail_propagates_failed_step", ["outputs_root_layout::root_has_only_category_dirs"], fail_payload.get("failed_steps"))

    return results, summary



def validate_local_regression_summary_contract_case(_base_params):
    case_id = "LOCAL_REGRESSION_SUMMARY_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    with tempfile.TemporaryDirectory(prefix="local_regression_summary_contract_") as temp_dir:
        temp_path = Path(temp_dir)
        run_dir = temp_path / "run_dir"
        run_dir.mkdir(parents=True, exist_ok=True)

        preflight_payload = {
            "status": "FAIL",
            "python_executable": "python",
            "duration_sec": 0.321,
            "failed_packages": ["coverage", "optuna"],
            "runtime_error": "",
        }
        preflight_text = run_all_module._safe_format_preflight_summary(preflight_payload)
        add_check(results, "output_contract", case_id, "preflight_summary_contract_lines", True, "status" in preflight_text and "failed_packages : coverage, optuna" in preflight_text)

        dataset_pass = run_all_module._write_dataset_prepare_summary(run_dir, {
            "status": "PASS",
            "duration_sec": 1.234,
            "dataset_dir": "data/tw_stock_data_vip_reduced",
            "source": "existing",
            "csv_count": 10,
            "csv_total_bytes": 2400,
            "csv_members_sha256": "a" * 64,
            "csv_content_sha256": "b" * 64,
            "fingerprint_algorithm": "sha256",
            "reused_existing": True,
            "extracted_files": 0,
        })
        dataset_pass_json = json.loads((run_dir / "dataset_prepare_summary.json").read_text(encoding="utf-8"))
        dataset_pass_text = (run_dir / "dataset_prepare_summary.txt").read_text(encoding="utf-8")
        add_check(results, "output_contract", case_id, "dataset_prepare_pass_required_keys", [], sorted(REQUIRED_DATASET_PREP_SUMMARY_KEYS - set(dataset_pass_json.keys())))
        add_check(results, "output_contract", case_id, "dataset_prepare_pass_text_contract", True, "dataset_dir : data/tw_stock_data_vip_reduced" in dataset_pass_text and f"csv_count   : {dataset_pass_json.get("csv_count")}" in dataset_pass_text and "members_sha : " in dataset_pass_text and "content_sha : " in dataset_pass_text)

        dataset_fail = run_all_module._write_dataset_prepare_summary(run_dir, {
            "status": "FAIL",
            "duration_sec": 0.456,
            "error_type": "RuntimeError",
            "error_message": "zip missing",
        })
        dataset_fail_text = (run_dir / "dataset_prepare_summary.txt").read_text(encoding="utf-8")
        add_check(results, "output_contract", case_id, "dataset_prepare_fail_status", "FAIL", dataset_fail.get("status"))
        add_check(results, "output_contract", case_id, "dataset_prepare_fail_text_contract", True, "error_type  : RuntimeError" in dataset_fail_text and "error_msg   : zip missing" in dataset_fail_text)

        chain_payload = {
            "status": "PASS",
            "dataset": "reduced",
            "dataset_info": {"csv_count": 10},
            "ticker_count": 10,
            "csv_count": 10,
            "duplicate_issue_count": 0,
            "skipped_ticker_count": 1,
            "detail_count": 2,
            "rows": [{"ticker": "2330"}],
            "highlights": {"blocked_by_counts": {"cash": 1}},
            "portfolio_snapshot": {"trade_rows": 2},
            "scanner_snapshot": {"candidate_count": 3},
            "rerun_consistency": {"enabled": True, "run_count": 2, "all_match": True, "runs": []},
            "failures": [],
            "peak_traced_memory_mb": 45.678,
        }
        write_json(run_dir / "chain_summary.json", chain_payload)
        write_csv(run_dir / "chain_summary.csv", [{field: "" for field in CHAIN_SUMMARY_CSV_FIELDS}], fieldnames=CHAIN_SUMMARY_CSV_FIELDS)
        chain_json = json.loads((run_dir / "chain_summary.json").read_text(encoding="utf-8"))
        chain_csv_fields, _ = _read_csv_rows(run_dir / "chain_summary.csv")
        add_check(results, "output_contract", case_id, "chain_summary_required_keys", [], sorted(REQUIRED_CHAIN_SUMMARY_KEYS - set(chain_json.keys())))
        add_check(results, "output_contract", case_id, "chain_summary_csv_header", CHAIN_SUMMARY_CSV_FIELDS, chain_csv_fields)

        ml_smoke_payload = {
            "status": "PASS",
            "dataset": "reduced",
            "dataset_info": {"csv_count": 10},
            "db_path": "outputs/ml_optimizer/demo.db",
            "db_trial_count": 1,
            "qualified_trial_count": 1,
            "best_trial_value": 1.23,
            "run_best_params_path": "models/run_best_params.json",
            "run_best_params_required": True,
            "best_params_keys": ["high_len", "atr_len"],
            "optimizer_profile_summary_path": "outputs/ml_optimizer/profile.json",
            "optimizer_profile_trial_count": 1,
            "optimizer_profile_avg_objective_wall_sec": 4.5,
            "optimizer_repro": {"enabled": True, "all_match": True},
            "failures": [],
            "peak_traced_memory_mb": 23.456,
        }
        write_json(run_dir / "ml_smoke_summary.json", ml_smoke_payload)
        ml_smoke_json = json.loads((run_dir / "ml_smoke_summary.json").read_text(encoding="utf-8"))
        add_check(results, "output_contract", case_id, "ml_smoke_summary_required_keys", [], sorted(REQUIRED_ML_SMOKE_SUMMARY_KEYS - set(ml_smoke_json.keys())))

        meta_payload = {
            "status": "PASS",
            "failures": [],
            "fail_count": 0,
            "coverage": {
                "ok": True,
                "status": "DONE",
                "totals": {"percent_covered": 52.5, "line_percent_covered": 55.0, "branch_percent_covered": 50.0, "line_min_percent": float(COVERAGE_LINE_MIN_FLOOR), "branch_min_percent": float(COVERAGE_BRANCH_MIN_FLOOR)},
                "line_percent_covered": 55.0,
                "branch_percent_covered": 50.0,
                "line_min_percent": float(COVERAGE_LINE_MIN_FLOOR),
                "branch_min_percent": float(COVERAGE_BRANCH_MIN_FLOOR),
                "missing_targets": [],
                "zero_covered_targets": [],
            },
            "checklist": {"ok": True, "status": "DONE", "partial_ids": [], "todo_ids": [], "done_ids": ["B11"], "unfinished_test_ids": []},
            "formal_entry": {
                "ok": True,
                "registry_steps": [],
                "registry_commands": [],
                "run_all_steps": [],
                "preflight_steps": [],
                "test_suite_steps": [],
            },
            "performance": {"ok": True, "skipped": False, "step_durations": {}},
            "results": [],
        }
        write_json(run_dir / "meta_quality_summary.json", meta_payload)
        meta_json = json.loads((run_dir / "meta_quality_summary.json").read_text(encoding="utf-8"))
        add_check(results, "output_contract", case_id, "meta_quality_summary_required_keys", [], sorted(REQUIRED_META_QUALITY_SUMMARY_KEYS - set(meta_json.keys())))
        add_check(results, "output_contract", case_id, "meta_quality_summary_coverage_threshold_keys", [], sorted([key for key in ["status", "line_percent_covered", "branch_percent_covered", "line_min_percent", "branch_min_percent", "missing_targets", "zero_covered_targets"] if key not in meta_json.get("coverage", {})]))
        add_check(results, "output_contract", case_id, "meta_quality_summary_checklist_status_key", True, meta_json.get("checklist", {}).get("status") == "DONE")
        add_check(results, "output_contract", case_id, "meta_quality_summary_formal_entry_required_keys", [], sorted(REQUIRED_META_QUALITY_FORMAL_ENTRY_KEYS - set(meta_json.get("formal_entry", {}).keys())))
        add_check(results, "output_contract", case_id, "meta_quality_summary_formal_entry_has_no_retired_steps_key", False, "steps" in meta_json.get("formal_entry", {}))

        selected_steps = ["quick_gate", "consistency", "chain_checks", "ml_smoke", "meta_quality"]
        completed_steps = ["quick_gate", "consistency"]
        failed_steps = ["ml_smoke"]
        not_run = run_all_module._compute_not_run_step_names(
            selected_step_names=selected_steps,
            failed_step_names=failed_steps,
            completed_script_names=completed_steps,
            include_dataset=True,
        )
        master_payload = {
            "overall_status": "FAIL",
            "dataset": "reduced",
            "dataset_info": {"csv_count": 10},
            "timestamp": "2026-04-02T00:00:00+08:00",
            "git_commit": "abc123",
            "scripts": [{"name": "quick_gate", "status": "PASS"}],
            "selected_steps": selected_steps,
            "failures": 1,
            "preflight": preflight_payload,
            "dataset_prepare": dataset_pass_json,
            "payload_failures": [],
            "not_run_step_names": not_run,
            "bundle_mode": "debug_bundle",
            "bundle_entries": run_all_module._build_bundle_entries(run_dir, [run_dir / "master_summary.json", run_dir / "preflight_summary.json"]),
            "failed_step_names": failed_steps,
            "suggested_rerun_command": "python tools/local_regression/run_all.py --only ml_smoke",
        }
        quick_gate_payload = {
            "status": "PASS",
            "dataset": "reduced",
            "dataset_info": {"csv_count": 10},
            "step_count": 3,
            "failed_count": 0,
            "failed_steps": [],
            "steps": [
                {"name": "py_compile", "status": "PASS"},
                {"name": "compileall", "status": "PASS"},
                {"name": "bare_except_scan", "status": "PASS"},
            ],
            "duration_sec": 1.23,
            "peak_traced_memory_mb": 12.34,
        }
        write_json(run_dir / "quick_gate_summary.json", quick_gate_payload)
        quick_gate_json = json.loads((run_dir / "quick_gate_summary.json").read_text(encoding="utf-8"))
        add_check(results, "output_contract", case_id, "quick_gate_summary_required_keys", [], sorted(REQUIRED_QUICK_GATE_SUMMARY_KEYS - set(quick_gate_json.keys())))

        compile_probe_calls = []

        def _fake_compile(source, cfile=None, doraise=False, *args, **kwargs):
            compile_probe_calls.append({"source": source, "cfile": cfile, "doraise": doraise})
            return cfile or source

        with patch.object(run_quick_gate_module, "iter_python_files", return_value=[run_quick_gate_module.PROJECT_ROOT / "apps" / "test_suite.py"]), \
             patch.object(run_quick_gate_module.py_compile, "compile", side_effect=_fake_compile):
            static_results = run_quick_gate_module.run_static_checks()

        compileall_result = next(item for item in static_results if item.get("name") == "compileall")
        compile_probe_path = Path(str(compile_probe_calls[0]["cfile"])) if compile_probe_calls else None
        compile_probe_resolved = str(compile_probe_path.resolve()) if compile_probe_path is not None else ""
        project_root_resolved = str(run_quick_gate_module.PROJECT_ROOT.resolve())
        add_check(results, "output_contract", case_id, "quick_gate_compileall_temp_cfile_present", True, bool(compile_probe_calls and compile_probe_calls[0].get("cfile")))
        add_check(results, "output_contract", case_id, "quick_gate_compileall_temp_cfile_outside_project_root", True, bool(compile_probe_resolved) and not compile_probe_resolved.startswith(project_root_resolved))
        add_check(results, "output_contract", case_id, "quick_gate_compileall_result_stays_pass_with_temp_cfile", "PASS", compileall_result.get("status"))

        runtime_run_dir = temp_path / "quick_gate_runtime_failure"
        runtime_run_dir.mkdir(parents=True, exist_ok=True)
        runtime_stdout = io.StringIO()
        with patch.object(run_quick_gate_module, "resolve_run_dir", return_value=runtime_run_dir):
            with patch.object(run_quick_gate_module, "load_manifest", return_value={**local_common.MANIFEST_DEFAULTS, "dataset": "reduced"}):
                with patch.object(run_quick_gate_module, "ensure_reduced_dataset", return_value={"csv_count": 10}):
                    with patch.object(run_quick_gate_module, "run_static_checks", side_effect=RuntimeError("synthetic quick gate crash")):
                        with patch("builtins.print"):
                            with redirect_stdout(runtime_stdout):
                                runtime_rc = run_quick_gate_module.main(["tools/local_regression/run_quick_gate.py"])

        runtime_json = json.loads((runtime_run_dir / "quick_gate_summary.json").read_text(encoding="utf-8"))
        runtime_console = (runtime_run_dir / "quick_gate_console.log").read_text(encoding="utf-8")
        runtime_stdout_text = runtime_stdout.getvalue()
        add_check(results, "output_contract", case_id, "quick_gate_runtime_failure_return_code", 1, runtime_rc)
        add_check(results, "output_contract", case_id, "quick_gate_runtime_failure_status", "FAIL", runtime_json.get("status"))
        add_check(results, "output_contract", case_id, "quick_gate_runtime_failure_marks_runtime_step", ["__runtime__"], runtime_json.get("failed_steps"))
        add_check(results, "output_contract", case_id, "quick_gate_runtime_failure_includes_error_fields", True, runtime_json.get("error_type") == "RuntimeError" and "synthetic quick gate crash" in runtime_json.get("runtime_error", ""))
        add_check(results, "output_contract", case_id, "quick_gate_runtime_failure_console_has_runtime_error", True, "synthetic quick gate crash" in runtime_console)
        add_check(results, "output_contract", case_id, "quick_gate_runtime_failure_does_not_leak_runtime_payload_to_parent_stdout", "", runtime_stdout_text)

        write_json(run_dir / "preflight_summary.json", preflight_payload)
        write_json(run_dir / "master_summary.json", master_payload)
        master_json = json.loads((run_dir / "master_summary.json").read_text(encoding="utf-8"))
        add_check(results, "output_contract", case_id, "master_summary_required_keys", [], sorted(REQUIRED_MASTER_SUMMARY_KEYS - set(master_json.keys())))
        add_check(results, "output_contract", case_id, "master_summary_not_run_steps_contract", ["chain_checks", "ml_smoke", "meta_quality"], master_json.get("not_run_step_names"))
        add_check(results, "output_contract", case_id, "master_summary_bundle_entries_posix", True, all("\\" not in item for item in master_json.get("bundle_entries", [])))

    summary["dataset_prepare_pass_status"] = dataset_pass_json.get("status")
    summary["master_selected_step_count"] = len(selected_steps)
    return results, summary




def _add_preflight_early_failure_dataset_contract_checks(results, case_id, *, work_dir: Path):
    early_failure_manifest = dict(run_all_module.MANIFEST_DEFAULTS)
    early_failure_manifest["dataset"] = "reduced"
    early_failure_manifest["bundle_name"] = "preflight_failed_bundle.zip"
    early_failure_dir = work_dir / "early_failure"
    early_failure_dir.mkdir(parents=True, exist_ok=True)
    early_preflight_payload = {
        "status": "FAIL",
        "python_executable": "python",
        "python_version": "3.x",
        "requirements_path": "requirements.txt",
        "checked_packages": ["coverage"],
        "failed_packages": ["coverage"],
        "checks": [],
        "duration_sec": 0.11,
    }
    write_json(early_failure_dir / "preflight_summary.json", early_preflight_payload)
    write_text(early_failure_dir / "preflight_summary.txt", "status : FAIL\n")
    with patch.object(run_all_module, "build_bundle_zip", return_value=early_failure_dir / "preflight_failed_bundle.zip"),          patch.object(run_all_module, "archive_bundle_history", side_effect=lambda path: path),          patch.object(run_all_module, "publish_root_bundle_copy", side_effect=lambda path: path),          patch.object(run_all_module, "_apply_output_retention", return_value={"removed_count": 0, "removed_bytes": 0, "removed_entries": []}),          patch.object(run_all_module, "resolve_git_commit", return_value="deadbeef"):
        early_result = run_all_module._finalize_early_failure(
            run_dir=early_failure_dir,
            manifest=early_failure_manifest,
            selected_step_names=["quick_gate", "consistency", "chain_checks", "ml_smoke", "meta_quality"],
            major_index=1,
            major_total=6,
            bundle_mode="preflight_failed",
            failed_step_names=["preflight"],
            preflight_payload=early_preflight_payload,
            include_dataset=True,
        )
    early_master = json.loads((early_failure_dir / "master_summary.json").read_text(encoding="utf-8"))
    expected_not_run = ["dataset_prepare", "quick_gate", "consistency", "chain_checks", "ml_smoke", "meta_quality"]
    add_check(results, "output_contract", case_id, "preflight_failed_early_summary_marks_dataset_prepare_not_run", expected_not_run, early_master.get("not_run_step_names"))
    add_check(results, "output_contract", case_id, "preflight_failed_early_result_marks_dataset_prepare_not_run", expected_not_run, early_result.get("not_run_step_names"))
    add_check(results, "output_contract", case_id, "preflight_failed_early_master_summary_required_keys", [], sorted(REQUIRED_MASTER_SUMMARY_KEYS - set(early_master.keys())))
    add_check(results, "output_contract", case_id, "preflight_failed_early_master_summary_dataset_prepare_status", "NOT_RUN", early_master.get("dataset_prepare", {}).get("status"))
    add_check(results, "output_contract", case_id, "preflight_failed_early_master_summary_dataset_prepare_blocked_by", "preflight", early_master.get("dataset_prepare", {}).get("blocked_by"))
    preflight_payload_failures = early_master.get("payload_failures") or []
    add_check(results, "output_contract", case_id, "preflight_failed_early_master_summary_payload_failure_names", ["preflight"], [item.get("name") for item in preflight_payload_failures])
    preflight_failure_reasons = preflight_payload_failures[0].get("failure_reasons", []) if preflight_payload_failures else []
    add_check(results, "output_contract", case_id, "preflight_failed_early_master_summary_payload_failure_has_status", True, "reported_status=FAIL" in preflight_failure_reasons)
    add_check(results, "output_contract", case_id, "preflight_failed_early_master_summary_payload_failure_has_failed_packages", True, "failed_packages=coverage" in preflight_failure_reasons)

    dataset_failure_dir = work_dir / "dataset_failure"
    dataset_failure_dir.mkdir(parents=True, exist_ok=True)
    passing_preflight_payload = {
        "status": "PASS",
        "python_executable": "python",
        "python_version": "3.x",
        "requirements_path": "requirements.txt",
        "checked_packages": ["coverage"],
        "failed_packages": [],
        "checks": [],
        "duration_sec": 0.05,
    }
    dataset_prepare_payload = {
        "status": "FAIL",
        "duration_sec": 0.23,
        "error_type": "RuntimeError",
        "error_message": "synthetic dataset prepare failure",
    }
    write_json(dataset_failure_dir / "preflight_summary.json", passing_preflight_payload)
    write_text(dataset_failure_dir / "preflight_summary.txt", "status : PASS\n")
    write_json(dataset_failure_dir / "dataset_prepare_summary.json", dataset_prepare_payload)
    write_text(dataset_failure_dir / "dataset_prepare_summary.txt", "status : FAIL\n")
    with patch.object(run_all_module, "build_bundle_zip", return_value=dataset_failure_dir / "dataset_prepare_failed_bundle.zip"),          patch.object(run_all_module, "archive_bundle_history", side_effect=lambda path: path),          patch.object(run_all_module, "publish_root_bundle_copy", side_effect=lambda path: path),          patch.object(run_all_module, "_apply_output_retention", return_value={"removed_count": 0, "removed_bytes": 0, "removed_entries": []}),          patch.object(run_all_module, "resolve_git_commit", return_value="deadbeef"):
        dataset_result = run_all_module._finalize_early_failure(
            run_dir=dataset_failure_dir,
            manifest=early_failure_manifest,
            selected_step_names=["quick_gate", "consistency", "chain_checks", "ml_smoke", "meta_quality"],
            major_index=2,
            major_total=6,
            bundle_mode="dataset_prepare_failed",
            failed_step_names=["dataset_prepare"],
            preflight_payload=passing_preflight_payload,
            include_dataset=True,
            dataset_prepare_payload=dataset_prepare_payload,
        )
    dataset_master = json.loads((dataset_failure_dir / "master_summary.json").read_text(encoding="utf-8"))
    add_check(results, "output_contract", case_id, "dataset_prepare_failed_early_result_marks_scripts_not_run", ["quick_gate", "consistency", "chain_checks", "ml_smoke", "meta_quality"], dataset_result.get("not_run_step_names"))
    dataset_payload_failures = dataset_master.get("payload_failures") or []
    add_check(results, "output_contract", case_id, "dataset_prepare_failed_early_master_summary_payload_failure_names", ["dataset_prepare"], [item.get("name") for item in dataset_payload_failures])
    dataset_failure_reasons = dataset_payload_failures[0].get("failure_reasons", []) if dataset_payload_failures else []
    add_check(results, "output_contract", case_id, "dataset_prepare_failed_early_master_summary_payload_failure_has_status", True, "reported_status=FAIL" in dataset_failure_reasons)
    add_check(results, "output_contract", case_id, "dataset_prepare_failed_early_master_summary_payload_failure_has_error_message", True, "error_message=synthetic dataset prepare failure" in dataset_failure_reasons)
    add_check(results, "output_contract", case_id, "dataset_prepare_failed_early_master_summary_payload_failure_not_marked_unreadable", False, "summary_unreadable" in dataset_failure_reasons)


def validate_run_all_preflight_early_failure_dataset_contract_case(_base_params):
    case_id = "RUN_ALL_PREFLIGHT_EARLY_FAILURE_DATASET_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    with tempfile.TemporaryDirectory(prefix="run_all_preflight_fail_contract_") as temp_dir:
        work_dir = Path(temp_dir)
        _add_preflight_early_failure_dataset_contract_checks(results, case_id, work_dir=work_dir)

    summary["checked_contract"] = "preflight_failed_dataset_prepare_not_run"
    return results, summary


def validate_run_all_manifest_failure_master_summary_contract_case(_base_params):
    case_id = "RUN_ALL_MANIFEST_FAILURE_MASTER_SUMMARY_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    with tempfile.TemporaryDirectory(prefix="run_all_manifest_fail_contract_") as temp_dir:
        run_dir = Path(temp_dir) / "manifest_failure"
        run_dir.mkdir(parents=True, exist_ok=True)
        with patch.object(run_all_module, "build_bundle_zip", return_value=run_dir / "manifest_failed_bundle.zip"),              patch.object(run_all_module, "archive_bundle_history", side_effect=lambda path: path),              patch.object(run_all_module, "publish_root_bundle_copy", side_effect=lambda path: path),              patch.object(run_all_module, "_apply_output_retention", return_value={"removed_count": 0, "removed_bytes": 0, "removed_entries": []}),              patch.object(run_all_module, "resolve_git_commit", return_value="deadbeef"):
            run_all_module._write_manifest_failure_bundle(
                run_dir=run_dir,
                selected_step_names=["quick_gate", "consistency", "chain_checks", "ml_smoke", "meta_quality"],
                include_dataset=True,
                manifest_error=ValueError("synthetic manifest failure"),
            )

        master_json = json.loads((run_dir / "master_summary.json").read_text(encoding="utf-8"))
        expected_not_run = ["preflight", "dataset_prepare", "quick_gate", "consistency", "chain_checks", "ml_smoke", "meta_quality"]
        add_check(results, "output_contract", case_id, "manifest_failed_master_summary_required_keys", [], sorted(REQUIRED_MASTER_SUMMARY_KEYS - set(master_json.keys())))
        add_check(results, "output_contract", case_id, "manifest_failed_master_summary_not_run_steps", expected_not_run, master_json.get("not_run_step_names"))
        add_check(results, "output_contract", case_id, "manifest_failed_master_summary_preflight_status", "NOT_RUN", master_json.get("preflight", {}).get("status"))
        add_check(results, "output_contract", case_id, "manifest_failed_master_summary_preflight_blocked_by", "manifest", master_json.get("preflight", {}).get("blocked_by"))
        add_check(results, "output_contract", case_id, "manifest_failed_master_summary_dataset_prepare_status", "NOT_RUN", master_json.get("dataset_prepare", {}).get("status"))
        add_check(results, "output_contract", case_id, "manifest_failed_master_summary_dataset_prepare_blocked_by", "manifest", master_json.get("dataset_prepare", {}).get("blocked_by"))
        add_check(results, "output_contract", case_id, "manifest_failed_master_summary_payload_failures_empty", [], master_json.get("payload_failures"))

    summary["checked_contract"] = "manifest_failed_master_summary_schema"
    return results, summary

def validate_artifact_lifecycle_contract_case(_base_params):
    case_id = "ARTIFACT_LIFECYCLE_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    with tempfile.TemporaryDirectory(prefix="artifact_lifecycle_") as temp_dir:
        temp_path = Path(temp_dir)
        project_root = temp_path / "project_root"
        run_dir = temp_path / "run_dir"
        output_root = project_root / "outputs" / "local_regression"
        run_dir.mkdir(parents=True, exist_ok=True)
        output_root.mkdir(parents=True, exist_ok=True)
        (project_root / "outputs" / "summary_tools").mkdir(parents=True, exist_ok=True)
        probe_master = run_dir / "master_summary.json"
        probe_console = run_dir / "console_tail.txt"
        probe_preflight = run_dir / "preflight_summary.json"
        probe_preflight_txt = run_dir / "preflight_summary.txt"
        probe_dataset = run_dir / "dataset_prepare_summary.json"
        probe_dataset_txt = run_dir / "dataset_prepare_summary.txt"
        probe_quick = run_dir / "quick_gate_summary.json"
        probe_validate = run_dir / "validate_consistency_summary.json"
        probe_chain_json = run_dir / "chain_summary.json"
        probe_chain_csv = run_dir / "chain_summary.csv"
        probe_ml = run_dir / "ml_smoke_summary.json"
        nested_fail = run_dir / "nested" / "debug.log"
        nested_fail.parent.mkdir(parents=True, exist_ok=True)

        for path, content in [
            (probe_master, '{"status":"PASS"}\n'),
            (probe_console, 'tail\n'),
            (probe_preflight, '{"status":"PASS"}\n'),
            (probe_preflight_txt, 'preflight ok\n'),
            (probe_dataset, '{"status":"PASS"}\n'),
            (probe_dataset_txt, 'dataset ok\n'),
            (probe_quick, '{"status":"PASS"}\n'),
            (probe_validate, '{"status":"PASS"}\n'),
            (probe_chain_json, '{"status":"PASS"}\n'),
            (probe_chain_csv, 'ticker,status\n1101,PASS\n'),
            (probe_ml, '{"status":"PASS"}\n'),
            (nested_fail, 'debug\n'),
        ]:
            path.write_text(content, encoding="utf-8")

        old_project_root = local_common.PROJECT_ROOT
        old_output_root = local_common.OUTPUT_ROOT
        local_common.PROJECT_ROOT = project_root
        local_common.OUTPUT_ROOT = output_root
        try:
            preferred_pass_paths = local_common.select_bundle_paths(run_dir, overall_ok=True)
            preferred_pass_names = [path.name for path in preferred_pass_paths]
            add_check(
                results,
                "artifact_contract",
                case_id,
                "select_bundle_paths_pass_minimum_set",
                [
                    "master_summary.json",
                    "preflight_summary.json",
                    "preflight_summary.txt",
                    "dataset_prepare_summary.json",
                    "dataset_prepare_summary.txt",
                    "quick_gate_summary.json",
                    "validate_consistency_summary.json",
                    "chain_summary.json",
                    "chain_summary.csv",
                    "ml_smoke_summary.json",
                    "console_tail.txt",
                ],
                preferred_pass_names,
            )
            preferred_fail_paths = local_common.select_bundle_paths(run_dir, overall_ok=False)
            preferred_fail_rel = sorted(str(path.relative_to(run_dir)).replace("\\", "/") for path in preferred_fail_paths)
            add_check(results, "artifact_contract", case_id, "select_bundle_paths_fail_includes_nested_debug_files", True, "nested/debug.log" in preferred_fail_rel)

            manifest_payload = local_common.build_artifacts_manifest(run_dir)
            add_check(results, "artifact_contract", case_id, "artifacts_manifest_counts_all_files", 12, manifest_payload.get("artifact_count"))
            manifest_paths = [item.get("relative_path") for item in manifest_payload.get("artifacts", [])]
            add_check(results, "artifact_contract", case_id, "artifacts_manifest_uses_posix_relative_paths", True, all(isinstance(item, str) and "\\" not in item and not item.startswith("/") for item in manifest_paths))
            add_check(results, "artifact_contract", case_id, "artifacts_manifest_includes_nested_debug_file", True, "nested/debug.log" in manifest_paths)

            bundle_path = local_common.build_bundle_zip(run_dir, "to_chatgpt_bundle.zip", include_paths=[probe_master, probe_console])
            add_check(results, "artifact_contract", case_id, "bundle_zip_exists", True, bundle_path.exists())
            with zipfile.ZipFile(bundle_path, "r") as zf:
                zip_members = sorted(zf.namelist())
            add_check(results, "artifact_contract", case_id, "bundle_zip_member_list", ["console_tail.txt", "master_summary.json"], zip_members)

            archived_bundle = local_common.archive_bundle_history(bundle_path)
            add_check(results, "artifact_contract", case_id, "archive_bundle_exists", True, archived_bundle.exists())
            add_check(results, "artifact_contract", case_id, "archive_bundle_removed_from_run_dir", False, bundle_path.exists())
            add_check(results, "artifact_contract", case_id, "archive_bundle_stays_under_output_root", str(output_root.resolve()), str(archived_bundle.parent.resolve()))

            stale_root_bundle = project_root / "to_chatgpt_bundle_old.zip"
            stale_root_bundle.write_text("stale\n", encoding="utf-8")
            root_copy_1 = local_common.publish_root_bundle_copy(archived_bundle)
            add_check(results, "artifact_contract", case_id, "root_bundle_copy_exists", True, root_copy_1.exists())
            root_bundles_after_first = sorted(path.name for path in project_root.glob("to_chatgpt_bundle*.zip"))
            add_check(results, "artifact_contract", case_id, "root_bundle_old_copy_removed_on_publish", [root_copy_1.name], root_bundles_after_first)

            probe_console.write_text("tail updated\n", encoding="utf-8")
            rebundle_path = local_common.build_bundle_zip(run_dir, "to_chatgpt_bundle.zip", include_paths=[probe_master, probe_console])
            with zipfile.ZipFile(rebundle_path, "r") as zf:
                rebundle_members = sorted(zf.namelist())
                rebundle_console = zf.read("console_tail.txt").decode("utf-8")
            rebundle_console_normalized = rebundle_console.replace("\r\n", "\n")
            add_check(results, "artifact_contract", case_id, "rerun_bundle_overwrites_run_dir_zip", ["console_tail.txt", "master_summary.json"], rebundle_members)
            add_check(results, "artifact_contract", case_id, "rerun_bundle_contains_latest_console_tail", "tail updated\n", rebundle_console_normalized)

            archived_bundle_2 = local_common.archive_bundle_history(rebundle_path)
            root_copy_2 = local_common.publish_root_bundle_copy(archived_bundle_2)
            root_bundles_after_second = sorted(path.name for path in project_root.glob("to_chatgpt_bundle*.zip"))
            add_check(results, "artifact_contract", case_id, "root_bundle_only_latest_copy_kept", [root_copy_2.name], root_bundles_after_second)
            with zipfile.ZipFile(root_copy_2, "r") as zf:
                root_console = zf.read("console_tail.txt").decode("utf-8")
            root_console_normalized = root_console.replace("\r\n", "\n")
            add_check(results, "artifact_contract", case_id, "root_bundle_copy_matches_latest_archive_contents", "tail updated\n", root_console_normalized)

            old_archive = output_root / "to_chatgpt_bundle_20260101_000000_old.zip"
            old_archive.write_text("old\n", encoding="utf-8")
            retention_now = local_common.taipei_now()
            old_ts = retention_now.timestamp() - 40 * 86400
            os.utime(old_archive, (old_ts, old_ts))
            new_ts = retention_now.timestamp()
            os.utime(archived_bundle, (new_ts, new_ts))
            os.utime(archived_bundle_2, (new_ts, new_ts))
            retention = apply_retention_rules(
                [
                    RetentionRule(
                        name="bundle_history",
                        target_dir=output_root,
                        patterns=["to_chatgpt_bundle*.zip"],
                        keep_last_n=2,
                        max_age_days=30,
                    )
                ],
                now=retention_now,
            )
            add_check(results, "artifact_contract", case_id, "retention_removed_old_archive_count", 1, retention.get("removed_count"))
            add_check(results, "artifact_contract", case_id, "retention_removed_old_archive_missing", False, old_archive.exists())
            remaining_archives = sorted(path.name for path in output_root.glob("to_chatgpt_bundle*.zip"))
            add_check(results, "artifact_contract", case_id, "retention_keeps_recent_archives", sorted([archived_bundle.name, archived_bundle_2.name]), remaining_archives)
        finally:
            local_common.PROJECT_ROOT = old_project_root
            local_common.OUTPUT_ROOT = old_output_root

    summary["archive_retention_checked"] = True
    summary["root_copy_overwrite_checked"] = True
    summary["pass_fail_bundle_selection_checked"] = True
    summary["artifacts_manifest_checked"] = True
    return results, summary


def validate_dataset_fingerprint_contract_case(_base_params):
    case_id = "DATASET_FINGERPRINT_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    dataset_info = local_common.ensure_reduced_dataset()
    add_check(results, "output_contract", case_id, "dataset_fingerprint_has_required_keys", [], sorted([
        key for key in ["csv_count", "csv_total_bytes", "csv_members_sha256", "csv_content_sha256", "fingerprint_algorithm"]
        if key not in dataset_info
    ]))
    add_check(results, "output_contract", case_id, "dataset_fingerprint_algorithm_is_sha256", "sha256", dataset_info.get("fingerprint_algorithm"))
    add_check(results, "output_contract", case_id, "dataset_member_hash_length_is_64", 64, len(str(dataset_info.get("csv_members_sha256", ""))))
    add_check(results, "output_contract", case_id, "dataset_content_hash_length_is_64", 64, len(str(dataset_info.get("csv_content_sha256", ""))))
    add_check(results, "output_contract", case_id, "dataset_total_bytes_positive", True, int(dataset_info.get("csv_total_bytes", 0)) > 0)

    with tempfile.TemporaryDirectory(prefix="dataset_fingerprint_contract_") as temp_dir:
        run_dir = Path(temp_dir) / "run_dir"
        run_dir.mkdir(parents=True, exist_ok=True)
        written = run_all_module._write_dataset_prepare_summary(run_dir, {
            "status": "PASS",
            "duration_sec": 0.123,
            **dataset_info,
        })
        summary_json = json.loads((run_dir / "dataset_prepare_summary.json").read_text(encoding="utf-8"))
        summary_txt = (run_dir / "dataset_prepare_summary.txt").read_text(encoding="utf-8")

    add_check(results, "output_contract", case_id, "dataset_prepare_summary_preserves_member_hash", dataset_info.get("csv_members_sha256"), summary_json.get("csv_members_sha256"))
    add_check(results, "output_contract", case_id, "dataset_prepare_summary_preserves_content_hash", dataset_info.get("csv_content_sha256"), summary_json.get("csv_content_sha256"))
    add_check(results, "output_contract", case_id, "dataset_prepare_summary_preserves_total_bytes", dataset_info.get("csv_total_bytes"), summary_json.get("csv_total_bytes"))
    add_check(results, "output_contract", case_id, "dataset_prepare_summary_text_displays_hashes", True, "members_sha : " in summary_txt and "content_sha : " in summary_txt)

    summary["csv_count"] = int(dataset_info.get("csv_count", 0))
    summary["csv_total_bytes"] = int(dataset_info.get("csv_total_bytes", 0))
    summary["member_hash_prefix"] = str(dataset_info.get("csv_members_sha256", ""))[:12]
    return results, summary


def validate_reduced_dataset_dynamic_contract_case(_base_params):
    case_id = "REDUCED_DATASET_DYNAMIC_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    dataset_info = local_common.ensure_reduced_dataset()
    actual_members = local_common._current_dataset_csv_members()
    add_check(results, "output_contract", case_id, "reduced_dataset_live_csv_count_matches_members", len(actual_members), int(dataset_info.get("csv_count", 0)))
    add_check(results, "output_contract", case_id, "reduced_dataset_live_total_bytes_positive", True, int(dataset_info.get("csv_total_bytes", 0)) > 0)

    with tempfile.TemporaryDirectory(prefix="reduced_dataset_dynamic_contract_") as temp_dir:
        temp_dataset_dir = Path(temp_dir) / "tw_stock_data_vip_reduced"
        temp_dataset_dir.mkdir(parents=True, exist_ok=True)
        synthetic_members = {
            "1101.csv": "Date,Open,High,Low,Close,Volume\n2024-01-02,10,11,9,10.5,1000\n",
            "2330.csv": "Date,Open,High,Low,Close,Volume\n2024-01-02,20,21,19,20.5,2000\n",
            "ETF/00631L.csv": "Date,Open,High,Low,Close,Volume\n2024-01-02,30,31,29,30.5,3000\n",
        }
        for rel_path, payload in synthetic_members.items():
            file_path = temp_dataset_dir / rel_path
            file_path.parent.mkdir(parents=True, exist_ok=True)
            file_path.write_text(payload, encoding="utf-8")

        original_dataset_dir = local_common.REDUCED_DATASET_DIR
        try:
            local_common.REDUCED_DATASET_DIR = temp_dataset_dir
            dynamic_members = local_common._current_dataset_csv_members()
            dynamic_info = local_common.ensure_reduced_dataset()
            expected_dynamic_fingerprint = local_common._build_dataset_fingerprint(temp_dataset_dir, dynamic_members)
        finally:
            local_common.REDUCED_DATASET_DIR = original_dataset_dir

    add_check(results, "output_contract", case_id, "reduced_dataset_contract_accepts_different_member_set", sorted(synthetic_members.keys()), dynamic_members)
    add_check(results, "output_contract", case_id, "reduced_dataset_contract_uses_runtime_member_count", len(synthetic_members), int(dynamic_info.get("csv_count", 0)))
    add_check(results, "output_contract", case_id, "reduced_dataset_contract_uses_runtime_fingerprint", expected_dynamic_fingerprint, {
        "csv_total_bytes": dynamic_info.get("csv_total_bytes"),
        "csv_members_sha256": dynamic_info.get("csv_members_sha256"),
        "csv_content_sha256": dynamic_info.get("csv_content_sha256"),
        "fingerprint_algorithm": dynamic_info.get("fingerprint_algorithm"),
    })

    summary["live_csv_count"] = int(dataset_info.get("csv_count", 0))
    summary["dynamic_csv_count"] = int(dynamic_info.get("csv_count", 0))
    summary["dynamic_members"] = dynamic_members
    return results, summary


def validate_run_all_dataset_prepare_pass_main_contract_case(_base_params):
    case_id = "RUN_ALL_DATASET_PREPARE_PASS_MAIN_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    manifest = dict(local_common.MANIFEST_DEFAULTS)
    manifest["bundle_name"] = "to_chatgpt_bundle.zip"
    fixed_now = local_common.taipei_now()

    with tempfile.TemporaryDirectory(prefix="run_all_dataset_prepare_pass_") as temp_dir:
        run_dir = Path(temp_dir) / "run_dir"
        run_dir.mkdir(parents=True, exist_ok=True)
        bundle_path = run_dir / manifest["bundle_name"]
        bundle_path.write_bytes(b"synthetic-bundle")

        dataset_info = {
            "dataset_dir": str(local_common.REDUCED_DATASET_DIR),
            "source": "repo_data_dir",
            "csv_count": 10,
            "reused_existing": True,
            "csv_total_bytes": 4096,
            "csv_members_sha256": "a" * 64,
            "csv_content_sha256": "b" * 64,
            "fingerprint_algorithm": "sha256",
            "extracted_files": [],
        }

        def _fake_run_preflight(run_dir_arg, *, selected_step_names):
            payload = {
                "status": "PASS",
                "python_executable": sys.executable,
                "duration_sec": 0.01,
                "failed_packages": [],
            }
            write_json(run_dir_arg / "preflight_summary.json", payload)
            write_text(run_dir_arg / "preflight_summary.txt", "status : PASS\n")
            return payload

        summary_name_by_script = {script: summary_name for _name, script, summary_name in run_all_module.SCRIPT_ORDER}

        def _fake_run_script(*, name, relative_script, timeout_sec, env, log_path, progress_callback, major_index, major_total, execution_mode="serial"):
            payload = {
                "status": "PASS",
                "failures": [],
                "failed_steps": [],
                "failed_count": 0,
                "fail_count": 0,
            }
            summary_name = summary_name_by_script[relative_script]
            write_json(run_dir / summary_name, payload)
            log_path.write_text(f"{name}: PASS\n", encoding="utf-8")
            return {
                "returncode": 0,
                "duration_sec": 0.02,
                "timed_out": False,
                "error_type": "",
                "error_message": "",
                "stdout": "",
                "stderr": "",
            }

        with patch.object(run_all_module, "create_staging_run_dir", return_value=run_dir), \
             patch.object(run_all_module, "load_manifest", return_value=manifest), \
             patch.object(run_all_module, "_run_preflight", side_effect=_fake_run_preflight), \
             patch.object(run_all_module, "ensure_reduced_dataset", return_value=dataset_info), \
             patch.object(run_all_module, "build_shared_prep_cache", return_value={"prepared_count": 0, "skipped_count": 0, "duplicate_issue_count": 0}), \
             patch.object(run_all_module, "_run_script", side_effect=_fake_run_script), \
             patch.object(run_all_module, "build_bundle_zip", return_value=bundle_path), \
             patch.object(run_all_module, "archive_bundle_history", side_effect=lambda path: path), \
             patch.object(run_all_module, "publish_root_bundle_copy", side_effect=lambda path: path), \
             patch.object(run_all_module, "_apply_output_retention", return_value={"removed_count": 0, "removed_bytes": 0, "removed_entries": []}), \
             patch.object(run_all_module, "cleanup_staging_dir", return_value=None), \
             patch.object(run_all_module, "resolve_git_commit", return_value="deadbeef"), \
             patch.object(run_all_module, "taipei_now", return_value=fixed_now), \
             patch.object(run_all_module, "gather_recent_console_tail", return_value="tail text"):
            stdout = io.StringIO()
            with redirect_stdout(stdout):
                rc = run_all_module.main(["run_all.py"])

        master_summary = json.loads((run_dir / "master_summary.json").read_text(encoding="utf-8"))
        dataset_prepare_summary = json.loads((run_dir / "dataset_prepare_summary.json").read_text(encoding="utf-8"))

    add_check(results, "output_contract", case_id, "run_all_main_dataset_prepare_pass_return_code", 0, rc)
    add_check(results, "output_contract", case_id, "run_all_main_dataset_prepare_pass_overall_status", "PASS", master_summary.get("overall_status"))
    add_check(results, "output_contract", case_id, "run_all_main_dataset_prepare_payload_status", "PASS", master_summary.get("dataset_prepare", {}).get("status"))
    add_check(results, "output_contract", case_id, "run_all_main_dataset_prepare_not_run_steps_empty", [], master_summary.get("not_run_step_names", []))
    add_check(results, "output_contract", case_id, "run_all_main_dataset_prepare_selected_steps", list(run_all_module.STEP_NAMES), master_summary.get("selected_steps"))
    for key in local_common.DATASET_INFO_KEYS:
        if key in dataset_info:
            add_check(results, "output_contract", case_id, f"run_all_main_dataset_prepare_preserves_{key}", dataset_info[key], dataset_prepare_summary.get(key))
            add_check(results, "output_contract", case_id, f"run_all_main_master_summary_dataset_info_preserves_{key}", dataset_info[key], master_summary.get("dataset_info", {}).get(key))

    summary["dataset_keys_checked"] = len([key for key in local_common.DATASET_INFO_KEYS if key in dataset_info])
    summary["selected_step_count"] = len(master_summary.get("selected_steps", []))
    return results, summary


def validate_atomic_write_contract_case(_base_params):
    case_id = "ATOMIC_WRITE_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    with tempfile.TemporaryDirectory(prefix="atomic_write_contract_") as temp_dir:
        run_dir = Path(temp_dir)
        json_path = run_dir / "atomic_summary.json"
        text_path = run_dir / "atomic_summary.txt"

        local_common.write_json(json_path, {"status": "original", "count": 1})
        local_common.write_text(text_path, "original-text\n")
        original_json = json_path.read_text(encoding="utf-8")
        original_text = text_path.read_text(encoding="utf-8")

        with patch.object(local_common.os, "replace", side_effect=OSError("simulated atomic replace failure")):
            try:
                local_common.write_json(json_path, {"status": "new", "count": 2})
            except OSError:
                pass
            else:
                raise AssertionError("write_json should raise when os.replace fails")

            try:
                local_common.write_text(text_path, "new-text\n")
            except OSError:
                pass
            else:
                raise AssertionError("write_text should raise when os.replace fails")

        remaining_temp_files = sorted(path.name for path in run_dir.glob(".*.tmp"))
        add_check(results, "output_contract", case_id, "atomic_write_json_preserves_previous_contents_on_replace_failure", original_json, json_path.read_text(encoding="utf-8"))
        add_check(results, "output_contract", case_id, "atomic_write_text_preserves_previous_contents_on_replace_failure", original_text, text_path.read_text(encoding="utf-8"))
        add_check(results, "output_contract", case_id, "atomic_write_cleanup_removes_temp_files_after_failure", [], remaining_temp_files)

        local_common.write_json(json_path, {"status": "final", "count": 3})
        local_common.write_text(text_path, "final-text\n")
        final_json = json.loads(json_path.read_text(encoding="utf-8"))
        final_text = text_path.read_text(encoding="utf-8")

    add_check(results, "output_contract", case_id, "atomic_write_json_success_after_failure", "final", final_json.get("status"))
    add_check(results, "output_contract", case_id, "atomic_write_text_success_after_failure", "final-text\n", final_text)

    summary["cleanup_checked"] = True
    return results, summary


def validate_atomic_write_retry_contract_case(_base_params):
    case_id = "ATOMIC_WRITE_RETRY_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    with tempfile.TemporaryDirectory(prefix="atomic_write_retry_contract_") as temp_dir:
        run_dir = Path(temp_dir)
        json_path = run_dir / "atomic_retry_summary.json"
        local_common.write_json(json_path, {"status": "original", "count": 1})

        original_replace = local_common.os.replace
        replace_attempts = {"count": 0}

        def flaky_replace(src, dst):
            replace_attempts["count"] += 1
            if replace_attempts["count"] < 3:
                raise PermissionError("simulated transient share violation")
            return original_replace(src, dst)

        with patch.object(local_common.os, "replace", side_effect=flaky_replace),              patch.object(local_common.time, "sleep", side_effect=lambda *_args, **_kwargs: None):
            local_common.write_json(json_path, {"status": "retried", "count": 2})

        payload = json.loads(json_path.read_text(encoding="utf-8"))
        remaining_temp_files = sorted(path.name for path in run_dir.glob(".*.tmp"))

    add_check(results, "output_contract", case_id, "atomic_write_retry_eventually_succeeds_after_transient_permission_error", "retried", payload.get("status"))
    add_check(results, "output_contract", case_id, "atomic_write_retry_attempt_count_matches_transient_failures", 3, replace_attempts["count"])
    add_check(results, "output_contract", case_id, "atomic_write_retry_cleans_temp_files_after_success", [], remaining_temp_files)

    summary["replace_attempts"] = replace_attempts["count"]
    return results, summary

def validate_atomic_write_cleanup_error_preserves_root_exception_case(_base_params):
    case_id = "ATOMIC_WRITE_CLEANUP_ERROR_PRESERVES_ROOT_EXCEPTION"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    with tempfile.TemporaryDirectory(prefix="atomic_write_cleanup_error_contract_") as temp_dir:
        run_dir = Path(temp_dir)
        json_path = run_dir / "atomic_cleanup_summary.json"
        local_common.write_json(json_path, {"status": "original", "count": 1})

        original_unlink = Path.unlink

        def failing_unlink(self, *args, **kwargs):
            if self.parent == run_dir and self.name.startswith(f".{json_path.name}.") and self.name.endswith(".tmp"):
                raise PermissionError("simulated temp cleanup failure")
            return original_unlink(self, *args, **kwargs)

        with patch.object(local_common.os, "replace", side_effect=OSError("simulated atomic replace failure")),              patch.object(Path, "unlink", new=failing_unlink):
            try:
                local_common.write_json(json_path, {"status": "new", "count": 2})
            except OSError as exc:
                raised_exc = exc
            else:
                raise AssertionError("write_json should raise original replace failure when temp cleanup also fails")

        remaining_temp_files = sorted(path.name for path in run_dir.glob(".*.tmp"))
        notes = list(getattr(raised_exc, "__notes__", []) or [])

    add_check(results, "output_contract", case_id, "atomic_write_cleanup_failure_preserves_root_exception_type", "OSError", type(raised_exc).__name__)
    add_check(results, "output_contract", case_id, "atomic_write_cleanup_failure_preserves_root_exception_message", True, "simulated atomic replace failure" in str(raised_exc))
    add_check(results, "output_contract", case_id, "atomic_write_cleanup_failure_records_cleanup_note", True, any("simulated temp cleanup failure" in note for note in notes))
    add_check(results, "output_contract", case_id, "atomic_write_cleanup_failure_leaves_temp_file_visible_for_manual_cleanup", 1, len(remaining_temp_files))

    summary["cleanup_notes_count"] = len(notes)
    return results, summary


def validate_validate_summary_atomic_write_contract_case(_base_params):
    case_id = "VALIDATE_SUMMARY_ATOMIC_WRITE"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    with tempfile.TemporaryDirectory(prefix="validate_summary_atomic_contract_") as temp_dir:
        run_dir = Path(temp_dir) / "run_dir"
        run_dir.mkdir(parents=True, exist_ok=True)
        summary_path = run_dir / "validate_consistency_summary.json"
        original_payload = {"status": "OLD", "elapsed_time_sec": 1.0}
        local_common.write_json(summary_path, original_payload)

        prev_run_dir = os.environ.get(LOCAL_REGRESSION_RUN_DIR_ENV)
        os.environ[LOCAL_REGRESSION_RUN_DIR_ENV] = str(run_dir)
        raised_exc = None
        try:
            df_results = pd.DataFrame([{"ticker": "1101", "status": "PASS"}])
            df_failed = pd.DataFrame(columns=["ticker", "module", "metric", "passed", "expected", "actual", "note"])
            with patch.object(local_common.os, "replace", side_effect=OSError("simulated validate summary replace failure")):
                try:
                    write_local_regression_summary(
                        dataset_profile_key="reduced",
                        dataset_source="direct",
                        data_dir="data/tw_stock_data_vip_reduced",
                        csv_path="outputs/validate_consistency/consistency_full_scan.csv",
                        xlsx_path="",
                        elapsed_time=3.21,
                        selected_tickers=["1101"],
                        df_results=df_results,
                        df_failed=df_failed,
                        output_dir="outputs/validate_consistency",
                        real_data_coverage_ok=True,
                        peak_traced_memory_mb=9.87,
                    )
                except OSError as exc:
                    raised_exc = exc
                else:
                    raise AssertionError("write_local_regression_summary should propagate atomic replace failure")
        finally:
            if prev_run_dir is None:
                os.environ.pop(LOCAL_REGRESSION_RUN_DIR_ENV, None)
            else:
                os.environ[LOCAL_REGRESSION_RUN_DIR_ENV] = prev_run_dir

        persisted_payload = json.loads(summary_path.read_text(encoding="utf-8"))
        remaining_temp_files = sorted(path.name for path in run_dir.glob(f".{summary_path.name}.*.tmp"))

    add_check(results, "output_contract", case_id, "validate_summary_atomic_write_raises_on_replace_failure", "OSError", type(raised_exc).__name__)
    add_check(results, "output_contract", case_id, "validate_summary_atomic_write_preserves_previous_payload", original_payload, persisted_payload)
    add_check(results, "output_contract", case_id, "validate_summary_atomic_write_cleans_temp_files_after_failure", [], remaining_temp_files)

    summary["cleanup_temp_files"] = len(remaining_temp_files)
    return results, summary


def validate_meta_quality_performance_memory_contract_case(_base_params):
    case_id = "META_QUALITY_PERFORMANCE_MEMORY_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    with tempfile.TemporaryDirectory(prefix="meta_quality_performance_memory_contract_") as temp_dir:
        run_dir = Path(temp_dir) / "run_dir"
        run_dir.mkdir(parents=True, exist_ok=True)
        write_json(run_dir / "quick_gate_summary.json", {"status": "PASS", "duration_sec": 1.0, "peak_traced_memory_mb": 12.5})
        write_json(run_dir / "validate_consistency_summary.json", {"status": "PASS", "elapsed_time_sec": 2.0, "peak_traced_memory_mb": 32.5})
        write_json(run_dir / "chain_summary.json", {"status": "PASS", "duration_sec": 3.0, "peak_traced_memory_mb": 64.0})
        write_json(run_dir / "ml_smoke_summary.json", {"status": "PASS", "duration_sec": 4.0, "peak_traced_memory_mb": 48.0, "optimizer_profile_trial_count": 1, "optimizer_profile_avg_objective_wall_sec": 1.5})

        manifest = dict(local_common.MANIFEST_DEFAULTS)
        with patch.dict(os.environ, {LOCAL_REGRESSION_RUN_DIR_ENV: str(run_dir)}):
            perf = build_performance_summary(
                run_dir,
                manifest,
                current_meta_quality_duration_sec=1.25,
                current_meta_quality_peak_traced_memory_mb=40.0,
            )

        add_check(results, "output_contract", case_id, "performance_memory_contract_ok", True, perf.get("ok"))
        add_check(results, "output_contract", case_id, "performance_step_peak_memory_keys", ["chain_checks", "consistency", "ml_smoke", "quick_gate"], sorted(perf.get("step_peak_traced_memory_mb", {}).keys()))
        add_check(results, "output_contract", case_id, "performance_max_peak_memory_value", 64.0, perf.get("max_step_peak_traced_memory_mb"))
        add_check(results, "output_contract", case_id, "performance_meta_quality_peak_memory_value", 40.0, perf.get("meta_quality_peak_traced_memory_mb"))

    summary["step_peak_memory_count"] = 4
    return results, summary


def validate_portfolio_sim_prepared_tool_contract_case(base_params):
    case_id = "PORTFOLIO_SIM_PREPARED_TOOL_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    case = build_synthetic_competing_candidates_case(base_params, make_synthetic_validation_params)
    ticker = case["primary_ticker"]
    frame = case["frames"][ticker]

    with tempfile.TemporaryDirectory(prefix="portfolio_sim_prepared_tool_") as temp_dir:
        write_synthetic_csv_bundle(temp_dir, {ticker: frame})
        file_path = str(Path(temp_dir) / f"{ticker}.csv")
        legacy_stats = run_portfolio_sim_tool_check_for_dir(
            temp_dir,
            case["params"],
            max_positions=1,
            enable_rotation=False,
            start_year=case["start_year"],
            benchmark_ticker=ticker,
        )
        prepared_stats = run_portfolio_sim_tool_check(ticker, file_path, case["params"])

    metric_names = [
        "trade_count",
        "win_rate",
        "pf_ev",
        "pf_payoff",
        "final_eq",
        "annual_return_pct",
        "reserved_buy_fill_rate",
        "portfolio_buy_rows",
        "portfolio_missed_buy_rows",
        "portfolio_half_take_profit_rows",
    ]
    for metric in metric_names:
        add_check(results, "output_contract", case_id, f"portfolio_sim_prepared_{metric}", legacy_stats[metric], prepared_stats[metric])

    add_check(results, "output_contract", case_id, "portfolio_sim_prepared_module_path", legacy_stats["module_path"], prepared_stats["module_path"])
    return results, summary



def validate_scanner_prepared_tool_contract_case(base_params):
    case_id = "SCANNER_PREPARED_TOOL_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    case = build_synthetic_competing_candidates_case(base_params, make_synthetic_validation_params)
    ticker = case["primary_ticker"]
    frame = case["frames"][ticker]

    with tempfile.TemporaryDirectory(prefix="scanner_prepared_tool_") as temp_dir:
        write_synthetic_csv_bundle(temp_dir, {ticker: frame})
        file_path = str(Path(temp_dir) / f"{ticker}.csv")
        legacy_result, legacy_module_path = run_scanner_tool_check(ticker, file_path, case["params"])
        min_rows_needed = get_required_min_rows(case["params"])
        clean_df, sanitize_stats = sanitize_ohlcv_dataframe(frame.copy(), ticker, min_rows=min_rows_needed)
        prepared_df, _standalone_logs, prepared_stats = prep_stock_data_and_trades(clean_df.copy(), case["params"], return_stats=True)
        prepared_result, prepared_module_path = run_scanner_tool_check(
            ticker,
            file_path,
            case["params"],
            prepared_df=prepared_df,
            sanitize_stats=sanitize_stats,
            precomputed_stats=prepared_stats,
        )

    add_check(results, "output_contract", case_id, "scanner_prepared_module_path", legacy_module_path, prepared_module_path)
    add_check(results, "output_contract", case_id, "scanner_prepared_payload", legacy_result, prepared_result)
    return results, summary


def validate_single_ticker_compounding_parity_contract_case(base_params):
    from core.backtest_core import run_v16_backtest
    from tools.validate.real_case_runners import build_single_ticker_portfolio_context, run_single_ticker_portfolio_check
    from tools.validate.scanner_expectations import build_consistency_parity_params

    params = make_synthetic_validation_params(base_params, tp_percent=0.0)
    params.initial_capital = 1000.0
    params.fixed_risk = 1.0
    params.buy_fee = 0.0
    params.sell_fee = 0.0
    params.tax_rate = 0.0
    params.min_fee = 0.0
    params.atr_buy_tol = 0.0
    params.atr_times_init = 1.0
    params.atr_times_trail = 1.0
    params.use_compounding = True
    params.max_position_cap_pct = 1.0

    case_id = "OUTPUT_SINGLE_TICKER_COMPOUNDING_PARITY"
    ticker = case_id
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    scenarios = [
        {
            "name": "after_loss",
            "expected_final_eq": 1040.0,
            "expected_total_return": 4.0,
            "df": pd.DataFrame(
                {
                    "Open": [100.0, 100.0, 95.0, 100.0, 110.0],
                    "High": [100.0, 100.0, 96.0, 100.0, 111.0],
                    "Low": [100.0, 100.0, 89.0, 100.0, 109.0],
                    "Close": [100.0, 100.0, 90.0, 100.0, 110.0],
                    "Volume": [1000.0, 1000.0, 1000.0, 1000.0, 1000.0],
                },
                index=pd.to_datetime([
                    "2024-01-01",
                    "2024-01-02",
                    "2024-01-03",
                    "2024-01-04",
                    "2024-01-05",
                ]),
            ),
            "precomputed_signals": (
                pd.Series([10.0, 10.0, 10.0, 10.0, 10.0], dtype=float).to_numpy(),
                pd.Series([True, False, True, False, False], dtype=bool).to_numpy(),
                pd.Series([False, True, False, True, False], dtype=bool).to_numpy(),
                pd.Series([100.0, float("nan"), 100.0, float("nan"), float("nan")], dtype=float).to_numpy(),
            ),
        },
        {
            "name": "after_gain",
            "expected_final_eq": 990.0,
            "expected_total_return": -1.0,
            "df": pd.DataFrame(
                {
                    "Open": [100.0, 100.0, 110.0, 100.0, 100.0, 90.0],
                    "High": [100.0, 100.0, 110.0, 100.0, 100.0, 90.0],
                    "Low": [100.0, 100.0, 110.0, 100.0, 100.0, 90.0],
                    "Close": [100.0, 100.0, 110.0, 100.0, 100.0, 90.0],
                    "Volume": [1000.0, 1000.0, 1000.0, 1000.0, 1000.0, 1000.0],
                },
                index=pd.to_datetime([
                    "2024-01-01",
                    "2024-01-02",
                    "2024-01-03",
                    "2024-01-04",
                    "2024-01-05",
                    "2024-01-06",
                ]),
            ),
            "precomputed_signals": (
                pd.Series([10.0, 10.0, 10.0, 10.0, 10.0, 10.0], dtype=float).to_numpy(),
                pd.Series([True, False, False, True, False, False], dtype=bool).to_numpy(),
                pd.Series([False, True, False, False, True, False], dtype=bool).to_numpy(),
                pd.Series([100.0, float("nan"), float("nan"), 100.0, float("nan"), float("nan")], dtype=float).to_numpy(),
            ),
        },
    ]

    for scenario in scenarios:
        single_stats, standalone_logs = run_v16_backtest(
            scenario["df"].copy(),
            params,
            return_logs=True,
            precomputed_signals=scenario["precomputed_signals"],
        )
        prepared_df = scenario["df"].copy()
        prepared_df["ATR"] = scenario["precomputed_signals"][0]
        prepared_df["is_setup"] = scenario["precomputed_signals"][1]
        prepared_df["ind_sell_signal"] = scenario["precomputed_signals"][2]
        prepared_df["buy_limit"] = scenario["precomputed_signals"][3]

        parity_params = build_consistency_parity_params(params)
        portfolio_context = build_single_ticker_portfolio_context(ticker, prepared_df, standalone_logs)
        portfolio_stats = run_single_ticker_portfolio_check(
            ticker,
            prepared_df,
            standalone_logs,
            parity_params,
            portfolio_context=portfolio_context,
        )
        portfolio_sim_stats = run_portfolio_sim_tool_check(
            ticker,
            f"{ticker}.csv",
            parity_params,
            prepared_df=prepared_df,
            standalone_logs=standalone_logs,
            packed_fast_data=portfolio_context["fast_data"],
            sorted_dates=portfolio_context["sorted_dates"],
            start_year=portfolio_context["start_year"],
        )

        metric_prefix = f"single_ticker_compounding_{scenario['name']}"
        add_check(results, "output_contract", case_id, f"{metric_prefix}_params_keep_compounding_enabled", True, bool(parity_params.use_compounding))
        add_check(results, "output_contract", case_id, f"{metric_prefix}_single_backtest_trade_count", 2, int(single_stats["trade_count"]))
        add_check(results, "output_contract", case_id, f"{metric_prefix}_single_backtest_final_equity_uses_compounding", scenario["expected_final_eq"], params.initial_capital * (1.0 + float(single_stats["asset_growth"]) / 100.0))
        add_check(results, "output_contract", case_id, f"{metric_prefix}_single_backtest_total_return_uses_compounding", scenario["expected_total_return"], float(single_stats["asset_growth"]))
        add_check(results, "output_contract", case_id, f"{metric_prefix}_portfolio_total_return_matches_single_backtest", float(single_stats["asset_growth"]), float(portfolio_stats["total_return"]))
        add_check(results, "output_contract", case_id, f"{metric_prefix}_portfolio_final_equity_matches_single_backtest", scenario["expected_final_eq"], float(portfolio_stats["final_eq"]))
        add_check(results, "output_contract", case_id, f"{metric_prefix}_portfolio_sim_total_return_matches_single_backtest", float(single_stats["asset_growth"]), float(portfolio_sim_stats["total_return"]))
        add_check(results, "output_contract", case_id, f"{metric_prefix}_portfolio_sim_final_equity_matches_single_backtest", scenario["expected_final_eq"], float(portfolio_sim_stats["final_eq"]))

        summary[f"single_asset_growth_{scenario['name']}"] = float(single_stats["asset_growth"])
        summary[f"portfolio_total_return_{scenario['name']}"] = float(portfolio_stats["total_return"])
        summary[f"portfolio_sim_total_return_{scenario['name']}"] = float(portfolio_sim_stats["total_return"])
    return results, summary


def validate_scanner_live_capital_contract_case(base_params):
    case_id = "SCANNER_LIVE_CAPITAL_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    params = make_synthetic_validation_params(base_params, tp_percent=0.0)
    params.initial_capital = 1_000_000.0
    params.fixed_risk = 0.01
    params.buy_fee = 0.0
    params.sell_fee = 0.0
    params.tax_rate = 0.0
    params.min_fee = 0.0
    params.scanner_live_capital = 2_000_000.0
    params.max_position_cap_pct = 1.0

    buy_limit = 100.0
    stop_loss = 90.0
    projected_qty = calc_reference_candidate_qty(buy_limit, stop_loss, params, ticker="2330", trade_date=pd.Timestamp("2026-01-02"))
    projected_cost = calc_entry_total_cost(buy_limit, projected_qty, params)

    scanner_result = build_scanner_response_from_stats(
        ticker="2330",
        stats={
            "is_candidate": True,
            "is_setup_today": True,
            "buy_limit": buy_limit,
            "stop_loss": stop_loss,
            "expected_value": 1.25,
            "trade_count": 8,
            "win_rate": 62.5,
            "max_drawdown": 9.5,
            "extended_candidate_today": None,
        },
        params=params,
        sanitize_stats={},
    )

    add_check(results, "output_contract", case_id, "scanner_live_capital_qty", 2000, projected_qty)
    add_check(results, "output_contract", case_id, "scanner_live_capital_proj_cost", 200000.0, projected_cost)
    add_check(results, "output_contract", case_id, "scanner_result_status", "buy", None if scanner_result is None else scanner_result[0])
    add_check(results, "output_contract", case_id, "scanner_result_proj_cost_uses_scanner_live_capital", 200000.0, None if scanner_result is None else float(scanner_result[1]))
    add_check(results, "output_contract", case_id, "scanner_result_message_mentions_scanner_live_capital_cost", True, False if scanner_result is None else "參考投入:200,000" in scanner_result[4])

    summary["projected_qty"] = projected_qty
    summary["projected_cost"] = projected_cost
    return results, summary


def validate_scanner_reference_clean_df_contract_case(base_params):
    case_id = "SCANNER_REFERENCE_CLEAN_DF_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    case = build_synthetic_competing_candidates_case(base_params, make_synthetic_validation_params)
    ticker = case["primary_ticker"]
    frame = case["frames"][ticker]

    with tempfile.TemporaryDirectory(prefix="scanner_reference_clean_df_") as temp_dir:
        write_synthetic_csv_bundle(temp_dir, {ticker: frame})
        file_path = str(Path(temp_dir) / f"{ticker}.csv")
        legacy_stats = run_scanner_reference_check(ticker, file_path, case["params"])
        min_rows_needed = get_required_min_rows(case["params"])
        clean_df, _sanitize_stats = sanitize_ohlcv_dataframe(frame.copy(), ticker, min_rows=min_rows_needed)
        clean_df_stats = run_scanner_reference_check_on_clean_df(ticker, clean_df, case["params"])

    metric_names = [
        "trade_count",
        "win_rate",
        "expected_value",
        "asset_growth",
        "max_drawdown",
        "missed_buys",
        "missed_sells",
        "is_candidate",
        "is_setup_today",
        "current_position",
    ]
    for metric in metric_names:
        add_check(results, "output_contract", case_id, f"scanner_reference_clean_df_{metric}", legacy_stats[metric], clean_df_stats[metric])

    add_check(results, "output_contract", case_id, "scanner_reference_clean_df_extended_candidate", legacy_stats.get("extended_candidate_today"), clean_df_stats.get("extended_candidate_today"))
    return results, summary


def validate_debug_chart_payload_without_html_export_contract_case(base_params):
    case_id = "DEBUG_CHART_PAYLOAD_WITHOUT_HTML_EXPORT_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    case = build_synthetic_competing_candidates_case(base_params, make_synthetic_validation_params)
    ticker = case["primary_ticker"]
    frame = case["frames"][ticker]
    min_rows_needed = get_required_min_rows(case["params"])
    clean_df, _sanitize_stats = sanitize_ohlcv_dataframe(frame.copy(), ticker, min_rows=min_rows_needed)

    debug_module = importlib.import_module("tools.trade_analysis.trade_log")
    with tempfile.TemporaryDirectory(prefix="debug_chart_payload_contract_") as temp_dir:
        analysis_result = debug_module.run_debug_analysis(
            clean_df.copy(),
            ticker,
            case["params"],
            export_excel=False,
            export_chart=False,
            return_chart_payload=True,
            verbose=False,
            output_dir=temp_dir,
        )

    chart_payload = analysis_result.get("chart_payload")
    add_check(results, "output_contract", case_id, "debug_chart_payload_available_without_html_export", True, chart_payload is not None)
    add_check(results, "output_contract", case_id, "debug_chart_path_omitted_without_html_export", None, analysis_result.get("chart_path"))
    add_check(results, "output_contract", case_id, "debug_chart_payload_total_bars_matches_clean_df", len(clean_df), 0 if chart_payload is None else len(chart_payload.get("x", [])))

    if chart_payload is not None:
        add_check(results, "output_contract", case_id, "debug_chart_payload_default_view_nonempty_without_html_export", True, int(chart_payload["default_view"]["end_idx"]) >= int(chart_payload["default_view"]["start_idx"]))

    summary["ticker"] = ticker
    return results, summary


def validate_debug_empty_price_df_chart_payload_contract_case(base_params):
    case_id = "DEBUG_EMPTY_PRICE_DF_CHART_PAYLOAD_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    case = build_synthetic_competing_candidates_case(base_params, make_synthetic_validation_params)
    params = case["params"]
    empty_df = pd.DataFrame(columns=["Open", "High", "Low", "Close", "Volume"])
    with tempfile.TemporaryDirectory(prefix="debug_empty_price_df_chart_payload_contract_") as temp_dir:
        analysis_result = importlib.import_module("tools.trade_analysis.trade_log").run_debug_analysis(
            empty_df.copy(),
            "EMPTY",
            params,
            export_excel=False,
            export_chart=False,
            return_chart_payload=True,
            verbose=False,
            output_dir=temp_dir,
        )

    chart_payload = analysis_result.get("chart_payload")
    add_check(results, "output_contract", case_id, "debug_empty_price_df_chart_payload_exists", True, chart_payload is not None)
    add_check(results, "output_contract", case_id, "debug_empty_price_df_chart_payload_is_placeholder_single_bar", 1, 0 if chart_payload is None else len(chart_payload.get("x", [])))
    if chart_payload is not None:
        add_check(results, "output_contract", case_id, "debug_empty_price_df_chart_payload_volume_zero", 0.0, float(chart_payload["volume"][0]))
        add_check(results, "output_contract", case_id, "debug_empty_price_df_chart_payload_open_zero", 0.0, float(chart_payload["open"][0]))
    summary["payload_bar_count"] = 0 if chart_payload is None else len(chart_payload.get("x", []))
    return results, summary


def validate_gui_embedded_chart_contract_case(base_params):
    case_id = "GUI_EMBEDDED_CHART_VIEWPORT_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    workbench_spec = importlib.import_module("tools.workbench_ui").build_workbench_spec()
    panel_spec = workbench_spec.get("panels", [])[0]
    add_check(results, "output_contract", case_id, "gui_embedded_chart_backend", "tools.trade_analysis.charting.create_matplotlib_trade_chart_figure", panel_spec.get("inline_chart_backend"))
    add_check(results, "output_contract", case_id, "gui_embedded_chart_default_show_volume", False, panel_spec.get("default_show_volume"))

    case = build_synthetic_competing_candidates_case(base_params, make_synthetic_validation_params)
    ticker = case["primary_ticker"]
    frame = case["frames"][ticker]
    min_rows_needed = get_required_min_rows(case["params"])
    clean_df, _sanitize_stats = sanitize_ohlcv_dataframe(frame.copy(), ticker, min_rows=min_rows_needed)

    debug_module = importlib.import_module("tools.trade_analysis.trade_log")
    with tempfile.TemporaryDirectory(prefix="gui_embedded_chart_contract_") as temp_dir:
        analysis_result = debug_module.run_debug_analysis(
            clean_df.copy(),
            ticker,
            case["params"],
            export_excel=False,
            export_chart=False,
            return_chart_payload=True,
            verbose=False,
            output_dir=temp_dir,
        )

    chart_payload = analysis_result.get("chart_payload")
    add_check(results, "output_contract", case_id, "gui_embedded_chart_payload_exists", True, chart_payload is not None)
    if chart_payload is None:
        return results, summary

    default_view = chart_payload.get("default_view", {})
    start_idx = int(default_view.get("start_idx", 0))
    end_idx = int(default_view.get("end_idx", 0))
    focus_positions = [int(pos) for pos in chart_payload.get("focus_positions", [])]
    add_check(results, "output_contract", case_id, "gui_embedded_chart_default_view_nonempty", True, end_idx >= start_idx)
    add_check(results, "output_contract", case_id, "gui_embedded_chart_focus_positions_in_view", True, all(start_idx <= pos <= end_idx for pos in focus_positions))

    visible_ranges = compute_visible_value_ranges(chart_payload, start_idx=start_idx, end_idx=end_idx)
    add_check(results, "output_contract", case_id, "gui_embedded_chart_visible_price_range_valid", True, visible_ranges["price_max"] > visible_ranges["price_min"])
    add_check(results, "output_contract", case_id, "gui_embedded_chart_visible_volume_range_valid", True, visible_ranges["volume_max"] > visible_ranges["volume_min"])

    outlier_window_start = max(1, len(chart_payload["x"]) - 15)
    outlier_window_end = len(chart_payload["x"]) - 1
    base_outlier_ranges = compute_visible_value_ranges(chart_payload, start_idx=outlier_window_start, end_idx=outlier_window_end)
    outlier_payload = dict(chart_payload)
    outlier_payload["high"] = np.array(chart_payload["high"], copy=True)
    outlier_payload["low"] = np.array(chart_payload["low"], copy=True)
    outlier_payload["high"][0] = outlier_payload["high"][0] + 100000.0
    outlier_payload["low"][0] = max(0.0, outlier_payload["low"][0] - 100000.0)
    outlier_ranges = compute_visible_value_ranges(outlier_payload, start_idx=outlier_window_start, end_idx=outlier_window_end)
    add_check(results, "output_contract", case_id, "gui_embedded_chart_offscreen_outlier_ignored", base_outlier_ranges["price_max"], outlier_ranges["price_max"])

    figure = create_matplotlib_trade_chart_figure(chart_payload=chart_payload, ticker=ticker, show_volume=False)
    contract = getattr(figure, "_stock_chart_contract", {})
    add_check(results, "output_contract", case_id, "gui_embedded_chart_figure_axes_count_without_volume", 1, len(figure.axes))
    add_check(results, "output_contract", case_id, "gui_embedded_chart_contract_volume_hidden", False, contract.get("volume_visible"))
    add_check(results, "output_contract", case_id, "gui_embedded_chart_contract_default_view_preserved", True, contract.get("default_view", {}).get("end_idx", -1) >= contract.get("default_view", {}).get("start_idx", 0))
    figure.clear()

    figure_with_volume = create_matplotlib_trade_chart_figure(chart_payload=chart_payload, ticker=ticker, show_volume=True)
    contract_with_volume = getattr(figure_with_volume, "_stock_chart_contract", {})
    add_check(results, "output_contract", case_id, "gui_embedded_chart_contract_volume_visible", True, contract_with_volume.get("volume_visible"))
    add_check(results, "output_contract", case_id, "gui_embedded_chart_volume_overlay_mode", "inset", contract_with_volume.get("volume_overlay_mode"))
    add_check(results, "output_contract", case_id, "gui_embedded_chart_volume_overlay_axis_present", True, contract_with_volume.get("volume_overlay_axis_present"))
    figure_with_volume.clear()

    large_dates = pd.date_range("2020-01-01", periods=900, freq="B")
    large_frame = pd.DataFrame(
        {
            "Open": np.linspace(50.0, 130.0, len(large_dates)),
            "High": np.linspace(51.0, 131.0, len(large_dates)),
            "Low": np.linspace(49.0, 129.0, len(large_dates)),
            "Close": np.linspace(50.5, 130.5, len(large_dates)),
            "Volume": np.linspace(1000.0, 5000.0, len(large_dates)),
        },
        index=large_dates,
    )
    large_chart_context = {
        "stop_line": np.full(len(large_dates), np.nan, dtype=np.float64),
        "tp_line": np.full(len(large_dates), np.nan, dtype=np.float64),
        "order_markers": [{"trace_name": "限價買進", "date": large_dates[-25], "price": 127.0, "hover_text": "限價買進"}],
        "trade_markers": [{"trace_name": "買進", "date": large_dates[-20], "price": 128.0, "hover_text": "買進"}],
    }
    large_chart_payload = build_debug_chart_payload(large_frame, large_chart_context)
    large_figure = create_matplotlib_trade_chart_figure(chart_payload=large_chart_payload, ticker="LARGE", show_volume=False)
    large_contract = getattr(large_figure, "_stock_chart_contract", {})
    add_check(results, "output_contract", case_id, "gui_embedded_chart_full_history_render_window", large_contract.get("total_bar_count", 0), large_contract.get("render_bar_count", 0))
    add_check(results, "output_contract", case_id, "gui_embedded_chart_full_history_navigation_enabled", True, large_contract.get("full_history_navigation_enabled"))
    add_check(results, "output_contract", case_id, "gui_embedded_chart_cjk_font_candidates_include_jhenghei", True, "Microsoft JhengHei" in get_matplotlib_cjk_font_candidates())
    add_check(results, "output_contract", case_id, "gui_embedded_chart_cjk_font_candidates_include_noto_tc", True, "Noto Sans CJK TC" in get_matplotlib_cjk_font_candidates())
    large_figure.clear()

    summary["default_view"] = {"start_idx": start_idx, "end_idx": end_idx}
    return results, summary


def validate_chart_payload_optional_overlay_keys_contract_case(_base_params):
    case_id = "CHART_PAYLOAD_OPTIONAL_OVERLAY_KEYS_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    minimal_payload = {
        "dates": pd.date_range("2024-01-01", periods=8, freq="D"),
        "date_labels": [f"2024-01-{day:02d}" for day in range(1, 9)],
        "x": np.arange(8, dtype=np.float64),
        "open": np.array([10, 11, 12, 11, 13, 14, 13, 15], dtype=np.float64),
        "high": np.array([11, 12, 13, 12, 14, 15, 14, 16], dtype=np.float64),
        "low": np.array([9, 10, 11, 10, 12, 13, 12, 14], dtype=np.float64),
        "close": np.array([10.5, 11.5, 11.8, 11.2, 13.4, 14.2, 13.1, 15.3], dtype=np.float64),
        "volume": np.array([100, 120, 130, 110, 150, 160, 140, 180], dtype=np.float64),
        "up_mask": np.array([True, True, False, True, True, True, False, True], dtype=bool),
        "marker_groups": {},
        "focus_positions": [2, 5],
        "default_view": {"start_idx": 1, "end_idx": 6},
    }
    normalized_payload = normalize_chart_payload_contract(minimal_payload)
    add_check(results, "output_contract", case_id, "chart_payload_optional_limit_line_filled", 8, int(len(normalized_payload.get("limit_line", []))))
    add_check(results, "output_contract", case_id, "chart_payload_optional_entry_line_filled", 8, int(len(normalized_payload.get("entry_line", []))))
    add_check(results, "output_contract", case_id, "chart_payload_optional_tp_line_filled", 8, int(len(normalized_payload.get("tp_line", []))))
    add_check(results, "output_contract", case_id, "chart_payload_optional_stop_line_filled", 8, int(len(normalized_payload.get("stop_line", []))))
    add_check(results, "output_contract", case_id, "chart_payload_optional_summary_box_defaulted", [], normalized_payload.get("summary_box"))
    add_check(results, "output_contract", case_id, "chart_payload_optional_status_box_defaulted", {}, normalized_payload.get("status_box"))

    figure = create_matplotlib_trade_chart_figure(chart_payload=minimal_payload, ticker="2330", show_volume=False)
    contract = getattr(figure, "_stock_chart_contract", {})
    add_check(results, "output_contract", case_id, "chart_payload_optional_overlay_keys_figure_builds_successfully", True, contract.get("total_bar_count", 0) == len(minimal_payload["x"]))
    figure.clear()
    return results, summary


def validate_gui_mouse_navigation_contract_case(_base_params):
    case_id = "GUI_MOUSE_NAVIGATION_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    inspector_source = build_project_absolute_path("tools", "workbench_ui", "single_stock_inspector.py").read_text(encoding="utf-8")
    charting_source = build_project_absolute_path("tools", "trade_analysis", "charting.py").read_text(encoding="utf-8")

    add_check(results, "output_contract", case_id, "gui_chart_does_not_require_navigation_toolbar", True, "NavigationToolbar2Tk" not in inspector_source)
    add_check(results, "output_contract", case_id, "gui_chart_binds_mouse_navigation", True, "bind_matplotlib_chart_navigation(figure, canvas)" in inspector_source)
    add_check(results, "output_contract", case_id, "charting_declares_mouse_navigation_binder", True, "def bind_matplotlib_chart_navigation(figure, canvas):" in charting_source)
    add_check(results, "output_contract", case_id, "charting_supports_wheel_zoom_event", True, '"scroll_event"' in charting_source)
    add_check(results, "output_contract", case_id, "charting_supports_left_drag_pan_event", True, '"motion_notify_event"' in charting_source and 'event.button != 1' in charting_source)

    minimal_payload = {
        "dates": pd.date_range("2024-01-01", periods=8, freq="D"),
        "date_labels": [f"2024-01-{day:02d}" for day in range(1, 9)],
        "x": np.arange(8, dtype=np.float64),
        "open": np.array([10, 11, 12, 11, 13, 14, 13, 15], dtype=np.float64),
        "high": np.array([11, 12, 13, 12, 14, 15, 14, 16], dtype=np.float64),
        "low": np.array([9, 10, 11, 10, 12, 13, 12, 14], dtype=np.float64),
        "close": np.array([10.5, 11.5, 11.8, 11.2, 13.4, 14.2, 13.1, 15.3], dtype=np.float64),
        "volume": np.array([100, 120, 130, 110, 150, 160, 140, 180], dtype=np.float64),
        "up_mask": np.array([True, True, False, True, True, True, False, True], dtype=bool),
        "stop_line": np.full(8, np.nan, dtype=np.float64),
        "tp_line": np.full(8, np.nan, dtype=np.float64),
        "marker_groups": {},
        "focus_positions": [2, 5],
        "default_view": {"start_idx": 1, "end_idx": 6},
    }
    figure = create_matplotlib_trade_chart_figure(chart_payload=minimal_payload, ticker="2330", show_volume=False)
    contract = getattr(figure, "_stock_chart_contract", {})
    add_check(results, "output_contract", case_id, "figure_contract_toolbar_not_required", False, contract.get("toolbar_required"))
    add_check(results, "output_contract", case_id, "figure_contract_mouse_zoom_flag_default_false_until_canvas_bind", False, contract.get("mouse_wheel_zoom_enabled"))
    add_check(results, "output_contract", case_id, "chart_navigation_binder_callable", True, callable(bind_matplotlib_chart_navigation))
    return results, summary


def _build_matplotlib_canvas_stub():
    class _CanvasWidget:
        def __init__(self):
            self.config_calls = []
        def configure(self, **kwargs):
            self.config_calls.append(kwargs)
        def focus_set(self):
            return None

    class _CanvasStub:
        def __init__(self):
            self._widget = _CanvasWidget()
            self._connections = {}
            self._grabbed_axis = None
            self.release_calls = []
            self.toolbar = None
        def mpl_connect(self, name, callback):
            self._connections[name] = callback
            return len(self._connections)
        def get_tk_widget(self):
            return self._widget
        def draw_idle(self):
            return None
        def grab_mouse(self, axis):
            self._grabbed_axis = axis
        def release_mouse(self, axis):
            self.release_calls.append(axis)
            if self._grabbed_axis is axis:
                self._grabbed_axis = None

    return _CanvasStub()


def validate_gui_navigation_canvas_stub_cleanup_contract_case(_base_params):
    case_id = "GUI_NAVIGATION_CANVAS_STUB_CLEANUP_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    canvas = _build_matplotlib_canvas_stub()
    add_check(results, "output_contract", case_id, "canvas_stub_declares_grab_mouse", True, callable(getattr(canvas, "grab_mouse", None)))
    add_check(results, "output_contract", case_id, "canvas_stub_declares_release_mouse", True, callable(getattr(canvas, "release_mouse", None)))

    sentinel_axis = object()
    canvas.grab_mouse(sentinel_axis)
    canvas.release_mouse(sentinel_axis)
    add_check(results, "output_contract", case_id, "canvas_stub_release_mouse_clears_grabbed_axis", None, canvas._grabbed_axis)
    add_check(results, "output_contract", case_id, "canvas_stub_release_mouse_records_axis", 1, len(canvas.release_calls))
    return results, summary


def validate_gui_navigation_canvas_stub_toolbar_contract_case(_base_params):
    case_id = "GUI_NAVIGATION_CANVAS_STUB_TOOLBAR_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    canvas = _build_matplotlib_canvas_stub()
    add_check(results, "output_contract", case_id, "canvas_stub_declares_toolbar_attribute", True, hasattr(canvas, "toolbar"))
    add_check(results, "output_contract", case_id, "canvas_stub_toolbar_defaults_to_none", None, getattr(canvas, "toolbar", None))
    return results, summary


def validate_gui_dark_theme_and_keyboard_pan_contract_case(_base_params):
    case_id = "GUI_DARK_THEME_AND_KEYBOARD_PAN_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    workbench_source = build_project_absolute_path("tools", "workbench_ui", "workbench.py").read_text(encoding="utf-8")
    charting_source = build_project_absolute_path("tools", "trade_analysis", "charting.py").read_text(encoding="utf-8")
    workbench_spec = importlib.import_module("tools.workbench_ui").build_workbench_spec()

    add_check(results, "output_contract", case_id, "gui_workbench_declares_deep_dark_theme", True, workbench_spec.get("ui_theme") == "deep_dark")
    add_check(results, "output_contract", case_id, "gui_workbench_configures_dark_ttk_theme", True, "def configure_workbench_theme(root):" in workbench_source and 'style.theme_use("clam")' in workbench_source)
    add_check(results, "output_contract", case_id, "charting_supports_keyboard_pan_event", True, '"key_press_event"' in charting_source)
    add_check(results, "output_contract", case_id, "charting_supports_status_chip_rendering", True, "def _render_status_chips(axis_price, status_box, label_font):" in charting_source)
    add_check(results, "output_contract", case_id, "charting_supports_dynamic_candle_widths", True, "def _compute_dynamic_linewidths(axis_price, figure):" in charting_source)

    minimal_payload = {
        "dates": pd.date_range("2024-01-01", periods=120, freq="D"),
        "date_labels": [f"2024-01-{(day % 30) + 1:02d}" for day in range(120)],
        "x": np.arange(120, dtype=np.float64),
        "open": np.linspace(30.0, 42.0, 120, dtype=np.float64),
        "high": np.linspace(31.0, 43.0, 120, dtype=np.float64),
        "low": np.linspace(29.0, 41.0, 120, dtype=np.float64),
        "close": np.linspace(30.5, 42.5, 120, dtype=np.float64),
        "volume": np.linspace(1000.0, 5000.0, 120, dtype=np.float64),
        "up_mask": np.array([True if idx % 2 == 0 else False for idx in range(120)], dtype=bool),
        "stop_line": np.full(120, np.nan, dtype=np.float64),
        "tp_line": np.full(120, np.nan, dtype=np.float64),
        "limit_line": np.full(120, np.nan, dtype=np.float64),
        "entry_line": np.full(120, np.nan, dtype=np.float64),
        "marker_groups": {},
        "signal_annotations": [],
        "summary_box": ["資產成長: 1.0%"],
        "status_box": {"lines": ["出現買入訊號", "歷史績效符合"], "ok": True},
        "focus_positions": [100],
        "default_view": {"start_idx": 84, "end_idx": 119},
    }
    figure = create_matplotlib_trade_chart_figure(chart_payload=minimal_payload, ticker="2330", show_volume=True)

    canvas = _build_matplotlib_canvas_stub()
    figure.canvas = canvas
    bind_matplotlib_chart_navigation(figure, canvas)
    contract = getattr(figure, "_stock_chart_contract", {})
    add_check(results, "output_contract", case_id, "figure_contract_keyboard_pan_enabled", True, contract.get("keyboard_pan_enabled"))
    add_check(results, "output_contract", case_id, "figure_contract_dynamic_candle_width_enabled", True, contract.get("dynamic_candle_width_enabled"))
    add_check(results, "output_contract", case_id, "figure_contract_status_chip_layout", "right_bottom", contract.get("status_chip_layout"))
    figure.clear()
    return results, summary


def validate_gui_chart_workspace_contract_case(_base_params):
    case_id = "GUI_CHART_WORKSPACE_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    inspector_source = build_project_absolute_path("tools", "workbench_ui", "single_stock_inspector.py").read_text(encoding="utf-8")
    add_check(results, "output_contract", case_id, "gui_chart_workspace_uses_notebook_tabs", True, 'ttk.Notebook(self, style="Workbench.TNotebook")' in inspector_source)
    add_check(results, "output_contract", case_id, "gui_chart_workspace_omits_summary_tab", False, 'text="執行摘要"' in inspector_source)
    add_check(results, "output_contract", case_id, "gui_chart_workspace_has_trade_detail_tab", True, 'text="交易明細"' in inspector_source)
    add_check(results, "output_contract", case_id, "gui_chart_workspace_default_volume_hidden", True, 'self._show_volume_var = tk.BooleanVar(value=False)' in inspector_source)
    add_check(results, "output_contract", case_id, "gui_chart_workspace_volume_toggle_rerenders_chart", True, 'show_volume=bool(self._show_volume_var.get())' in inspector_source)

    workbench_spec = importlib.import_module("tools.workbench_ui").build_workbench_spec()
    panel_spec = workbench_spec.get("panels", [])[0]
    add_check(results, "output_contract", case_id, "gui_chart_workspace_panel_default_show_volume", False, panel_spec.get("default_show_volume"))
    add_check(results, "output_contract", case_id, "gui_chart_workspace_startup_window_mode", "maximized", workbench_spec.get("startup_window_mode"))

    large_dates = pd.date_range("2010-01-01", periods=2400, freq="B")
    large_frame = pd.DataFrame(
        {
            "Open": np.linspace(50.0, 150.0, len(large_dates)),
            "High": np.linspace(51.0, 151.0, len(large_dates)),
            "Low": np.linspace(49.0, 149.0, len(large_dates)),
            "Close": np.linspace(50.5, 150.5, len(large_dates)),
            "Volume": np.linspace(5000.0, 25000.0, len(large_dates)),
        },
        index=large_dates,
    )
    chart_payload = build_debug_chart_payload(large_frame, create_debug_chart_context(large_frame))
    figure = create_matplotlib_trade_chart_figure(chart_payload=chart_payload, ticker="LARGE", show_volume=True)
    contract = getattr(figure, "_stock_chart_contract", {})
    add_check(results, "output_contract", case_id, "gui_chart_workspace_full_history_bar_count", len(chart_payload["x"]), contract.get("render_bar_count"))
    add_check(results, "output_contract", case_id, "gui_chart_workspace_volume_overlay_ratio_leq_quarter", True, contract.get("volume_overlay_ratio", 1.0) <= 0.25)
    add_check(results, "output_contract", case_id, "gui_chart_workspace_volume_overlay_mode", "inset", contract.get("volume_overlay_mode"))
    add_check(results, "output_contract", case_id, "gui_chart_workspace_volume_overlay_axis_present", True, contract.get("volume_overlay_axis_present"))
    figure.clear()

    summary["panel_id"] = panel_spec.get("panel_id")
    return results, summary


def validate_record_signal_annotation_meta_contract_case(_base_params):
    case_id = "DEBUG_RECORD_SIGNAL_ANNOTATION_META_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    dates = pd.date_range("2024-01-01", periods=3, freq="D")
    frame = pd.DataFrame(
        {
            "Open": [10.0, 10.2, 10.4],
            "High": [10.4, 10.5, 10.8],
            "Low": [9.8, 10.0, 10.1],
            "Close": [10.1, 10.3, 10.6],
            "Volume": [1000.0, 1200.0, 1500.0],
        },
        index=dates,
    )
    chart_context = create_debug_chart_context(frame)
    meta_payload = {"profit_pct": 3.25, "source": "synthetic"}
    record_signal_annotation(
        chart_context,
        current_date=dates[1],
        signal_type="sell",
        anchor_price=float(frame.iloc[1]["Low"]),
        title="賣訊",
        detail_lines=["本次績效: +3.25%"],
        meta=meta_payload,
    )
    chart_payload = build_debug_chart_payload(frame, chart_context)
    normalized_annotations = chart_payload.get("signal_annotations", [])
    first_annotation = normalized_annotations[0] if normalized_annotations else {}

    add_check(results, "output_contract", case_id, "record_signal_annotation_accepts_meta_keyword", 1, len(chart_context["signal_annotations"]))
    add_check(results, "output_contract", case_id, "signal_annotation_meta_retained_in_chart_context", meta_payload, chart_context["signal_annotations"][0].get("meta") if chart_context["signal_annotations"] else None)
    add_check(results, "output_contract", case_id, "signal_annotation_meta_retained_in_chart_payload", meta_payload, first_annotation.get("meta"))
    add_check(results, "output_contract", case_id, "signal_annotation_profit_pct_available_for_sell_face_resolution", 3.25, (first_annotation.get("meta") or {}).get("profit_pct"))
    return results, summary


def validate_workbench_theme_accent_symbol_contract_case(_base_params):
    case_id = "GUI_WORKBENCH_THEME_ACCENT_SYMBOL_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    import tools.workbench_ui.workbench as workbench_module

    workbench_source = build_project_absolute_path("tools", "workbench_ui", "workbench.py").read_text(encoding="utf-8")
    add_check(results, "output_contract", case_id, "workbench_declares_accent_constant", True, hasattr(workbench_module, "WORKBENCH_ACCENT"))
    add_check(results, "output_contract", case_id, "workbench_accent_constant_is_hex_string", True, isinstance(getattr(workbench_module, "WORKBENCH_ACCENT", None), str) and str(getattr(workbench_module, "WORKBENCH_ACCENT", "")).startswith("#"))
    add_check(results, "output_contract", case_id, "workbench_theme_references_declared_accent_constant", True, "WORKBENCH_ACCENT" in workbench_source)
    return results, summary


def validate_gui_scanner_console_and_latest_contract_case(_base_params):
    case_id = "GUI_SCANNER_CONSOLE_AND_LATEST_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    inspector_source = build_project_absolute_path("tools", "workbench_ui", "single_stock_inspector.py").read_text(encoding="utf-8")
    charting_source = build_project_absolute_path("tools", "trade_analysis", "charting.py").read_text(encoding="utf-8")
    scanner_source = build_project_absolute_path("tools", "scanner", "scan_runner.py").read_text(encoding="utf-8")
    stock_processor_source = build_project_absolute_path("tools", "scanner", "stock_processor.py").read_text(encoding="utf-8")
    scanner_display_source = build_project_absolute_path("core", "scanner_display.py").read_text(encoding="utf-8")

    add_check(results, "output_contract", case_id, "gui_panel_uses_full_dataset_only", False, 'text="資料集"' in inspector_source)
    add_check(results, "output_contract", case_id, "gui_panel_exposes_scanner_button", True, 'text="計算候選股"' in inspector_source)
    add_check(results, "output_contract", case_id, "gui_panel_exposes_candidate_dropdown", True, '_candidate_combo' in inspector_source)
    add_check(results, "output_contract", case_id, "gui_panel_exposes_history_button", True, 'text="計算歷史績效股"' in inspector_source)
    add_check(results, "output_contract", case_id, "gui_panel_exposes_history_dropdown", True, '_history_combo' in inspector_source)
    add_check(results, "output_contract", case_id, "gui_history_dropdown_uses_sort_probe_text", True, 'build_scanner_sort_probe_text(' in inspector_source)
    add_check(results, "output_contract", case_id, "gui_panel_exposes_console_tab", True, 'text="Console"' in inspector_source)
    add_check(results, "output_contract", case_id, "gui_panel_exposes_latest_button", True, 'text="回到最新K線"' in inspector_source)
    add_check(results, "output_contract", case_id, "gui_panel_calls_scanner_runner", True, 'run_daily_scanner(' in inspector_source)
    add_check(results, "output_contract", case_id, "gui_panel_calls_history_scanner_runner", True, 'run_history_qualified_scanner(' in inspector_source)
    add_check(results, "output_contract", case_id, "charting_declares_scroll_to_latest_helper", True, 'def scroll_chart_to_latest(' in charting_source)
    add_check(results, "output_contract", case_id, "charting_declares_right_padding_bars", True, 'CHART_RIGHT_PADDING_BARS' in charting_source)
    add_check(results, "output_contract", case_id, "scanner_runner_returns_candidate_rows_payload", True, '"candidate_rows": list(candidate_rows)' in scanner_source or "'candidate_rows': list(candidate_rows)" in scanner_source)
    add_check(results, "output_contract", case_id, "scanner_runner_returns_history_rows_payload", True, "\"history_qualified_rows\": list(scan_state['rows'])" in scanner_source or "'history_qualified_rows': list(scan_state['rows'])" in scanner_source)
    add_check(results, "output_contract", case_id, "scanner_display_declares_asset_growth_probe", True, '資產成長' in scanner_display_source and 'build_scanner_sort_probe_text' in scanner_display_source)
    add_check(results, "output_contract", case_id, "scanner_row_text_uses_sort_probe_text", True, 'build_scanner_sort_probe_text(' in stock_processor_source)
    add_check(results, "output_contract", case_id, "scanner_history_rows_preserve_asset_growth_metric", True, 'asset_growth' in stock_processor_source and "'asset_growth': asset_growth_pct" in stock_processor_source)
    return results, summary


def validate_gui_sidebar_latest_preview_contract_case(_base_params):
    case_id = "GUI_SIDEBAR_LATEST_PREVIEW_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    inspector_source = build_project_absolute_path("tools", "workbench_ui", "single_stock_inspector.py").read_text(encoding="utf-8")
    charting_source = build_project_absolute_path("tools", "trade_analysis", "charting.py").read_text(encoding="utf-8")
    backtest_source = build_project_absolute_path("tools", "trade_analysis", "backtest.py").read_text(encoding="utf-8")

    add_check(results, "output_contract", case_id, "gui_panel_removes_execute_button", False, 'text="執行回測"' in inspector_source)
    add_check(results, "output_contract", case_id, "gui_panel_runs_backtest_on_enter", True, 'ticker_entry.bind("<Return>"' in inspector_source)
    add_check(results, "output_contract", case_id, "gui_panel_runs_backtest_on_candidate_select", True, 'self.after_idle(self._run_analysis)' in inspector_source)
    add_check(results, "output_contract", case_id, "gui_panel_declares_right_sidebar_signal_chip", True, '_signal_chip' in inspector_source and '歷史績效表' in inspector_source)
    add_check(results, "output_contract", case_id, "gui_panel_omits_html_button", False, '開啟 HTML K 線圖' in inspector_source)
    add_check(results, "output_contract", case_id, "gui_panel_declares_selected_date_value_block", True, '選取日線值' in inspector_source and '_selected_open_var' in inspector_source and '_selected_high_var' in inspector_source and '_selected_low_var' in inspector_source and '_selected_close_var' in inspector_source and '_selected_volume_var' in inspector_source)
    add_check(results, "output_contract", case_id, "charting_declares_future_preview_contract", True, 'future_preview' in charting_source and '_render_future_preview_lines' in charting_source)
    add_check(results, "output_contract", case_id, "charting_declares_dynamic_right_padding", True, 'extra_right_padding = max(' in charting_source and 'CHART_RIGHT_PADDING_RATIO' in charting_source)
    add_check(results, "output_contract", case_id, "backtest_uses_future_preview_for_latest_signal", True, 'set_chart_future_preview(' in backtest_source)
    add_check(results, "output_contract", case_id, "charting_reduces_latest_blank_space_to_about_one_sixth", True, "CHART_RIGHT_PADDING_RATIO = 0.1666666667" in charting_source)
    return results, summary


def validate_gui_single_stock_refined_visual_contract_case(_base_params):
    case_id = "GUI_SINGLE_STOCK_REFINED_VISUAL_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    inspector_source = build_project_absolute_path("tools", "workbench_ui", "single_stock_inspector.py").read_text(encoding="utf-8")
    charting_source = build_project_absolute_path("tools", "trade_analysis", "charting.py").read_text(encoding="utf-8")
    entry_flow_source = build_project_absolute_path("tools", "trade_analysis", "entry_flow.py").read_text(encoding="utf-8")
    workbench_source = build_project_absolute_path("tools", "workbench_ui", "workbench.py").read_text(encoding="utf-8")

    add_check(results, "output_contract", case_id, "single_stock_chart_uses_pure_black_bg", True, 'MATPLOTLIB_DARK_BG = "#000000"' in charting_source)
    add_check(results, "output_contract", case_id, "single_stock_sidebar_signal_chip_uses_fixed_text", True, 'self._sidebar_signal_var = tk.StringVar(value=SIDEBAR_SIGNAL_CHIP_TEXT)' in inspector_source or 'self._sidebar_signal_var = tk.StringVar(value="出現買入訊號")' in inspector_source)
    add_check(results, "output_contract", case_id, "single_stock_sidebar_gate_chip_uses_fixed_text", True, 'self._sidebar_history_var = tk.StringVar(value=SIDEBAR_HISTORY_CHIP_TEXT)' in inspector_source or 'self._sidebar_history_var = tk.StringVar(value="符合歷史績效")' in inspector_source)
    add_check(results, "output_contract", case_id, "single_stock_gui_removes_chart_hint_footer_label", False, 'textvariable=self._chart_hint_var' in inspector_source)
    add_check(results, "output_contract", case_id, "single_stock_gui_moves_ohlcv_to_sidebar", True, '_selected_open_var' in inspector_source and '_selected_volume_var' in inspector_source)
    add_check(results, "output_contract", case_id, "single_stock_hover_overlay_no_longer_repeats_top_left_text_box", True, 'hover_text_artist.set_visible(False)' in charting_source)
    add_check(results, "output_contract", case_id, "single_stock_sell_trade_label_includes_max_drawdown", True, '最大回撤:' in charting_source)
    add_check(results, "output_contract", case_id, "single_stock_entry_preview_lines_start_next_day_before_fill", True, '_record_entry_plan_preview_levels(' in entry_flow_source and 'record_active_levels(' in entry_flow_source)
    add_check(results, "output_contract", case_id, "single_stock_sidebar_fonts_enlarged", True, 'font=("Microsoft JhengHei", 16, "bold")' in workbench_source and 'font=("Microsoft JhengHei", 14)' in workbench_source)
    return results, summary


def validate_gui_extended_preview_continuity_contract_case(base_params):
    case_id = "GUI_EXTENDED_PREVIEW_CONTINUITY_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    inspector_source = build_project_absolute_path("tools", "workbench_ui", "single_stock_inspector.py").read_text(encoding="utf-8")
    entry_flow_source = build_project_absolute_path("tools", "trade_analysis", "entry_flow.py").read_text(encoding="utf-8")

    add_check(results, "output_contract", case_id, "single_stock_signal_chip_runtime_text_stays_fixed", True, 'self._sidebar_signal_var.set(SIDEBAR_SIGNAL_CHIP_TEXT)' in inspector_source and 'self._sidebar_signal_var.set(signal_text)' not in inspector_source)
    add_check(results, "output_contract", case_id, "single_stock_history_chip_runtime_text_stays_fixed", True, 'self._sidebar_history_var.set(SIDEBAR_HISTORY_CHIP_TEXT)' in inspector_source and 'self._sidebar_history_var.set(history_text)' not in inspector_source)
    add_check(results, "output_contract", case_id, "entry_preview_uses_candidate_layer_for_new_signals", True, 'preview_candidate_plan = build_normal_candidate_plan' in entry_flow_source)
    add_check(results, "output_contract", case_id, "entry_preview_uses_candidate_layer_for_extended_days", True, 'preview_candidate_plan = build_extended_candidate_plan_from_signal' in entry_flow_source)
    add_check(results, "output_contract", case_id, "entry_preview_extended_candidate_plan_stays_counterfactual_signature", True, 'build_extended_candidate_plan_from_signal(active_extended_signal, sizing_cap, params, ticker=ticker, security_profile=security_profile, trade_date=effective_trade_date)' in entry_flow_source)

    params = make_synthetic_validation_params(base_params)
    params.initial_capital = 1.0

    signal_dates = pd.date_range("2024-01-01", periods=7, freq="D")
    frame = pd.DataFrame(
        {
            "Open": [10.10, 10.20, 10.46, 10.44, 10.40, 10.36, 10.32],
            "High": [10.20, 10.35, 10.58, 10.55, 10.51, 10.47, 10.42],
            "Low": [10.00, 10.08, 10.45, 10.41, 10.37, 10.33, 10.29],
            "Close": [10.12, 10.30, 10.44, 10.40, 10.36, 10.32, 10.28],
            "Volume": [1000, 1000, 1000, 1000, 1000, 1000, 1000],
        },
        index=signal_dates,
    )
    precomputed_signals = (
        np.array([0.30, 0.30, 0.30, 0.30, 0.30, 0.30, 0.30], dtype=np.float64),
        np.array([False, True, False, False, False, False, False], dtype=bool),
        np.array([False, False, False, False, False, False, False], dtype=bool),
        np.array([np.nan, 10.50, np.nan, np.nan, np.nan, np.nan, np.nan], dtype=np.float64),
    )

    debug_module = importlib.import_module("tools.trade_analysis.trade_log")
    with tempfile.TemporaryDirectory(prefix="gui_extended_preview_continuity_contract_") as temp_dir:
        analysis_result = debug_module.run_debug_analysis(
            frame.copy(),
            "2330",
            params,
            export_excel=False,
            export_chart=False,
            return_chart_payload=True,
            verbose=False,
            precomputed_signals=precomputed_signals,
            output_dir=temp_dir,
        )
    payload = analysis_result.get("chart_payload") or {}
    limit_line = np.asarray(payload.get("limit_line", []), dtype=np.float64)
    stop_line = np.asarray(payload.get("stop_line", []), dtype=np.float64)
    tp_line = np.asarray(payload.get("tp_line", []), dtype=np.float64)
    continued_preview_days = [2, 3, 4, 5]

    add_check(results, "output_contract", case_id, "extended_preview_limit_line_continues_across_candidate_days", True, bool(limit_line.size > max(continued_preview_days) and np.isfinite(limit_line[continued_preview_days]).all()))
    add_check(results, "output_contract", case_id, "extended_preview_stop_line_stays_absent_before_fill", True, bool(stop_line.size > max(continued_preview_days) and np.isnan(stop_line[continued_preview_days]).all()))
    add_check(results, "output_contract", case_id, "extended_preview_tp_line_stays_absent_before_fill", True, bool(tp_line.size > max(continued_preview_days) and np.isnan(tp_line[continued_preview_days]).all()))
    add_check(results, "output_contract", case_id, "extended_preview_stays_visual_even_when_entry_plan_not_orderable", 0, len((payload.get("marker_groups") or {}).get("買進", [])))

    summary["continued_preview_days"] = [signal_dates[idx].strftime("%Y-%m-%d") for idx in continued_preview_days]
    return results, summary


def validate_gui_signal_annotation_and_forced_close_visual_contract_case(_base_params):
    case_id = "GUI_SIGNAL_ANNOTATION_AND_FORCED_CLOSE_VISUAL_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    charting_source = build_project_absolute_path("tools", "trade_analysis", "charting.py").read_text(encoding="utf-8")
    backtest_source = build_project_absolute_path("tools", "trade_analysis", "backtest.py").read_text(encoding="utf-8")

    sell_fn_start = backtest_source.index('def _record_sell_signal_annotation(')
    sell_fn_end = backtest_source.index('\n\ndef _apply_chart_sidebars', sell_fn_start)
    sell_fn_source = backtest_source[sell_fn_start:sell_fn_end]

    add_check(results, "output_contract", case_id, "single_stock_sell_signal_annotation_uses_title_only", True, "title='賣訊'" in sell_fn_source and 'detail_lines=[]' in sell_fn_source)
    add_check(results, "output_contract", case_id, "single_stock_sell_signal_annotation_omits_executed_sell_fields", True, '訊號日收盤:' not in sell_fn_source and '僅代表賣訊，不代表已成交' not in sell_fn_source and '股數:' not in sell_fn_source and '金額:' not in sell_fn_source and '損益:' not in sell_fn_source and '報酬率:' not in sell_fn_source)
    add_check(results, "output_contract", case_id, "forced_close_marker_uses_yellow_visual", True, '"期末強制結算": {"plotly_symbol": "square", "mpl_marker": "s", "color": "#facc15"}' in charting_source)

    dates = pd.date_range("2024-01-01", periods=4, freq="D")
    chart_payload = normalize_chart_payload_contract(
        {
            "dates": dates,
            "x": np.arange(4, dtype=np.float32),
            "open": np.array([100.0, 101.0, 102.0, 103.0], dtype=np.float32),
            "high": np.array([101.0, 102.0, 103.0, 104.0], dtype=np.float32),
            "low": np.array([99.0, 100.0, 101.0, 102.0], dtype=np.float32),
            "close": np.array([100.5, 101.5, 102.5, 103.5], dtype=np.float32),
            "volume": np.array([1000.0, 1000.0, 1000.0, 1000.0], dtype=np.float32),
            "stop_line": np.full(4, np.nan, dtype=np.float32),
            "tp_line": np.full(4, np.nan, dtype=np.float32),
            "limit_line": np.full(4, np.nan, dtype=np.float32),
            "entry_line": np.full(4, np.nan, dtype=np.float32),
            "marker_groups": {},
            "signal_annotations": [],
            "future_preview": {"limit_price": 103.0, "stop_price": 99.0, "tp_half_price": 132.0, "entry_price": 103.0},
            "summary_box": [],
            "status_box": {"lines": [], "ok": True},
            "default_view": {"start_idx": 0, "end_idx": 3},
        }
    )
    visible_ranges = compute_visible_value_ranges(chart_payload, start_idx=0.0, end_idx=4.8)
    add_check(results, "output_contract", case_id, "future_preview_tp_participates_in_visible_price_range", True, float(visible_ranges.get("price_max", 0.0)) > 130.0)

    summary["future_preview_price_max"] = float(visible_ranges.get("price_max", 0.0))
    return results, summary






def validate_gui_buy_signal_annotation_anchor_price_contract_case(base_params):
    case_id = "GUI_BUY_SIGNAL_ANNOTATION_ANCHOR_PRICE_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    backtest_source = build_project_absolute_path("tools", "trade_analysis", "backtest.py").read_text(encoding="utf-8")
    entry_flow_source = build_project_absolute_path("tools", "trade_analysis", "entry_flow.py").read_text(encoding="utf-8")
    add_check(results, "output_contract", case_id, "buy_signal_annotation_anchor_keeps_signal_low_even_when_entry_plan_exists", True, "anchor_price = float(signal_low)" in backtest_source and "anchor_price = float(entry_plan['limit_price'])" not in backtest_source)
    add_check(results, "output_contract", case_id, "buy_trade_marker_uses_fill_price", True, 'action="買進"' in entry_flow_source and "price=entry_result['buy_price']" in entry_flow_source)

    case = build_synthetic_competing_candidates_case(base_params, make_synthetic_validation_params)
    frame = case["frames"][case["primary_ticker"]].copy()
    params = case["params"]
    chart_context = create_debug_chart_context(frame)
    entry_plan = {
        "qty": 1000,
        "limit_price": 22.2,
        "init_sl": 20.65,
    }
    history_snapshot = {"current_capital": 2110780.0}
    expected_signal_low = float(frame.iloc[2]["Low"])
    _record_buy_signal_annotation(
        chart_context=chart_context,
        signal_date=frame.index[2],
        signal_low=expected_signal_low,
        entry_plan=entry_plan,
        history_snapshot=history_snapshot,
        params=params,
    )
    annotation = (chart_context.get("signal_annotations") or [None])[0]
    add_check(results, "output_contract", case_id, "buy_signal_annotation_anchor_matches_signal_low", expected_signal_low, None if annotation is None else float(annotation.get("anchor_price")))
    add_check(results, "output_contract", case_id, "buy_signal_annotation_meta_keeps_entry_price_preview", float(entry_plan["limit_price"]), None if annotation is None else float((annotation.get("meta") or {}).get("entry_price", np.nan)))

    from tools.trade_analysis.charting import record_trade_marker
    fill_price = 21.85
    record_trade_marker(
        chart_context,
        current_date=frame.index[2],
        action="買進",
        price=fill_price,
        qty=entry_plan["qty"],
        meta={"entry_price": fill_price},
    )
    trade_marker = (chart_context.get("trade_markers") or [None])[0]
    add_check(results, "output_contract", case_id, "buy_trade_marker_price_matches_fill_price", fill_price, None if trade_marker is None else float(trade_marker.get("price")))

    _record_buy_signal_annotation(
        chart_context=chart_context,
        signal_date=frame.index[3],
        signal_low=float(frame.iloc[3]["Low"]),
        entry_plan=None,
        history_snapshot=history_snapshot,
        params=params,
    )
    fallback_annotation = (chart_context.get("signal_annotations") or [None, None])[1]
    expected_fallback_signal_low = float(frame.iloc[3]["Low"])
    add_check(results, "output_contract", case_id, "buy_signal_annotation_without_entry_plan_falls_back_to_signal_low", expected_fallback_signal_low, None if fallback_annotation is None else float(fallback_annotation.get("anchor_price")))

    summary["entry_plan_signal_anchor_price"] = None if annotation is None else float(annotation.get("anchor_price"))
    summary["buy_trade_marker_price"] = None if trade_marker is None else float(trade_marker.get("price"))
    summary["fallback_signal_low_anchor_price"] = None if fallback_annotation is None else float(fallback_annotation.get("anchor_price"))
    return results, summary
def validate_gui_trade_marker_and_tp_visual_contract_case(_base_params):
    case_id = "GUI_TRADE_MARKER_AND_TP_VISUAL_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    charting_source = build_project_absolute_path("tools", "trade_analysis", "charting.py").read_text(encoding="utf-8")
    exit_flow_source = build_project_absolute_path("tools", "trade_analysis", "exit_flow.py").read_text(encoding="utf-8")
    backtest_source = build_project_absolute_path("tools", "trade_analysis", "backtest.py").read_text(encoding="utf-8")

    add_check(results, "output_contract", case_id, "indicator_sell_marker_uses_green_horizontal_line", True, '"指標賣出": {"plotly_symbol": "line-ew-open", "mpl_marker": "_", "color": MATPLOTLIB_INDICATOR_SELL_COLOR}' in charting_source)
    add_check(results, "output_contract", case_id, "tp_visual_uses_yellow_color", True, 'MATPLOTLIB_TP_COLOR = "#facc15"' in charting_source)
    add_check(results, "output_contract", case_id, "exit_trade_labels_include_trade_count", True, 'lines.append(f"交易次數: {int(trade_count)}")' in charting_source and "'trade_count': None if history_snapshot is None else int(history_snapshot.get('trade_count', 0) or 0)" in exit_flow_source)
    add_check(results, "output_contract", case_id, "buy_signal_annotation_exports_limit_and_reserved_for_autoscale", True, "'limit_price': float(entry_plan['limit_price'])" in backtest_source and "'reserved_capital': float(reserved_capital)" in backtest_source and "'tp_price': float(tp_line)" not in backtest_source)

    dates = pd.date_range("2024-01-01", periods=4, freq="D")
    chart_payload = normalize_chart_payload_contract(
        {
            "dates": dates,
            "x": np.arange(4, dtype=np.float32),
            "open": np.array([100.0, 101.0, 102.0, 103.0], dtype=np.float32),
            "high": np.array([101.0, 102.0, 103.0, 104.0], dtype=np.float32),
            "low": np.array([99.0, 100.0, 101.0, 102.0], dtype=np.float32),
            "close": np.array([100.5, 101.5, 102.5, 103.5], dtype=np.float32),
            "volume": np.array([1000.0, 1000.0, 1000.0, 1000.0], dtype=np.float32),
            "stop_line": np.full(4, np.nan, dtype=np.float32),
            "tp_line": np.full(4, np.nan, dtype=np.float32),
            "limit_line": np.full(4, np.nan, dtype=np.float32),
            "entry_line": np.full(4, np.nan, dtype=np.float32),
            "marker_groups": {},
            "signal_annotations": [
                {
                    "date": dates[3],
                    "x": 3,
                    "anchor_price": 103.0,
                    "signal_type": "buy",
                    "title": "買訊",
                    "detail_text": "預留: 132,000",
                    "meta": {"limit_price": 132.0, "reserved_capital": 132000.0},
                }
            ],
            "future_preview": {},
            "summary_box": [],
            "status_box": {"lines": [], "ok": True},
            "default_view": {"start_idx": 0, "end_idx": 3},
        }
    )
    visible_ranges = compute_visible_value_ranges(chart_payload, start_idx=0.0, end_idx=3.0)
    add_check(results, "output_contract", case_id, "buy_signal_limit_preview_participates_in_visible_price_range", True, float(visible_ranges.get("price_max", 0.0)) > 130.0)

    summary["price_max_with_signal_meta"] = float(visible_ranges.get("price_max", 0.0))
    return results, summary




def validate_gui_trade_box_capital_and_round_trip_contract_case(_base_params):
    case_id = "GUI_TRADE_BOX_CAPITAL_AND_ROUND_TRIP_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    charting_source = build_project_absolute_path("tools", "trade_analysis", "charting.py").read_text(encoding="utf-8")
    backtest_source = build_project_absolute_path("tools", "trade_analysis", "backtest.py").read_text(encoding="utf-8")
    exit_flow_source = build_project_absolute_path("tools", "trade_analysis", "exit_flow.py").read_text(encoding="utf-8")
    trade_log_source = build_project_absolute_path("tools", "trade_analysis", "trade_log.py").read_text(encoding="utf-8")
    inspector_source = build_project_absolute_path("tools", "workbench_ui", "single_stock_inspector.py").read_text(encoding="utf-8")

    add_check(results, "output_contract", case_id, "debug_view_initial_capital_uses_scanner_live_capital_basis", True, 'cloned_params.initial_capital = resolve_scanner_live_capital(cloned_params)' in trade_log_source)
    add_check(results, "output_contract", case_id, "buy_signal_annotation_includes_current_capital_and_reserved_capital", True, '資金:' in backtest_source and '預留:' in backtest_source)
    add_check(results, "output_contract", case_id, "final_exit_marker_uses_completed_snapshot_after_same_day_exits", True, 'include_current_date_exits=True' in exit_flow_source)

    charting_module = importlib.import_module("tools.trade_analysis.charting")
    build_trade_label_text = getattr(charting_module, "_build_trade_label_text")

    buy_label_text = build_trade_label_text(
        "買進",
        {
            "qty": 1000,
            "meta": {
                "limit_price": 56.0,
                "buy_capital": 123456.0,
                "tp_price": 62.4,
                "entry_price": 55.2,
                "stop_price": 47.65,
            },
        },
    )
    tp_label_text = build_trade_label_text(
        "半倉停利",
        {
            "qty": 520,
            "meta": {
                "sell_capital": 32100.0,
                "pnl_value": 4567.0,
            },
        },
    )
    exit_label_text = build_trade_label_text(
        "停損殺出",
        {
            "qty": 480,
            "meta": {
                "current_capital": 2012345.0,
                "sell_capital": 28000.0,
                "pnl_value": -1234.0,
                "total_pnl": 3333.0,
                "pnl_pct": 4.56,
                "payoff_ratio": 3.21,
                "win_rate": 66.6,
                "expected_value": 0.78,
                "max_drawdown": 5.43,
                "trade_count": 11,
            },
        },
    )

    add_check(results, "output_contract", case_id, "buy_trade_label_uses_trade_info_wording", True, "停利: 62.40" in buy_label_text and "限價: 56.00" in buy_label_text and "成交: 55.20" in buy_label_text and "停損: 47.65" in buy_label_text and "實支: 123,456" in buy_label_text)
    add_check(results, "output_contract", case_id, "buy_trade_label_includes_actual_spend_wording_last", "實支: 123,456", buy_label_text.split("\n")[-1] if buy_label_text else "")
    add_check(results, "output_contract", case_id, "buy_trade_label_overlay_includes_buy_traces", True, 'supported_traces = {"買進", "買進(延續候選)", "半倉停利", "停損殺出", "指標賣出", "期末強制結算"}' in charting_source)
    add_check(results, "output_contract", case_id, "buy_trade_label_overlay_places_buy_box_below_kbar", True, '_resolve_trade_label_offsets(slot_index, placement=placement, trace_name=trace_name)' in charting_source and 'base_y = -64' in charting_source and 'va = "top" if placement == "below" else "bottom"' in charting_source)
    add_check(results, "output_contract", case_id, "buy_trade_label_overlay_offsets_when_signal_or_nearby_trade_boxes_cluster", True, '_count_nearby_annotation_slots(x_value, placement, occupied_positions, collision_window=collision_window)' in charting_source and 'collision_window = 5 if placement == "below" else 4' in charting_source and 'occupied_positions.append({"x": x_value, "placement": placement})' in charting_source)
    add_check(results, "output_contract", case_id, "single_stock_sidebar_trade_info_includes_actual_spend_field", True, 'text="交易資訊"' in inspector_source and '_selected_tp_var' in inspector_source and '_selected_limit_var' in inspector_source and '_selected_entry_var' in inspector_source and '_selected_stop_var' in inspector_source and '_selected_actual_spend_var' in inspector_source)
    add_check(results, "output_contract", case_id, "single_stock_sidebar_sets_actual_spend_from_hover_snapshot", True, 'self._selected_actual_spend_var.set(self._format_sidebar_amount_value("實支", snapshot.get("buy_capital")))' in inspector_source and '"buy_capital": buy_trade_meta.get("buy_capital")' in charting_source)
    add_check(results, "output_contract", case_id, "tp_trade_label_includes_sell_amount_and_leg_pnl", True, "金額: 32,100" in tp_label_text and "損益: +4,567" in tp_label_text)
    add_check(results, "output_contract", case_id, "exit_trade_label_includes_total_pnl_and_trade_count_last", "交易次數: 11", exit_label_text.split("\n")[-1] if exit_label_text else "")
    add_check(results, "output_contract", case_id, "exit_trade_label_includes_current_capital_and_total_pnl", True, "資金: 2,012,345" in exit_label_text and "總損益: +3,333" in exit_label_text)

    summary["buy_label"] = buy_label_text
    summary["tp_label"] = tp_label_text
    summary["exit_label_last_line"] = exit_label_text.split("\n")[-1] if exit_label_text else ""
    return results, summary
def validate_gui_trade_count_and_sidebar_sync_contract_case(base_params):
    case_id = "GUI_TRADE_COUNT_AND_SIDEBAR_SYNC_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    inspector_source = build_project_absolute_path("tools", "workbench_ui", "single_stock_inspector.py").read_text(encoding="utf-8")
    charting_source = build_project_absolute_path("tools", "trade_analysis", "charting.py").read_text(encoding="utf-8")
    add_check(results, "output_contract", case_id, "single_stock_sidebar_declares_trade_info_block", True, 'text="交易資訊"' in inspector_source and '_selected_tp_var' in inspector_source and '_selected_entry_var' in inspector_source and '_selected_stop_var' in inspector_source)
    add_check(results, "output_contract", case_id, "single_stock_sidebar_omits_buy_signal_info_block", True, 'text="買訊資訊"' not in inspector_source and '_selected_reserved_var' not in inspector_source and '_selected_signal_capital_var' not in inspector_source)
    add_check(results, "output_contract", case_id, "chart_hover_snapshot_keeps_trade_line_values_for_sidebar", True, '"tp_price": float(chart_payload["tp_line"][idx])' in charting_source and '"limit_price": float(chart_payload["limit_line"][idx])' in charting_source and '"entry_price": float(chart_payload["entry_line"][idx])' in charting_source and '"stop_price": float(chart_payload["stop_line"][idx])' in charting_source)

    params = make_synthetic_validation_params(base_params)
    forced_close_date = pd.Timestamp("2024-01-03")
    stats_index = build_trade_stats_index([
        {
            "exit_date": pd.Timestamp("2024-01-02"),
            "pnl": 1000.0,
            "r_mult": 1.0,
        },
        {
            "exit_date": forced_close_date,
            "pnl": 500.0,
            "r_mult": 0.5,
        },
    ])
    trade_logs = []
    chart_context = {"trade_markers": []}
    append_debug_forced_closeout(
        position={
            "entry": 100.0,
            "qty": 100,
            "initial_qty": 100,
            "realized_pnl": 0.0,
            "close_price": 105.0,
            "sl": 95.0,
        },
        current_date=forced_close_date,
        atr_last=1.5,
        params=params,
        trade_logs=trade_logs,
        chart_context=chart_context,
        current_capital_before_event=float(params.initial_capital),
        stats_index=stats_index,
        overall_max_drawdown=6.78,
    )
    forced_close_marker = chart_context["trade_markers"][-1] if chart_context.get("trade_markers") else {}
    forced_close_meta = forced_close_marker.get("meta") or {}
    expected_trade_count = int(stats_index["cum_trade_count"][-1]) if len(stats_index.get("cum_trade_count", [])) > 0 else 0
    add_check(results, "output_contract", case_id, "final_exit_trade_count_uses_completed_round_trip_index", expected_trade_count, forced_close_meta.get("trade_count"))
    add_check(results, "output_contract", case_id, "final_exit_marker_action_is_forced_close", "期末強制結算", forced_close_marker.get("trace_name"))

    charting_module = importlib.import_module("tools.trade_analysis.charting")
    build_trade_label_text = getattr(charting_module, "_build_trade_label_text")
    exit_label_text = build_trade_label_text(
        "停損殺出",
        {
            "qty": 100,
            "meta": {
                "sell_capital": 12345.0,
                "pnl_value": -321.0,
                "pnl_pct": -1.23,
                "payoff_ratio": 4.56,
                "win_rate": 78.9,
                "expected_value": 2.34,
                "max_drawdown": 6.78,
                "trade_count": 14,
            },
        },
    )
    exit_lines = exit_label_text.split("\n")
    add_check(results, "output_contract", case_id, "exit_trade_label_places_trade_count_last", "交易次數: 14", exit_lines[-1] if exit_lines else "")
    add_check(results, "output_contract", case_id, "exit_trade_label_trade_count_matches_latest_completed_trade", True, "交易次數: 14" in exit_label_text)

    summary["expected_trade_count"] = expected_trade_count
    summary["forced_close_trade_count"] = forced_close_meta.get("trade_count")
    summary["exit_label_last_line"] = exit_lines[-1] if exit_lines else ""
    return results, summary
    return results, summary


def validate_gui_chart_margin_and_latest_extended_preview_contract_case(base_params):
    case_id = "GUI_CHART_MARGIN_AND_LATEST_EXTENDED_PREVIEW_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    charting_source = build_project_absolute_path("tools", "trade_analysis", "charting.py").read_text(encoding="utf-8")
    backtest_source = build_project_absolute_path("tools", "trade_analysis", "backtest.py").read_text(encoding="utf-8")

    add_check(results, "output_contract", case_id, "single_stock_chart_tightens_left_bottom_margins_for_more_candle_space", True, 'MATPLOTLIB_SUBPLOT_LEFT = 0.046' in charting_source and 'MATPLOTLIB_SUBPLOT_BOTTOM = 0.058' in charting_source and 'axis_price.tick_params(axis="y", colors=MATPLOTLIB_TEXT_COLOR, labelsize=11, pad=6)' in charting_source)
    add_check(results, "output_contract", case_id, "single_stock_legend_keeps_small_inset_gap_from_top_left_boundary", True, 'bbox_to_anchor=(0.012, 1.012)' in charting_source and 'MATPLOTLIB_SUBPLOT_TOP = 0.986' in charting_source)
    add_check(results, "output_contract", case_id, "single_stock_latest_extended_candidate_uses_future_preview_helper", True, 'build_extended_candidate_plan_from_signal(' in backtest_source and 'latest_extended_preview = build_extended_candidate_plan_from_signal(' in backtest_source and 'security_profile=resolved_security_profile' in backtest_source and 'trade_date=dates[-1]' in backtest_source and '_apply_chart_future_preview_from_plan(chart_context, latest_extended_preview)' in backtest_source)

    params = make_synthetic_validation_params(base_params)
    params.initial_capital = 2_000_000.0

    signal_dates = pd.date_range("2024-01-01", periods=7, freq="D")
    frame = pd.DataFrame(
        {
            "Open": [10.10, 10.20, 10.55, 10.56, 10.57, 10.58, 10.59],
            "High": [10.18, 10.36, 10.60, 10.62, 10.63, 10.64, 10.65],
            "Low": [10.02, 10.12, 10.52, 10.53, 10.54, 10.55, 10.56],
            "Close": [10.12, 10.30, 10.56, 10.57, 10.58, 10.59, 10.60],
            "Volume": [1000, 1000, 1000, 1000, 1000, 1000, 1000],
        },
        index=signal_dates,
    )
    precomputed_signals = (
        np.array([0.30, 0.30, 0.30, 0.30, 0.30, 0.30, 0.30], dtype=np.float64),
        np.array([False, True, False, False, False, False, False], dtype=bool),
        np.array([False, False, False, False, False, False, False], dtype=bool),
        np.array([np.nan, 10.50, np.nan, np.nan, np.nan, np.nan, np.nan], dtype=np.float64),
    )

    debug_module = importlib.import_module("tools.trade_analysis.trade_log")
    backtest_module = importlib.import_module("tools.trade_analysis.backtest")
    forced_history_snapshot = {
        'trade_count': 12,
        'win_rate': 58.0,
        'expected_value': 0.35,
        'payoff_ratio': 1.6,
        'is_candidate': True,
        'asset_growth_pct': 8.0,
        'max_drawdown': 4.2,
    }
    with tempfile.TemporaryDirectory(prefix="gui_chart_margin_latest_extended_preview_contract_") as temp_dir:
        with patch.object(backtest_module, '_build_pit_history_snapshot', return_value=forced_history_snapshot):
            analysis_result = debug_module.run_debug_analysis(
                frame.copy(),
                "2330",
                params,
                export_excel=False,
                export_chart=False,
                return_chart_payload=True,
                verbose=False,
                precomputed_signals=precomputed_signals,
                output_dir=temp_dir,
            )
    payload = analysis_result.get("chart_payload") or {}
    future_preview = dict(payload.get("future_preview") or {})
    limit_line = np.asarray(payload.get("limit_line", []), dtype=np.float64)
    stop_line = np.asarray(payload.get("stop_line", []), dtype=np.float64)
    tp_line = np.asarray(payload.get("tp_line", []), dtype=np.float64)

    add_check(results, "output_contract", case_id, "extended_preview_limit_line_reaches_latest_actual_bar", True, bool(limit_line.size > 6 and np.isfinite(limit_line[6])))
    add_check(results, "output_contract", case_id, "extended_preview_stop_line_stays_absent_at_latest_actual_bar", True, bool(stop_line.size > 6 and np.isnan(stop_line[6])))
    add_check(results, "output_contract", case_id, "extended_preview_tp_line_stays_absent_at_latest_actual_bar", True, bool(tp_line.size > 6 and np.isnan(tp_line[6])))
    add_check(results, "output_contract", case_id, "latest_extended_candidate_future_preview_limit_exists", True, bool(np.isfinite(float(future_preview.get("limit_price", np.nan)))))
    add_check(results, "output_contract", case_id, "latest_extended_candidate_future_preview_stop_absent_before_fill", True, bool(np.isnan(float(future_preview.get("stop_price", np.nan)))))
    add_check(results, "output_contract", case_id, "latest_extended_candidate_future_preview_tp_absent_before_fill", True, bool(np.isnan(float(future_preview.get("tp_half_price", np.nan)))))

    summary["latest_future_preview"] = {
        key: float(value) for key, value in future_preview.items() if value is not None and np.isfinite(value)
    }
    return results, summary


def validate_gui_latest_raw_signal_preview_helper_contract_case(_base_params):
    case_id = "GUI_LATEST_RAW_SIGNAL_PREVIEW_HELPER_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    backtest_path = build_project_absolute_path("tools", "trade_analysis", "backtest.py")
    backtest_source = backtest_path.read_text(encoding="utf-8")
    backtest_ast = ast.parse(backtest_source)

    imported_entry_plan_symbols = set()
    referenced_names = set()
    for node in ast.walk(backtest_ast):
        if isinstance(node, ast.ImportFrom) and node.module == "core.entry_plans":
            imported_entry_plan_symbols.update(alias.name for alias in node.names)
        elif isinstance(node, ast.Name):
            referenced_names.add(node.id)

    add_check(results, "output_contract", case_id, "latest_raw_signal_preview_uses_normal_entry_plan_helper", True, "build_normal_entry_plan" in backtest_source and "entry_plan_preview = build_normal_entry_plan(" in backtest_source and "security_profile=resolved_security_profile" in backtest_source)
    add_check(results, "output_contract", case_id, "latest_raw_signal_preview_helper_is_imported_from_core_entry_plans", True, "build_normal_entry_plan" in referenced_names and "build_normal_entry_plan" in imported_entry_plan_symbols)
    add_check(results, "output_contract", case_id, "latest_tail_preview_candidate_helper_is_imported_from_core_entry_plans", True, "build_normal_candidate_plan" in referenced_names and "build_normal_candidate_plan" in imported_entry_plan_symbols)
    add_check(results, "output_contract", case_id, "latest_tail_preview_threads_security_profile", True, "latest_entry_plan_preview = build_normal_candidate_plan(" in backtest_source and "security_profile=resolved_security_profile" in backtest_source)
    return results, summary

def validate_gui_chart_overlay_layout_and_pan_contract_case(_base_params):
    case_id = "GUI_CHART_OVERLAY_LAYOUT_AND_PAN_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    inspector_source = build_project_absolute_path("tools", "workbench_ui", "single_stock_inspector.py").read_text(encoding="utf-8")
    workbench_source = build_project_absolute_path("tools", "workbench_ui", "workbench.py").read_text(encoding="utf-8")
    charting_source = build_project_absolute_path("tools", "trade_analysis", "charting.py").read_text(encoding="utf-8")

    add_check(
        results,
        "output_contract",
        case_id,
        "gui_panel_uses_explicit_workbench_dark_styles",
        True,
        'style="Workbench.TNotebook"' in inspector_source
        and 'style="Workbench.TFrame"' in inspector_source
        and 'style="Workbench.TLabel"' in inspector_source,
    )
    add_check(results, "output_contract", case_id, "gui_workbench_uses_palette_force_for_dark_theme", True, "root.tk_setPalette(" in workbench_source)
    add_check(results, "output_contract", case_id, "chart_mouse_pan_uses_pixel_anchor", True, "anchor_px" in charting_source and "bars_per_pixel = (float(origin_right) - float(origin_left)) / axis_width_px" in charting_source and 'interaction_flags["dragging"]' in charting_source)

    dates = pd.date_range("2024-01-01", periods=80, freq="B")
    payload = {
        "dates": dates,
        "date_labels": [dt.strftime("%Y-%m-%d") for dt in dates],
        "x": np.arange(len(dates), dtype=np.float64),
        "open": np.linspace(30.0, 38.0, len(dates), dtype=np.float64),
        "high": np.linspace(30.8, 38.8, len(dates), dtype=np.float64),
        "low": np.linspace(29.4, 37.4, len(dates), dtype=np.float64),
        "close": np.linspace(30.3, 38.3, len(dates), dtype=np.float64),
        "volume": np.linspace(1000.0, 3000.0, len(dates), dtype=np.float64),
        "up_mask": np.array([idx % 2 == 0 for idx in range(len(dates))], dtype=bool),
        "stop_line": np.full(len(dates), np.nan, dtype=np.float64),
        "tp_line": np.full(len(dates), np.nan, dtype=np.float64),
        "limit_line": np.full(len(dates), np.nan, dtype=np.float64),
        "entry_line": np.full(len(dates), np.nan, dtype=np.float64),
        "marker_groups": {
            "買進": [{"trace_name": "買進", "date": dates[20], "x": 20, "price": 32.1, "qty": 1000, "note": "", "hover_text": "買進", "meta": {"limit_price": 32.0, "entry_price": 32.1}}],
            "指標賣出": [{"trace_name": "指標賣出", "date": dates[30], "x": 30, "price": 34.2, "qty": 1000, "note": "本次績效: -2.00%", "hover_text": "賣出", "meta": {"pnl_pct": -2.0, "pnl_value": -2000.0}}],
        },
        "signal_annotations": [{"date": dates[20], "x": 20, "anchor_price": 31.8, "signal_type": "buy", "title": "買訊", "detail_text": "限價: 32.00", "note": "", "meta": {}}],
        "summary_box": ["資產成長: 10.0%"],
        "status_box": {"lines": ["出現買入訊號", "歷史績效符合"], "ok": True},
        "focus_positions": [20, 30],
        "default_view": {"start_idx": 10, "end_idx": 45},
    }
    figure = create_matplotlib_trade_chart_figure(chart_payload=payload, ticker="2330", show_volume=True)
    contract = getattr(figure, "_stock_chart_contract", {})
    state = getattr(figure, "_stock_chart_navigation_state", {})
    add_check(results, "output_contract", case_id, "figure_contract_status_chip_layout_right_bottom", "right_bottom", contract.get("status_chip_layout"))
    add_check(results, "output_contract", case_id, "figure_contract_buy_trade_label_boxes_enabled", True, contract.get("buy_trade_label_boxes_enabled"))
    add_check(results, "output_contract", case_id, "figure_contract_mouse_drag_pan_mode_pixel_anchor", "pixel_anchor", contract.get("mouse_drag_pan_mode"))
    add_check(results, "output_contract", case_id, "figure_contract_grid_alpha_subtle", True, float(contract.get("grid_alpha", 1.0)) <= 0.40)
    add_check(results, "output_contract", case_id, "figure_trade_labels_include_buy_and_sell_markers", 2, len(state.get("trade_label_artists") or []))
    figure.clear()
    return results, summary


def validate_gui_chart_recent_view_signal_overlay_contract_case(base_params):
    case_id = "GUI_CHART_RECENT_VIEW_SIGNAL_OVERLAY_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    inspector_source = build_project_absolute_path("tools", "workbench_ui", "single_stock_inspector.py").read_text(encoding="utf-8")
    add_check(results, "output_contract", case_id, "gui_inline_chart_skips_html_export_by_default", True, "export_chart=False" in inspector_source and "return_chart_payload=True" in inspector_source)

    large_dates = pd.date_range("2018-01-01", periods=1800, freq="B")
    large_frame = pd.DataFrame(
        {
            "Open": np.linspace(50.0, 150.0, len(large_dates)),
            "High": np.linspace(51.0, 151.0, len(large_dates)),
            "Low": np.linspace(49.0, 149.0, len(large_dates)),
            "Close": np.linspace(50.5, 150.5, len(large_dates)),
            "Volume": np.linspace(5000.0, 25000.0, len(large_dates)),
        },
        index=large_dates,
    )
    large_chart_context = create_debug_chart_context(large_frame)
    record_signal_annotation(
        large_chart_context,
        current_date=large_dates[-15],
        signal_type="buy",
        anchor_price=float(large_frame.iloc[-15]["Low"]),
        title="買訊",
        detail_lines=["股數: 1,000", "限價: 120.00", "預留: 120,000"],
    )
    set_chart_summary_box(large_chart_context, summary_lines=["資產成長: 10.0%", "交易次數: 5"])
    set_chart_status_box(large_chart_context, status_lines=["無賣訊", "出現買入訊號", "歷史績效符合"], ok=True)
    chart_payload = build_debug_chart_payload(large_frame, large_chart_context)
    figure = create_matplotlib_trade_chart_figure(chart_payload=chart_payload, ticker="LARGE", show_volume=True)
    contract = getattr(figure, "_stock_chart_contract", {})
    default_view = contract.get("default_view", {})

    add_check(results, "output_contract", case_id, "gui_chart_default_recent_view_ends_at_last_bar", len(chart_payload["x"]) - 1, int(default_view.get("end_idx", -1)))
    add_check(results, "output_contract", case_id, "gui_chart_default_recent_view_narrower_than_full_history", True, int(default_view.get("start_idx", 0)) > 0)
    add_check(results, "output_contract", case_id, "gui_chart_contract_default_lookback_months", 6, contract.get("default_lookback_months"))
    add_check(results, "output_contract", case_id, "gui_chart_contract_hover_value_display_enabled", True, contract.get("hover_value_display_enabled"))
    add_check(results, "output_contract", case_id, "gui_chart_contract_twse_up_color", "#ff5b6e", contract.get("twse_up_color"))
    add_check(results, "output_contract", case_id, "gui_chart_contract_twse_down_color", "#18b26b", contract.get("twse_down_color"))
    add_check(results, "output_contract", case_id, "gui_chart_contract_summary_box_present", True, contract.get("summary_box_present"))
    add_check(results, "output_contract", case_id, "gui_chart_contract_status_box_present", True, contract.get("status_box_present"))
    add_check(results, "output_contract", case_id, "gui_chart_contract_signal_annotation_present", True, contract.get("signal_annotation_count", 0) >= 1)
    figure.clear()

    case = build_synthetic_competing_candidates_case(base_params, make_synthetic_validation_params)
    params = case["params"]
    signal_dates = pd.date_range("2024-01-01", periods=6, freq="D")
    signal_frame = pd.DataFrame(
        {
            "Open": [10.0, 10.2, 10.8, 11.2, 10.7, 10.5],
            "High": [10.4, 10.8, 11.3, 11.4, 10.9, 10.7],
            "Low": [9.8, 10.0, 10.1, 10.9, 10.2, 10.0],
            "Close": [10.1, 10.7, 11.2, 11.0, 10.4, 10.3],
            "Volume": [1000, 1200, 1500, 1400, 1300, 1600],
        },
        index=signal_dates,
    )
    precomputed_signals = (
        np.array([1.0, 1.0, 1.0, 1.0, 1.0, 1.0], dtype=np.float64),
        np.array([False, True, False, False, False, False], dtype=bool),
        np.array([False, False, False, True, False, False], dtype=bool),
        np.array([np.nan, 10.9, np.nan, np.nan, np.nan, np.nan], dtype=np.float64),
    )
    debug_module = importlib.import_module("tools.trade_analysis.trade_log")
    with tempfile.TemporaryDirectory(prefix="gui_chart_signal_overlay_contract_") as temp_dir:
        analysis_result = debug_module.run_debug_analysis(
            signal_frame.copy(),
            "2330",
            params,
            export_excel=False,
            export_chart=False,
            return_chart_payload=True,
            verbose=False,
            precomputed_signals=precomputed_signals,
            output_dir=temp_dir,
        )
    payload = analysis_result.get("chart_payload") or {}
    signal_annotations = payload.get("signal_annotations", [])
    buy_signal_dates = [item["date"].strftime("%Y-%m-%d") for item in signal_annotations if item.get("signal_type") == "buy"]
    sell_signal_dates = [item["date"].strftime("%Y-%m-%d") for item in signal_annotations if item.get("signal_type") == "sell"]
    marker_groups = payload.get("marker_groups", {})
    buy_trade_dates = [item["date"].strftime("%Y-%m-%d") for item in marker_groups.get("買進", [])]
    sell_trade_dates = [item["date"].strftime("%Y-%m-%d") for item in marker_groups.get("指標賣出", [])]

    add_check(results, "output_contract", case_id, "gui_chart_buy_signal_annotation_precedes_execution_by_one_bar", signal_dates[1].strftime("%Y-%m-%d"), buy_signal_dates[0] if buy_signal_dates else "")
    add_check(results, "output_contract", case_id, "gui_chart_buy_execution_occurs_on_next_bar", signal_dates[2].strftime("%Y-%m-%d"), buy_trade_dates[0] if buy_trade_dates else "")
    add_check(results, "output_contract", case_id, "gui_chart_sell_signal_annotation_precedes_execution_by_one_bar", signal_dates[3].strftime("%Y-%m-%d"), sell_signal_dates[0] if sell_signal_dates else "")
    add_check(results, "output_contract", case_id, "gui_chart_sell_execution_occurs_on_next_bar", signal_dates[4].strftime("%Y-%m-%d"), sell_trade_dates[0] if sell_trade_dates else "")

    summary["signal_annotation_count"] = len(signal_annotations)
    return results, summary


def validate_debug_entry_plan_marker_optional_contract_case(_base_params):
    case_id = "DEBUG_ENTRY_PLAN_MARKER_OPTIONAL_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    dates = pd.date_range("2024-01-01", periods=6, freq="D")
    frame = pd.DataFrame(
        {
            "Open": np.linspace(10.0, 15.0, len(dates)),
            "High": np.linspace(11.0, 16.0, len(dates)),
            "Low": np.linspace(9.0, 14.0, len(dates)),
            "Close": np.linspace(10.5, 15.5, len(dates)),
            "Volume": np.full(len(dates), 1000.0),
        },
        index=dates,
    )
    chart_context = create_debug_chart_context(frame)
    _record_entry_plan_marker(
        chart_context,
        current_date=dates[2],
        entry_plan=None,
        entry_type="normal",
        entry_result={"filled": False, "count_as_missed_buy": False},
        note="",
    )
    add_check(results, "output_contract", case_id, "debug_entry_plan_marker_none_plan_noop", 0, len(chart_context["order_markers"]))
    return results, summary


def validate_gui_workbench_contract_case(base_params):
    case_id = "GUI_WORKBENCH_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    apps_gui = importlib.import_module("apps.workbench")
    tools_gui = importlib.import_module("tools.workbench_ui")
    workbench_spec = tools_gui.build_workbench_spec()

    add_check(results, "output_contract", case_id, "gui_app_thin_entry_main", apps_gui.main, tools_gui.main)
    add_check(results, "output_contract", case_id, "gui_workbench_entry_module", "apps.workbench", workbench_spec.get("entry_module"))
    add_check(results, "output_contract", case_id, "gui_workbench_title", "股票工具工作台", workbench_spec.get("title"))
    add_check(results, "output_contract", case_id, "gui_workbench_startup_window_mode", "maximized", workbench_spec.get("startup_window_mode"))

    panel_specs = workbench_spec.get("panels", [])
    panel_ids = [panel.get("panel_id") for panel in panel_specs]
    add_check(results, "output_contract", case_id, "gui_workbench_panel_ids", ["single_stock_backtest_inspector", "portfolio_backtest_inspector"], panel_ids)
    if panel_specs:
        panel_spec = panel_specs[0]
        add_check(results, "output_contract", case_id, "gui_workbench_panel_tab_label", "單股回測檢視", panel_spec.get("tab_label"))
        add_check(results, "output_contract", case_id, "gui_workbench_backend_runner", "tools.trade_analysis.trade_log.run_ticker_analysis", panel_spec.get("backend_runner"))
        add_check(results, "output_contract", case_id, "gui_workbench_artifact_keys", ["excel_path"], panel_spec.get("artifact_keys"))
        add_check(results, "output_contract", case_id, "gui_workbench_inline_chart_backend", "tools.trade_analysis.charting.create_matplotlib_trade_chart_figure", panel_spec.get("inline_chart_backend"))
        add_check(results, "output_contract", case_id, "gui_workbench_default_show_volume", False, panel_spec.get("default_show_volume"))
        add_check(results, "output_contract", case_id, "gui_workbench_single_jump_to_trade_enabled", True, panel_spec.get("jump_to_trade_enabled"))

    if len(panel_specs) > 1:
        portfolio_panel_spec = panel_specs[1]
        add_check(results, "output_contract", case_id, "gui_workbench_portfolio_panel_tab_label", "投組回測檢視", portfolio_panel_spec.get("tab_label"))
        add_check(results, "output_contract", case_id, "gui_workbench_portfolio_backend_runner", "tools.portfolio_sim.simulation_runner.run_portfolio_simulation_prepared", portfolio_panel_spec.get("backend_runner"))
        add_check(results, "output_contract", case_id, "gui_workbench_portfolio_artifact_keys", ["dashboard_html_path", "report_xlsx_path"], portfolio_panel_spec.get("artifact_keys"))
        add_check(results, "output_contract", case_id, "gui_workbench_portfolio_jump_to_trade_enabled", True, portfolio_panel_spec.get("jump_to_trade_enabled"))

    inspector_source = build_project_absolute_path("tools", "workbench_ui", "single_stock_inspector.py").read_text(encoding="utf-8")
    portfolio_inspector_source = build_project_absolute_path("tools", "workbench_ui", "portfolio_backtest_inspector.py").read_text(encoding="utf-8")
    portfolio_runner_source = build_project_absolute_path("tools", "portfolio_sim", "simulation_runner.py").read_text(encoding="utf-8")
    workbench_source = build_project_absolute_path("tools", "workbench_ui", "workbench.py").read_text(encoding="utf-8")
    meta_quality_coverage_source = build_project_absolute_path("tools", "local_regression", "meta_quality_coverage.py").read_text(encoding="utf-8")
    validate_main_source = build_project_absolute_path("tools", "validate", "main.py").read_text(encoding="utf-8")
    charting_source = build_project_absolute_path("tools", "trade_analysis", "charting.py").read_text(encoding="utf-8")
    add_check(results, "output_contract", case_id, "gui_workbench_registry_lazy_panel_imports", True, "from tools.workbench_ui.single_stock_inspector import" not in workbench_source and "from tools.workbench_ui.portfolio_backtest_inspector import" not in workbench_source)
    add_check(results, "output_contract", case_id, "gui_headless_coverage_omits_runtime_ui_files", True, "HEADLESS_COVERAGE_OMIT_PATTERNS" in meta_quality_coverage_source and "tools\" / \"workbench_ui\" / \"*.py\"" in meta_quality_coverage_source)
    add_check(results, "output_contract", case_id, "gui_validate_coverage_omits_runtime_ui_files", True, "HEADLESS_COVERAGE_OMIT_PATTERNS" in validate_main_source and "workbench_ui" in validate_main_source and "omit=HEADLESS_COVERAGE_OMIT_PATTERNS" in validate_main_source)
    add_check(results, "output_contract", case_id, "gui_trade_navigation_forces_immediate_redraw", True, "figure.canvas.draw()" in charting_source and "figure.canvas.flush_events()" in charting_source)
    add_check(results, "output_contract", case_id, "gui_trade_navigation_uses_compact_center_window", True, "window_width = max(" in charting_source and "min(current_width" in charting_source)
    add_check(results, "output_contract", case_id, "gui_single_trade_navigation_uses_buy_markers", True, "trace_names=BUY_TRADE_TRACE_NAMES" in inspector_source)
    add_check(results, "output_contract", case_id, "gui_portfolio_trade_navigation_uses_buy_markers", True, "trace_names=BUY_TRADE_TRACE_NAMES" in portfolio_inspector_source)
    portfolio_tab_close_block = portfolio_inspector_source.split("    def _refresh_performance_tab_close_buttons(", 1)[1].split("    def _destroy_performance_close_button", 1)[0]
    close_button_factory_block = portfolio_inspector_source.split("    def _get_or_create_performance_close_button(", 1)[1].split("    def _refresh_performance_tab_close_buttons", 1)[0]
    add_check(results, "output_contract", case_id, "gui_portfolio_performance_tab_close_button_is_positioned", True, "button.place(" in portfolio_tab_close_block and "button.lift()" in portfolio_tab_close_block)
    add_check(results, "output_contract", case_id, "gui_portfolio_performance_tab_close_button_is_real_button", True, "tk.Button(" in close_button_factory_block and "command=lambda" in close_button_factory_block)
    add_check(results, "output_contract", case_id, "gui_portfolio_performance_tab_close_button_is_panel_overlay", True, "tk.Button(\n            self," in close_button_factory_block and "button.place(in_=self" in portfolio_tab_close_block)
    add_check(results, "output_contract", case_id, "gui_portfolio_performance_tab_close_refresh_preserves_live_buttons", True, "if str(tab_id) in live_tab_ids:" in portfolio_tab_close_block and "continue" in portfolio_tab_close_block)
    add_check(results, "output_contract", case_id, "gui_single_trade_indexes_survive_chart_clear", True, "trade_indexes = extract_trade_marker_indexes" in inspector_source and "self._current_chart_trade_indexes = trade_indexes" in inspector_source)
    add_check(results, "output_contract", case_id, "gui_portfolio_trade_indexes_survive_chart_clear", True, "trade_indexes = _extract_trade_marker_indexes" in portfolio_inspector_source and "self._current_chart_trade_indexes = trade_indexes" in portfolio_inspector_source)
    add_check(results, "output_contract", case_id, "inspector_prefers_canonical_trade_analysis_runner", True, "run_ticker_analysis" in inspector_source and "run_debug_ticker_analysis" not in inspector_source)
    add_check(results, "output_contract", case_id, "inspector_prefers_canonical_trade_analysis_data_dir_helper", True, "resolve_trade_analysis_data_dir" in inspector_source and "resolve_debug_data_dir" not in inspector_source)
    add_check(results, "output_contract", case_id, "inspector_prefers_canonical_trade_chart_figure_alias", True, "create_matplotlib_trade_chart_figure" in inspector_source and "create_matplotlib_debug_chart_figure" not in inspector_source)
    add_check(results, "output_contract", case_id, "workbench_panel_registry_omits_legacy_debug_runner_alias", False, "run_debug_ticker_analysis" in workbench_source)
    add_check(results, "output_contract", case_id, "workbench_panel_registry_omits_legacy_debug_chart_alias", False, "create_matplotlib_debug_chart_figure" in workbench_source)

    single_start_scanner_block = inspector_source.split("    def _start_scanner(", 1)[1].split("    def _finish_scanner_success", 1)[0]
    single_prepare_console_block = inspector_source.split("    def _prepare_console_for_new_task(", 1)[1].split("    def _clear_console", 1)[0]
    portfolio_run_backtest_block = portfolio_inspector_source.split("    def _run_portfolio_backtest(", 1)[1].split("    def _run_portfolio_worker", 1)[0]
    portfolio_prepare_console_block = portfolio_inspector_source.split("    def _prepare_console_for_new_task(", 1)[1].split("    def _clear_console", 1)[0]
    add_check(results, "output_contract", case_id, "gui_single_new_scanner_task_preserves_console_history", True, "_prepare_console_for_new_task()" in single_start_scanner_block and "_clear_console()" not in single_start_scanner_block)
    add_check(results, "output_contract", case_id, "gui_single_prepare_console_task_does_not_delete_history", False, '.delete("1.0", "end")' in single_prepare_console_block)
    add_check(results, "output_contract", case_id, "gui_portfolio_new_backtest_task_preserves_console_history", True, "_prepare_console_for_new_task()" in portfolio_run_backtest_block and "_clear_console()" not in portfolio_run_backtest_block)
    add_check(results, "output_contract", case_id, "gui_portfolio_prepare_console_task_does_not_delete_history", False, '.delete("1.0", "end")' in portfolio_prepare_console_block)

    add_check(results, "output_contract", case_id, "gui_single_scanner_console_preserves_scanner_ansi_colors", True, "SCANNER_CONSOLE_COLORS" in inspector_source and "_insert_ansi_text" in inspector_source and "ANSI_ESCAPE_SEQUENCE_PATTERN" not in inspector_source)
    add_check(results, "output_contract", case_id, "gui_portfolio_has_end_year_dropdown", True, "END_YEAR_LATEST_LABEL" in portfolio_inspector_source and 'text="結束年"' in portfolio_inspector_source and "_end_year_combo" in portfolio_inspector_source)
    add_check(results, "output_contract", case_id, "gui_portfolio_passes_end_year_to_backend", True, 'end_year=options["end_year"]' in portfolio_inspector_source and 'export_portfolio_reports(df_eq, df_tr, df_yearly, options["benchmark_ticker"], options["start_year"], end_year=options["end_year"])' in portfolio_inspector_source)
    add_check(results, "output_contract", case_id, "portfolio_runner_filters_by_optional_end_year", True, "def _filter_market_dates_by_end_year" in portfolio_runner_source and "end_year=None" in portfolio_runner_source and "resolved_sorted_dates" in portfolio_runner_source)

    with io.StringIO() as stdout_buffer:
        with redirect_stdout(stdout_buffer):
            exit_code = tools_gui.main(["apps/workbench.py", "--help"])
        help_output = stdout_buffer.getvalue()
    add_check(results, "output_contract", case_id, "gui_help_exit_code", 0, exit_code)
    add_check(results, "output_contract", case_id, "gui_help_mentions_entry", True, "apps/workbench.py" in help_output)

    case = build_synthetic_competing_candidates_case(base_params, make_synthetic_validation_params)
    ticker = case["primary_ticker"]
    frame = case["frames"][ticker]
    min_rows_needed = get_required_min_rows(case["params"])
    clean_df, _sanitize_stats = sanitize_ohlcv_dataframe(frame.copy(), ticker, min_rows=min_rows_needed)

    debug_module = importlib.import_module("tools.trade_analysis.trade_log")
    charting_module = importlib.import_module("tools.trade_analysis.charting")
    add_check(results, "output_contract", case_id, "trade_log_exposes_canonical_trade_analysis_data_dir_helper", True, hasattr(debug_module, "resolve_trade_analysis_data_dir"))
    add_check(results, "output_contract", case_id, "charting_exposes_canonical_trade_chart_alias", True, hasattr(charting_module, "create_matplotlib_trade_chart_figure"))
    add_check(results, "output_contract", case_id, "charting_exposes_index_navigation_helper", True, hasattr(charting_module, "scroll_chart_to_index"))
    add_check(results, "output_contract", case_id, "charting_exposes_adjacent_trade_navigation_helper", True, hasattr(charting_module, "scroll_chart_to_adjacent_trade"))
    add_check(results, "output_contract", case_id, "charting_adjacent_trade_prev_from_latest", 12, charting_module.resolve_adjacent_trade_index(19, [4, 12], direction=-1))
    add_check(results, "output_contract", case_id, "charting_adjacent_trade_next_wraps_from_latest", 4, charting_module.resolve_adjacent_trade_index(19, [4, 12], direction=1))
    add_check(results, "output_contract", case_id, "charting_adjacent_trade_next_from_first", 12, charting_module.resolve_adjacent_trade_index(4, [4, 12], direction=1))
    add_check(results, "output_contract", case_id, "charting_adjacent_trade_empty_returns_none", None, charting_module.resolve_adjacent_trade_index(4, [], direction=1))
    legacy_df = debug_module.run_debug_backtest(clean_df.copy(), ticker, case["params"], export_excel=False, verbose=False)
    with tempfile.TemporaryDirectory(prefix="gui_workbench_contract_") as temp_dir:
        analysis_result = debug_module.run_debug_analysis(
            clean_df.copy(),
            ticker,
            case["params"],
            export_excel=True,
            export_chart=False,
            return_chart_payload=True,
            verbose=False,
            output_dir=temp_dir,
        )
        excel_exists = os.path.exists(analysis_result["excel_path"]) if analysis_result.get("excel_path") else False
        add_check(results, "output_contract", case_id, "gui_backend_trade_logs_payload", _normalize_nan_records(legacy_df), _normalize_nan_records(analysis_result["trade_logs_df"]))
        add_check(results, "output_contract", case_id, "gui_backend_chart_artifact_omitted", None, analysis_result.get("chart_path"))
        add_check(results, "output_contract", case_id, "gui_backend_excel_artifact_exists", True, excel_exists)
        add_check(results, "output_contract", case_id, "gui_backend_chart_payload_exists", True, analysis_result.get("chart_payload") is not None)
        add_check(results, "output_contract", case_id, "gui_backend_excel_artifact_name", f"Debug_TradeLog_{ticker}.xlsx", os.path.basename(analysis_result["excel_path"]))

    summary["workbench_title"] = workbench_spec.get("title")
    summary["panel_ids"] = panel_ids
    return results, summary


def validate_debug_trade_log_prepared_tool_contract_case(base_params):
    case_id = "DEBUG_TRADE_LOG_PREPARED_TOOL_CONTRACT"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    case = build_synthetic_competing_candidates_case(base_params, make_synthetic_validation_params)
    ticker = case["primary_ticker"]
    frame = case["frames"][ticker]
    min_rows_needed = get_required_min_rows(case["params"])
    clean_df, _sanitize_stats = sanitize_ohlcv_dataframe(frame.copy(), ticker, min_rows=min_rows_needed)
    prepared_df, _standalone_logs, _prepared_stats = prep_stock_data_and_trades(clean_df.copy(), case["params"], return_stats=True)

    legacy_df, legacy_module_path = run_debug_trade_log_check(ticker, clean_df, case["params"])
    prepared_result_df, prepared_module_path = run_debug_trade_log_check(ticker, clean_df, case["params"], prepared_df=prepared_df)

    legacy_records = _normalize_nan_records(legacy_df)
    prepared_records = _normalize_nan_records(prepared_result_df)

    add_check(results, "output_contract", case_id, "debug_prepared_module_path", legacy_module_path, prepared_module_path)
    add_check(results, "output_contract", case_id, "debug_prepared_payload", legacy_records, prepared_records)
    return results, summary



def validate_debug_trade_log_chart_context_optional_case(base_params):
    case_id = "DEBUG_TRADE_LOG_CHART_CONTEXT_OPTIONAL"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    case = build_synthetic_competing_candidates_case(base_params, make_synthetic_validation_params)
    ticker = case["primary_ticker"]
    frame = case["frames"][ticker]
    min_rows_needed = get_required_min_rows(case["params"])
    clean_df, _sanitize_stats = sanitize_ohlcv_dataframe(frame.copy(), ticker, min_rows=min_rows_needed)

    result_df, module_path = run_debug_trade_log_check(ticker, clean_df, case["params"])
    records = _normalize_nan_records(result_df)

    add_check(results, "output_contract", case_id, "debug_chart_context_optional_module_path", "tools/trade_analysis/trade_log.py", module_path)
    add_check(results, "output_contract", case_id, "debug_chart_context_optional_has_rows", True, bool(records))
    add_check(results, "output_contract", case_id, "debug_chart_context_optional_has_buy_row", True, any(str(row.get("動作", "")).startswith("買進") for row in records))
    return results, summary



def validate_tool_module_path_normalization_case(_base_params):
    case_id = "OUTPUT_TOOL_MODULE_PATH_NORMALIZATION"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    abs_debug_path = build_project_absolute_path("tools", "trade_analysis", "trade_log.py")
    backslash_debug_path = str(abs_debug_path).replace("/", "\\")

    add_check(results, "output_contract", case_id, "normalize_module_path_from_absolute", "tools/trade_analysis/trade_log.py", normalize_project_relative_path(abs_debug_path))
    add_check(results, "output_contract", case_id, "normalize_module_path_from_backslash_absolute", "tools/trade_analysis/trade_log.py", normalize_project_relative_path(backslash_debug_path))
    add_check(results, "output_contract", case_id, "normalize_module_path_relative_passthrough", "apps/workbench.py", normalize_project_relative_path("apps/workbench.py"))
    return results, summary



def validate_module_path_normalizer_accepts_path_objects_case(_base_params):
    case_id = "OUTPUT_TOOL_MODULE_PATH_NORMALIZER_ACCEPTS_PATH_OBJECTS"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    debug_path = build_project_absolute_path("tools", "trade_analysis", "trade_log.py")
    gui_path = build_project_absolute_path("apps", "workbench.py")

    add_check(results, "output_contract", case_id, "normalize_module_path_from_path_object_debug", "tools/trade_analysis/trade_log.py", normalize_project_relative_path(debug_path))
    add_check(results, "output_contract", case_id, "normalize_module_path_from_path_object_gui", "apps/workbench.py", normalize_project_relative_path(gui_path))
    return results, summary



def validate_module_loader_project_root_string_patch_case(_base_params):
    case_id = "OUTPUT_TOOL_MODULE_LOADER_PROJECT_ROOT_STRING_PATCH"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    with tempfile.TemporaryDirectory(prefix="v16_module_loader_root_patch_") as tmp_dir:
        runtime_root = Path(tmp_dir)
        with patch.object(module_loader_module, "PROJECT_ROOT", str(runtime_root)):
            patched_abs_path = module_loader_module.build_project_absolute_path("tools", "trade_analysis", "trade_log.py")
            add_check(results, "output_contract", case_id, "build_project_absolute_path_accepts_string_project_root", str(runtime_root / "tools" / "trade_analysis" / "trade_log.py"), str(patched_abs_path))
            add_check(results, "output_contract", case_id, "normalize_module_path_under_string_project_root", "tools/trade_analysis/trade_log.py", module_loader_module.normalize_project_relative_path(patched_abs_path))
            add_check(results, "output_contract", case_id, "normalize_absolute_string_path_under_string_project_root", "tools/trade_analysis/trade_log.py", module_loader_module.normalize_project_relative_path(str(runtime_root / "tools" / "trade_analysis" / "trade_log.py")))
    return results, summary

def validate_meta_quality_reuses_existing_coverage_artifacts_case(base_params):
    case_id = "META_QUALITY_REUSE_COVERAGE_ARTIFACTS"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    from tools.validate import synthetic_cases as synthetic_cases_module

    manifest = local_common.load_manifest()
    overall_line_floor = float(manifest["coverage_line_min_percent"])
    overall_branch_floor = float(manifest["coverage_branch_min_percent"])
    critical_line_floor = float(manifest["coverage_critical_line_min_percent"])
    critical_branch_floor = float(manifest["coverage_critical_branch_min_percent"])
    synthetic_case_count = len(synthetic_cases_module.get_synthetic_validators())

    with tempfile.TemporaryDirectory(prefix="meta_quality_reuse_cov_") as temp_dir:
        run_dir = Path(temp_dir)
        coverage_dir = run_dir / "coverage_artifacts"
        coverage_dir.mkdir(parents=True, exist_ok=True)
        coverage_payload = {
            "totals": {
                "covered_lines": int(max(overall_line_floor, 60.0) * 2),
                "num_statements": 200,
                "covered_branches": int(max(overall_branch_floor, 60.0) * 2),
                "num_branches": 200,
                "percent_covered": float((max(overall_line_floor, 60.0) + max(overall_branch_floor, 60.0)) / 2.0),
            },
            "files": {
                rel_path: {
                    "summary": {
                        "covered_lines": int(critical_line_floor) if rel_path in CRITICAL_COVERAGE_TARGETS else 1,
                        "num_statements": 100 if rel_path in CRITICAL_COVERAGE_TARGETS else 1,
                        "percent_covered": float(critical_line_floor) if rel_path in CRITICAL_COVERAGE_TARGETS else 100.0,
                        "covered_branches": int(critical_branch_floor) if rel_path in CRITICAL_COVERAGE_TARGETS else 1,
                        "num_branches": 100 if rel_path in CRITICAL_COVERAGE_TARGETS else 1,
                    }
                }
                for rel_path in COVERAGE_TARGETS
            },
        }
        write_text(coverage_dir / "coverage_synthetic.json", json.dumps(coverage_payload, ensure_ascii=False, indent=2))
        write_json(coverage_dir / "coverage_run_info.json", {
            "source": "validate_consistency",
            "returncode": 0,
            "stdout": "cached",
            "stderr": "",
            "timed_out": False,
            "synthetic_fail_count": 0,
            "synthetic_case_count": synthetic_case_count,
            "json_generated": True,
        })

        with patch("tools.validate.synthetic_cases.run_synthetic_consistency_suite", side_effect=AssertionError("should reuse existing coverage artifacts")):
            coverage_summary = build_coverage_summary(run_dir, manifest)

    add_check(results, "output_contract", case_id, "meta_quality_reuse_coverage_ok", True, coverage_summary.get("ok"))
    add_check(results, "output_contract", case_id, "meta_quality_reuse_coverage_flag", True, coverage_summary.get("reused_existing"))
    add_check(results, "output_contract", case_id, "meta_quality_reuse_coverage_returncode", 0, coverage_summary.get("run_result", {}).get("returncode"))
    return results, summary



def validate_dataset_prepare_fallback_write_traceability_case(_base_params):
    case_id = "DATASET_PREPARE_FALLBACK_WRITE_TRACEABILITY"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    with tempfile.TemporaryDirectory(prefix="dataset_prepare_fallback_trace_") as temp_dir:
        run_dir = Path(temp_dir)
        stderr_buffer = io.StringIO()

        def _fail_json(_path, _payload):
            raise OSError("json fallback disk full")

        def _fail_text(_path, _text):
            raise OSError("text fallback permission denied")

        with patch.object(run_all_module, "_write_dataset_prepare_summary", side_effect=ValueError("primary summary write failed")),              patch.object(run_all_module, "write_json", side_effect=_fail_json),              patch.object(run_all_module, "write_text", side_effect=_fail_text),              patch("sys.stderr", stderr_buffer):
            fallback_summary = run_all_module._safe_write_dataset_prepare_summary(
                run_dir,
                {
                    "status": "FAIL",
                    "duration_sec": 1.234,
                    "error_type": "RuntimeError",
                    "error_message": "dataset missing",
                },
            )

    fallback_write_errors = fallback_summary.get("fallback_write_errors", [])
    stderr_text = stderr_buffer.getvalue()
    add_check(results, "output_contract", case_id, "fallback_summary_preserves_primary_write_error", "ValueError: primary summary write failed", fallback_summary.get("summary_write_error"))
    add_check(results, "output_contract", case_id, "fallback_summary_records_json_write_failure", True, any(str(item).startswith("json:OSError: json fallback disk full") for item in fallback_write_errors))
    add_check(results, "output_contract", case_id, "fallback_summary_records_text_write_failure", True, any(str(item).startswith("txt:OSError: text fallback permission denied") for item in fallback_write_errors))
    add_check(results, "output_contract", case_id, "fallback_summary_emits_stderr_when_no_fallback_file_written", True, "dataset prepare fallback summary write failed" in stderr_text and "json fallback disk full" in stderr_text and "text fallback permission denied" in stderr_text)

    summary["fallback_error_count"] = len(fallback_write_errors)
    return results, summary



def validate_console_tail_read_error_traceability_case(_base_params):
    case_id = "CONSOLE_TAIL_READ_ERROR_TRACEABILITY"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    with tempfile.TemporaryDirectory(prefix="console_tail_trace_") as temp_dir:
        run_dir = Path(temp_dir)
        ok_log = run_dir / "ok.log"
        bad_log = run_dir / "bad.log"
        ok_log.write_text("line-1\nline-2\n", encoding="utf-8")
        bad_log.write_text("unreadable\n", encoding="utf-8")
        original_read_text = Path.read_text

        def _patched_read_text(self, *args, **kwargs):
            if self == bad_log:
                raise OSError("simulated tail read failure")
            return original_read_text(self, *args, **kwargs)

        with patch.object(Path, "read_text", _patched_read_text):
            tail_text = local_common.gather_recent_console_tail(run_dir, limit_lines=5)

    add_check(results, "output_contract", case_id, "console_tail_keeps_readable_log_tail", True, "===== ok.log =====" in tail_text and "line-2" in tail_text)
    add_check(results, "output_contract", case_id, "console_tail_surfaces_read_error_instead_of_silent_skip", True, "===== bad.log =====" in tail_text and "[read_error] OSError: simulated tail read failure" in tail_text)

    summary["tail_has_read_error_marker"] = "[read_error]" in tail_text
    return results, summary
