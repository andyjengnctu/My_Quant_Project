import io
import sys
import tempfile
import types
from contextlib import redirect_stdout
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from apps import test_suite as test_suite_module
from core.display_common import _strip_ansi
from tools.portfolio_sim import reporting as portfolio_reporting
from tools.validate.reporting import print_console_summary

from .checks import add_check


def _capture_output(callable_obj):
    buffer = io.StringIO()
    with redirect_stdout(buffer):
        callable_obj()
    return _strip_ansi(buffer.getvalue())


def validate_validate_console_summary_reporting_case(_base_params):
    case_id = "VALIDATE_CONSOLE_SUMMARY_REPORTING"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    df_results = pd.DataFrame(
        [
            {"ticker": "1101", "module": "stats", "metric": "win_rate", "status": "PASS", "passed": True, "expected": "58.3", "actual": "58.3", "note": ""},
            {"ticker": "1102", "module": "stats", "metric": "ev", "status": "FAIL", "passed": False, "expected": "0.8", "actual": "0.4", "note": "EV mismatch"},
            {"ticker": "SYN_CASE", "module": "synthetic", "metric": "rule", "status": "FAIL", "passed": False, "expected": "PASS", "actual": "FAIL", "note": "Synthetic fail"},
            {"ticker": "REAL_DATA_COVERAGE", "module": "system", "metric": "coverage", "status": "SKIP", "passed": False, "expected": "full", "actual": "synthetic-only", "note": "skip"},
        ]
    )
    df_failed = df_results[df_results["status"] == "FAIL"].copy()
    df_summary = pd.DataFrame(
        [
            {"ticker": "1101", "synthetic": False},
            {"ticker": "1102", "synthetic": False},
            {"ticker": "SYN_CASE", "synthetic": True},
            {"ticker": "REAL_DATA_COVERAGE", "synthetic": False},
        ]
    )

    console_text = _capture_output(
        lambda: print_console_summary(
            df_results=df_results,
            df_failed=df_failed,
            df_summary=df_summary,
            csv_path="outputs/validate/consistency_full_scan_20260402.csv",
            xlsx_path="outputs/validate/consistency_issues_20260402.xlsx",
            elapsed_time=12.34,
            real_summary_count=2,
            real_tickers=["1101", "1102"],
            normalize_ticker_text=lambda value: str(value).strip(),
            max_console_fail_preview=5,
        )
    )

    add_check(results, "reporting_schema", case_id, "console_summary_has_title", True, "一致性回歸摘要" in console_text)
    add_check(results, "reporting_schema", case_id, "console_summary_has_counts", True, "PASS 數: 1" in console_text and "SKIP 數: 1" in console_text and "FAIL 數: 2" in console_text)
    add_check(results, "reporting_schema", case_id, "console_summary_has_problem_counts", True, "有問題真實股票數: 1" in console_text and "有問題 synthetic case 數: 1" in console_text and "有問題 system 項目數: 0" in console_text)
    add_check(results, "reporting_schema", case_id, "console_summary_has_paths", True, "完整 CSV: outputs/validate/consistency_full_scan_20260402.csv" in console_text and "問題 Excel: outputs/validate/consistency_issues_20260402.xlsx" in console_text)
    add_check(results, "reporting_schema", case_id, "console_summary_has_fail_preview_sections", True, "失敗項前覽：" in console_text and "失敗真實股票前覽：" in console_text and "失敗 synthetic/system 前覽：" in console_text)
    add_check(results, "reporting_schema", case_id, "console_summary_has_fail_preview_details", True, "EV mismatch" in console_text and "Synthetic fail" in console_text)

    summary["console_summary_lines"] = len([line for line in console_text.splitlines() if line.strip()])
    return results, summary


def validate_portfolio_yearly_report_schema_case(_base_params):
    case_id = "PORTFOLIO_YEARLY_REPORT_SCHEMA"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    yearly_rows = [
        {"year": 2023, "year_return_pct": 12.345, "is_full_year": True, "start_date": "2023-01-03", "end_date": "2023-12-29"},
        {"year": 2024, "year_return_pct": -3.21, "is_full_year": False, "start_date": "2024-01-02", "end_date": "2024-06-28"},
    ]

    captured_df = None

    def _run_non_empty():
        nonlocal captured_df
        captured_df = portfolio_reporting.print_yearly_return_report(yearly_rows)

    report_text = _capture_output(_run_non_empty)
    add_check(results, "reporting_schema", case_id, "yearly_report_has_title", True, "各年度報酬率" in report_text)
    add_check(results, "reporting_schema", case_id, "yearly_report_has_percent_and_type", True, "12.35%" in report_text and "完整" in report_text and "-3.21%" in report_text and "非完整" in report_text)
    add_check(results, "reporting_schema", case_id, "yearly_report_returns_dataframe_columns", ["year", "year_return_pct", "is_full_year", "start_date", "end_date", "year_label", "year_type"], list(captured_df.columns))
    add_check(results, "reporting_schema", case_id, "yearly_report_year_label_and_type", True, captured_df.loc[0, "year_label"] == "2023" and captured_df.loc[1, "year_type"] == "非完整")

    empty_df = None

    def _run_empty():
        nonlocal empty_df
        empty_df = portfolio_reporting.print_yearly_return_report([])

    empty_text = _capture_output(_run_empty)
    add_check(results, "reporting_schema", case_id, "yearly_report_empty_warning", True, "無年度報酬率資料" in empty_text)
    add_check(results, "reporting_schema", case_id, "yearly_report_empty_schema", ["year", "year_return_pct", "is_full_year", "start_date", "end_date"], list(empty_df.columns))

    summary["yearly_rows"] = int(len(captured_df)) if captured_df is not None else 0
    return results, summary


def _base_test_suite_result_payload():
    return {
        "overall_status": "PASS",
        "failures": 0,
        "selected_steps": ["quick_gate", "consistency", "chain_checks", "ml_smoke", "meta_quality"],
        "failed_step_names": [],
        "not_run_step_names": [],
        "scripts": [
            {"name": "quick_gate", "status": "PASS", "duration_sec": 8.42, "failure_reasons": []},
            {"name": "consistency", "status": "PASS", "duration_sec": 10.23, "failure_reasons": []},
            {"name": "chain_checks", "status": "PASS", "duration_sec": 11.22, "failure_reasons": []},
            {"name": "ml_smoke", "status": "PASS", "duration_sec": 4.01, "failure_reasons": []},
            {"name": "meta_quality", "status": "PASS", "duration_sec": 1.81, "failure_reasons": []},
        ],
        "step_payloads": {
            "dataset_prepare": {"status": "PASS", "duration_sec": 0.62, "csv_count": 617, "source": "data/tw_stock_data_vip_reduced"},
            "quick_gate": {"status": "PASS", "step_count": 88, "failed_count": 0},
            "consistency": {"status": "PASS", "total_checks": 2752, "fail_count": 0, "skip_count": 0, "real_ticker_count": 24},
            "chain_checks": {"status": "PASS", "ticker_count": 24, "highlights": {"traded_ticker_count": 7, "missed_buy_ticker_count": 2, "blocked_by_counts": {"cash": 3, "slots": 1}}, "portfolio_snapshot": {"trade_rows": 14, "reserved_buy_fill_rate": 83.33}},
            "ml_smoke": {"status": "PASS", "db_trial_count": 1},
            "meta_quality": {"status": "PASS", "fail_count": 0, "coverage": {"totals": {"percent_covered": 52.41}}, "checklist": {"todo_ids": []}},
        },
        "preflight": {"status": "PASS", "duration_sec": 0.45, "failed_packages": []},
        "bundle_mode": "minimum_set",
        "archived_bundle": "outputs/local_regression/to_chatgpt_bundle_20260402_123456_abcd1234.zip",
        "root_bundle_copy": "to_chatgpt_bundle_20260402_123456_abcd1234.zip",
        "bundle_entries": ["master_summary.json", "preflight_summary.json", "quick_gate_summary.json"],
        "retention": {"removed_count": 2, "removed_bytes": 4096},
    }


def validate_test_suite_summary_reporting_case(_base_params):
    case_id = "TEST_SUITE_SUMMARY_REPORTING"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    result_payload = _base_test_suite_result_payload()

    summary_text = _capture_output(lambda: test_suite_module._print_human_summary(result_payload))
    add_check(results, "reporting_schema", case_id, "test_suite_summary_has_title_and_bundle", True, "Test Suite 結果整理" in summary_text and "bundle 模式 : minimum_set" in summary_text)
    add_check(results, "reporting_schema", case_id, "test_suite_summary_has_step_table", True, "[步驟摘要]" in summary_text and "quick gate" in summary_text and "meta quality" in summary_text)
    add_check(results, "reporting_schema", case_id, "test_suite_summary_has_highlights", True, "step_count=88" in summary_text and "total_checks=2752" in summary_text and "db_trial_count=1" in summary_text)
    add_check(results, "reporting_schema", case_id, "test_suite_summary_has_chain_and_meta_details", True, "blocked_by : cash:3, slots:1" in summary_text and "coverage_percent=52.41" in summary_text)
    add_check(results, "reporting_schema", case_id, "test_suite_summary_has_retention", True, "retention  : removed=2 | bytes=4096" in summary_text)

    summary["test_suite_summary_lines"] = len([line for line in summary_text.splitlines() if line.strip()])
    return results, summary


def validate_issue_excel_report_schema_case(_base_params):
    case_id = "VALIDATE_ISSUE_EXCEL_REPORT_SCHEMA"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    from tools.validate.reporting import write_issue_excel_report

    df_failed = pd.DataFrame([
        {"ticker": 23, "module": "stats", "metric": "ev", "expected": "1.0", "actual": "0.2", "note": "EV mismatch"},
        {"ticker": 1101, "module": "flow", "metric": "missed_sell", "expected": "0", "actual": "1", "note": "unexpected miss"},
    ])
    df_failed_summary = pd.DataFrame([
        {"ticker": 23, "failed_checks": 1},
        {"ticker": 1101, "failed_checks": 1},
    ])
    df_failed_module = pd.DataFrame([
        {"module": "stats", "failed_checks": 1},
        {"module": "flow", "failed_checks": 1},
    ])

    with tempfile.TemporaryDirectory() as tmp_dir:
        report_path = write_issue_excel_report(
            df_failed=df_failed,
            df_failed_summary=df_failed_summary,
            df_failed_module=df_failed_module,
            timestamp="20260402_120000",
            output_dir=tmp_dir,
            normalize_ticker=lambda value: str(value).strip().zfill(4),
        )
        workbook = load_workbook(report_path)
        failed_only = workbook["failed_only"]
        failed_tickers = workbook["failed_tickers"]
        failed_modules = workbook["failed_modules"]

        add_check(results, "reporting_schema", case_id, "issue_excel_path_exists", True, Path(report_path).is_file())
        add_check(results, "reporting_schema", case_id, "issue_excel_filename_pattern", True, str(report_path).endswith("consistency_issues_20260402_120000.xlsx"))
        add_check(results, "reporting_schema", case_id, "issue_excel_sheet_names", ["failed_only", "failed_tickers", "failed_modules"], workbook.sheetnames)
        add_check(results, "reporting_schema", case_id, "issue_excel_failed_only_header", ["ticker", "module", "metric", "expected", "actual", "note"], [cell.value for cell in failed_only[1]])
        add_check(results, "reporting_schema", case_id, "issue_excel_failed_summary_header", ["ticker", "failed_checks"], [cell.value for cell in failed_tickers[1]])
        add_check(results, "reporting_schema", case_id, "issue_excel_failed_module_header", ["module", "failed_checks"], [cell.value for cell in failed_modules[1]])
        add_check(results, "reporting_schema", case_id, "issue_excel_ticker_normalized_and_text", True, failed_only["A2"].value == "0023" and failed_only["A2"].number_format == "@" and failed_tickers["A2"].value == "0023" and failed_tickers["A2"].number_format == "@")

    summary["issue_excel_rows"] = int(len(df_failed))
    return results, summary


def validate_portfolio_export_report_artifacts_case(_base_params):
    case_id = "PORTFOLIO_EXPORT_REPORT_ARTIFACTS"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    df_eq = pd.DataFrame([
        {"Date": "2024-01-02", "Strategy_Return_Pct": 0.0, "Benchmark_0050_Pct": 0.0},
        {"Date": "2024-01-03", "Strategy_Return_Pct": 1.2, "Benchmark_0050_Pct": 0.5},
    ])
    df_tr = pd.DataFrame([
        {"ticker": "2330", "entry_date": "2024-01-02", "exit_date": "2024-01-03", "net_pnl": 1234.0},
    ])
    df_yearly = pd.DataFrame([
        {"year": 2024, "year_return_pct": 1.2, "is_full_year": False, "start_date": "2024-01-02", "end_date": "2024-01-03", "year_label": "2024", "year_type": "非完整"},
    ])

    class _FakeFigure:
        def __init__(self):
            self.traces = []
            self.layout = {}

        def add_trace(self, trace):
            self.traces.append(trace)

        def update_layout(self, **kwargs):
            self.layout.update(kwargs)

        def write_html(self, path):
            Path(path).write_text(
                f"title={self.layout.get('title','')}\ntrace_count={len(self.traces)}\n",
                encoding="utf-8",
            )

    fake_go = types.ModuleType("plotly.graph_objects")
    fake_go.Figure = _FakeFigure
    fake_go.Scatter = lambda **kwargs: {"kind": "scatter", **kwargs}
    fake_plotly = types.ModuleType("plotly")
    fake_plotly.graph_objects = fake_go

    with tempfile.TemporaryDirectory() as tmp_dir:
        xlsx_path = Path(tmp_dir) / "portfolio_report.xlsx"
        html_path = Path(tmp_dir) / "portfolio_dashboard.html"
        with patch.object(portfolio_reporting, "REPORT_XLSX_PATH", str(xlsx_path)), \
             patch.object(portfolio_reporting, "DASHBOARD_HTML_PATH", str(html_path)), \
             patch.dict(sys.modules, {"plotly": fake_plotly, "plotly.graph_objects": fake_go}):
            export_text = _capture_output(
                lambda: portfolio_reporting.export_portfolio_reports(
                    df_eq=df_eq,
                    df_tr=df_tr,
                    df_yearly=df_yearly,
                    benchmark_ticker="0050",
                    start_year=2024,
                )
            )

        workbook = load_workbook(xlsx_path)
        eq_ws = workbook["Equity Curve"]
        tr_ws = workbook["Trade History"]
        yr_ws = workbook["Yearly Returns"]
        html_text = html_path.read_text(encoding="utf-8")

        add_check(results, "reporting_schema", case_id, "portfolio_export_xlsx_exists", True, xlsx_path.is_file())
        add_check(results, "reporting_schema", case_id, "portfolio_export_html_exists", True, html_path.is_file())
        add_check(results, "reporting_schema", case_id, "portfolio_export_sheet_names", ["Equity Curve", "Trade History", "Yearly Returns"], workbook.sheetnames)
        add_check(results, "reporting_schema", case_id, "portfolio_export_equity_header", ["Date", "Strategy_Return_Pct", "Benchmark_0050_Pct"], [cell.value for cell in eq_ws[1]])
        add_check(results, "reporting_schema", case_id, "portfolio_export_trade_header", ["ticker", "entry_date", "exit_date", "net_pnl"], [cell.value for cell in tr_ws[1]])
        add_check(results, "reporting_schema", case_id, "portfolio_export_yearly_header", ["year", "year_return_pct", "is_full_year", "start_date", "end_date", "year_label", "year_type"], [cell.value for cell in yr_ws[1]])
        add_check(results, "reporting_schema", case_id, "portfolio_export_html_trace_count", True, "trace_count=2" in html_text)
        add_check(results, "reporting_schema", case_id, "portfolio_export_console_paths", True, str(xlsx_path) in export_text and str(html_path) in export_text)

    summary["portfolio_export_rows"] = int(len(df_eq))
    return results, summary


def validate_test_suite_summary_failure_reporting_case(_base_params):
    case_id = "TEST_SUITE_SUMMARY_FAILURE_REPORTING"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    result_payload = _base_test_suite_result_payload()
    result_payload.update({
        "overall_status": "FAIL",
        "failures": 1,
        "failed_step_names": ["chain_checks"],
        "scripts": [
            {"name": "quick_gate", "status": "PASS", "duration_sec": 8.42, "failure_reasons": []},
            {"name": "consistency", "status": "PASS", "duration_sec": 10.23, "failure_reasons": []},
            {"name": "chain_checks", "status": "FAIL", "duration_sec": 11.22, "failure_reasons": ["returncode=1", "failed_steps=scanner_snapshot", "summary_error=snapshot mismatch"]},
            {"name": "ml_smoke", "status": "PASS", "duration_sec": 4.01, "failure_reasons": []},
            {"name": "meta_quality", "status": "PASS", "duration_sec": 1.81, "failure_reasons": []},
        ],
        "step_payloads": {
            **result_payload["step_payloads"],
            "chain_checks": {
                "status": "FAIL",
                "error_message": "snapshot mismatch",
                "failed_steps": ["scanner_snapshot"],
                "ticker_count": 24,
            },
        },
        "bundle_mode": "debug_bundle",
        "suggested_rerun_command": "python tools/local_regression/run_all.py --only chain_checks",
    })

    summary_text = _capture_output(lambda: test_suite_module._print_human_summary(result_payload))
    add_check(results, "reporting_schema", case_id, "test_suite_failure_summary_has_failure_line", True, "失敗步驟 : chain checks(chain_checks)" in summary_text)
    add_check(results, "reporting_schema", case_id, "test_suite_failure_summary_has_step_reason", True, "chain checks" in summary_text and "reported_status=FAIL" in summary_text and "failed_steps=scanner_snapshot" in summary_text)
    add_check(results, "reporting_schema", case_id, "test_suite_failure_summary_has_rerun_command", True, "建議重跑 : python tools/local_regression/run_all.py --only chain_checks" in summary_text)

    summary["failure_summary_lines"] = len([line for line in summary_text.splitlines() if line.strip()])
    return results, summary


def validate_test_suite_summary_manifest_failure_reporting_case(_base_params):
    case_id = "TEST_SUITE_SUMMARY_MANIFEST_FAILURE_REPORTING"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    result_payload = {
        "overall_status": "FAIL",
        "failures": 1,
        "selected_steps": ["quick_gate", "consistency", "chain_checks", "ml_smoke", "meta_quality"],
        "failed_step_names": ["manifest"],
        "not_run_step_names": ["preflight", "dataset_prepare", "quick_gate", "consistency", "chain_checks", "ml_smoke", "meta_quality"],
        "scripts": [],
        "step_payloads": {
            "manifest": {"status": "FAIL", "error_type": "LocalRegressionError", "error_message": "manifest 欄位 bundle_name 不可空白"},
        },
        "preflight": {},
        "bundle_mode": "manifest_failed",
        "archived_bundle": "outputs/local_regression/to_chatgpt_bundle_20260402_123456_abcd1234.zip",
        "root_bundle_copy": "to_chatgpt_bundle_20260402_123456_abcd1234.zip",
        "bundle_entries": ["master_summary.json", "manifest_summary.json"],
        "retention": {"removed_count": 0, "removed_bytes": 0},
    }

    summary_text = _capture_output(lambda: test_suite_module._print_human_summary(result_payload))
    add_check(results, "reporting_schema", case_id, "test_suite_manifest_summary_has_manifest_error", True, "manifest    : FAIL | LocalRegressionError: manifest 欄位 bundle_name 不可空白" in summary_text)
    add_check(results, "reporting_schema", case_id, "test_suite_manifest_summary_marks_blocked_steps", True, "preflight   : NOT_RUN | blocked_by_manifest" in summary_text and "dataset prep: NOT_RUN | blocked_by_manifest" in summary_text and "quick gate    NOT_RUN" in summary_text and "blocked_by_manifest" in summary_text)
    add_check(results, "reporting_schema", case_id, "test_suite_manifest_summary_lists_not_run_steps", True, "未執行步驟 : dataset prepare(dataset_prepare), quick gate(quick_gate), consistency, chain checks(chain_checks), ml smoke(ml_smoke), meta quality(meta_quality)" in summary_text)

    summary["manifest_summary_lines"] = len([line for line in summary_text.splitlines() if line.strip()])
    return results, summary


def validate_test_suite_summary_optional_dataset_skip_case(_base_params):
    case_id = "TEST_SUITE_SUMMARY_OPTIONAL_DATASET_SKIP"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    result_payload = {
        "overall_status": "PASS",
        "failures": 0,
        "selected_steps": ["meta_quality"],
        "failed_step_names": [],
        "not_run_step_names": [],
        "scripts": [
            {"name": "meta_quality", "status": "PASS", "duration_sec": 1.81, "failure_reasons": []},
        ],
        "step_payloads": {
            "meta_quality": {"status": "PASS", "fail_count": 0, "coverage": {"totals": {"percent_covered": 51.23}}, "checklist": {"todo_ids": ["B15", "B19"]}},
        },
        "preflight": {"status": "PASS", "duration_sec": 0.18, "failed_packages": []},
        "bundle_mode": "minimum_set",
        "archived_bundle": "outputs/local_regression/to_chatgpt_bundle_20260402_123456_abcd1234.zip",
        "root_bundle_copy": "to_chatgpt_bundle_20260402_123456_abcd1234.zip",
        "bundle_entries": ["master_summary.json", "meta_quality_summary.json"],
        "retention": {"removed_count": 0, "removed_bytes": 0},
    }

    summary_text = _capture_output(lambda: test_suite_module._print_human_summary(result_payload))
    add_check(results, "reporting_schema", case_id, "test_suite_partial_summary_marks_selected_step", True, "執行步驟 : meta quality" in summary_text)
    add_check(results, "reporting_schema", case_id, "test_suite_partial_summary_marks_dataset_not_required", True, "dataset prep: SKIP | not_required" in summary_text)
    add_check(results, "reporting_schema", case_id, "test_suite_partial_summary_marks_unselected_steps", True, "quick gate    SKIP" in summary_text and "consistency   SKIP" in summary_text and "ml smoke      SKIP" in summary_text and "not_selected" in summary_text)

    summary["partial_summary_lines"] = len([line for line in summary_text.splitlines() if line.strip()])
    return results, summary


def validate_test_suite_summary_preflight_failure_reporting_case(_base_params):
    case_id = "TEST_SUITE_SUMMARY_PREFLIGHT_FAILURE_REPORTING"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    result_payload = {
        "overall_status": "FAIL",
        "failures": 1,
        "selected_steps": ["quick_gate", "consistency", "chain_checks", "ml_smoke", "meta_quality"],
        "failed_step_names": ["preflight"],
        "not_run_step_names": ["dataset_prepare", "quick_gate", "consistency", "chain_checks", "ml_smoke", "meta_quality"],
        "scripts": [],
        "step_payloads": {},
        "preflight": {"status": "FAIL", "duration_sec": 0.31, "failed_packages": ["numpy", "pandas"], "error_message": "missing packages"},
        "bundle_mode": "debug_bundle",
        "archived_bundle": None,
        "root_bundle_copy": None,
        "bundle_entries": ["master_summary.json", "preflight_summary.json"],
        "retention": {"removed_count": 0, "removed_bytes": 0},
    }

    summary_text = _capture_output(lambda: test_suite_module._print_human_summary(result_payload))
    add_check(results, "reporting_schema", case_id, "test_suite_preflight_failure_has_failure_detail", True, "preflight   : FAIL | reported_status=FAIL" in summary_text and "failed_packages=numpy,pandas" in summary_text and "error_message=missing packages" in summary_text)
    add_check(results, "reporting_schema", case_id, "test_suite_preflight_failure_marks_blocked_steps", True, "dataset prep: NOT_RUN | blocked_by_preflight" in summary_text and "quick gate    NOT_RUN" in summary_text and "blocked_by_preflight" in summary_text)
    add_check(results, "reporting_schema", case_id, "test_suite_preflight_failure_formats_named_steps_and_empty_bundle", True, "未執行步驟 : dataset prepare(dataset_prepare), quick gate(quick_gate)" in summary_text and "歷史 bundle : (none)" in summary_text and "根目錄 bundle : (none)" in summary_text)

    summary["preflight_failure_summary_lines"] = len([line for line in summary_text.splitlines() if line.strip()])
    return results, summary


def validate_test_suite_summary_dataset_prepare_failure_reporting_case(_base_params):
    case_id = "TEST_SUITE_SUMMARY_DATASET_PREPARE_FAILURE_REPORTING"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    result_payload = _base_test_suite_result_payload()
    result_payload.update({
        "overall_status": "FAIL",
        "failures": 1,
        "failed_step_names": ["dataset_prepare"],
        "not_run_step_names": ["quick_gate", "consistency", "chain_checks", "ml_smoke", "meta_quality"],
        "scripts": [],
        "step_payloads": {
            "dataset_prepare": {
                "status": "FAIL",
                "duration_sec": 0.77,
                "error_message": "reduced dataset missing",
                "summary_write_error": "disk full",
            },
        },
        "preflight": {"status": "PASS", "duration_sec": 0.18, "failed_packages": []},
        "bundle_mode": "debug_bundle",
        "suggested_rerun_command": "",
    })

    summary_text = _capture_output(lambda: test_suite_module._print_human_summary(result_payload))
    add_check(results, "reporting_schema", case_id, "test_suite_dataset_failure_has_detail", True, "dataset prep: FAIL | reported_status=FAIL" in summary_text and "error_message=reduced dataset missing" in summary_text and "summary_write_error=disk full" in summary_text)
    add_check(results, "reporting_schema", case_id, "test_suite_dataset_failure_marks_regression_blocked", True, "quick gate    NOT_RUN" in summary_text and "blocked_by_dataset_prepare" in summary_text and "未執行步驟 : quick gate(quick_gate), consistency, chain checks(chain_checks), ml smoke(ml_smoke), meta quality(meta_quality)" in summary_text)

    summary["dataset_failure_summary_lines"] = len([line for line in summary_text.splitlines() if line.strip()])
    return results, summary


def validate_test_suite_summary_unreadable_payload_reporting_case(_base_params):
    case_id = "TEST_SUITE_SUMMARY_UNREADABLE_PAYLOAD_REPORTING"
    results = []
    summary = {"ticker": case_id, "synthetic": True}

    result_payload = _base_test_suite_result_payload()
    result_payload.update({
        "overall_status": "FAIL",
        "failures": 1,
        "failed_step_names": ["chain_checks"],
        "scripts": [
            {"name": "quick_gate", "status": "PASS", "duration_sec": 8.42, "failure_reasons": []},
            {"name": "consistency", "status": "PASS", "duration_sec": 10.23, "failure_reasons": []},
            {"name": "chain_checks", "status": "FAIL", "duration_sec": 11.22, "failure_reasons": ["summary_unreadable", "error_type=JSONDecodeError"]},
            {"name": "ml_smoke", "status": "PASS", "duration_sec": 4.01, "failure_reasons": []},
            {"name": "meta_quality", "status": "PASS", "duration_sec": 1.81, "failure_reasons": []},
        ],
        "step_payloads": {
            **result_payload["step_payloads"],
            "chain_checks": {
                "status": "FAIL",
                "error_type": "JSONDecodeError",
                "error_message": "Expecting value: line 1 column 1 (char 0)",
            },
        },
        "bundle_mode": "debug_bundle",
        "suggested_rerun_command": "python tools/local_regression/run_all.py --only chain_checks",
    })

    summary_text = _capture_output(lambda: test_suite_module._print_human_summary(result_payload))
    add_check(results, "reporting_schema", case_id, "test_suite_unreadable_payload_surfaces_reason", True, "summary_unreadable" in summary_text and "error_type=JSONDecodeError" in summary_text and "Expecting value: line 1 column 1 (char 0)" in summary_text)
    add_check(results, "reporting_schema", case_id, "test_suite_unreadable_payload_formats_failed_step_labels", True, "失敗步驟 : chain checks(chain_checks)" in summary_text and "建議重跑 : python tools/local_regression/run_all.py --only chain_checks" in summary_text)

    summary["unreadable_payload_summary_lines"] = len([line for line in summary_text.splitlines() if line.strip()])
    return results, summary
