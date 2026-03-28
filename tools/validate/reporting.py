import os

import pandas as pd


def write_issue_excel_report(df_failed, df_failed_summary, df_failed_module, timestamp, *, output_dir, normalize_ticker):
    from openpyxl.styles import numbers

    if df_failed.empty:
        return None

    report_path = os.path.join(output_dir, f"consistency_issues_{timestamp}.xlsx")

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        df_failed.to_excel(writer, sheet_name="failed_only", index=False)
        df_failed_summary.to_excel(writer, sheet_name="failed_tickers", index=False)
        df_failed_module.to_excel(writer, sheet_name="failed_modules", index=False)

        for sheet_name in ["failed_only", "failed_tickers"]:
            ws = writer.book[sheet_name]
            header_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
            if "ticker" in header_map:
                col_idx = header_map["ticker"]
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.number_format = numbers.FORMAT_TEXT
                    cell.value = normalize_ticker(cell.value)

    return report_path


def print_console_summary(
    df_results,
    df_failed,
    df_summary,
    csv_path,
    xlsx_path,
    elapsed_time,
    real_summary_count,
    real_tickers,
    *,
    normalize_ticker_text,
    max_console_fail_preview,
):
    synthetic_summary_count = 0
    synthetic_ticker_set = set()
    if not df_summary.empty and "synthetic" in df_summary.columns:
        synthetic_mask = df_summary["synthetic"].astype("boolean").fillna(False).astype(bool)
        synthetic_summary_count = int(synthetic_mask.sum())
        synthetic_ticker_set = {
            normalize_ticker_text(ticker)
            for ticker in df_summary.loc[synthetic_mask, "ticker"].dropna().tolist()
        }

    system_summary_count = max(len(df_summary) - real_summary_count - synthetic_summary_count, 0)

    real_ticker_set = {normalize_ticker_text(ticker) for ticker in real_tickers}
    failed_ticker_series = (
        df_failed["ticker"].fillna("").map(normalize_ticker_text)
        if not df_failed.empty else pd.Series(dtype="object")
    )
    failed_real_mask = failed_ticker_series.isin(real_ticker_set) if not df_failed.empty else pd.Series(dtype=bool)
    failed_synthetic_mask = failed_ticker_series.isin(synthetic_ticker_set) if not df_failed.empty else pd.Series(dtype=bool)
    failed_system_mask = ~(failed_real_mask | failed_synthetic_mask) if not df_failed.empty else pd.Series(dtype=bool)

    failed_real_tickers = int(failed_ticker_series[failed_real_mask].nunique()) if not df_failed.empty else 0
    failed_synthetic_cases = int(failed_ticker_series[failed_synthetic_mask].nunique()) if not df_failed.empty else 0
    failed_system_items = int(failed_ticker_series[failed_system_mask].nunique()) if not df_failed.empty else 0

    pass_count = int((df_results["status"] == "PASS").sum()) if not df_results.empty else 0
    skip_count = int((df_results["status"] == "SKIP").sum()) if not df_results.empty else 0
    fail_count = int((df_results["status"] == "FAIL").sum()) if not df_results.empty else 0

    print("\n================================================================================")
    print("一致性回歸摘要")
    print("================================================================================")
    print(f"耗時: {elapsed_time:.2f} 秒")
    print(f"成功進入 summary 的真實股票數: {real_summary_count}")
    print(f"synthetic case 數: {synthetic_summary_count}")
    print(f"system 檢查列數: {system_summary_count}")
    print(f"summary 總列數: {len(df_summary)}")
    print(f"總檢查數: {len(df_results)}")
    print(f"PASS 數: {pass_count}")
    print(f"SKIP 數: {skip_count}")
    print(f"FAIL 數: {fail_count}")
    print(f"有問題真實股票數: {failed_real_tickers}")
    print(f"有問題 synthetic case 數: {failed_synthetic_cases}")
    print(f"有問題 system 項目數: {failed_system_items}")
    print(f"完整 CSV: {csv_path}")
    print(f"問題 Excel: {xlsx_path if xlsx_path else '無，因為沒有 failed 項'}")

    if df_failed.empty:
        print("\n失敗項摘要：無")
        return

    print("\n失敗項前覽：")
    show_cols = ["ticker", "module", "metric", "expected", "actual", "note"]
    preview_df = df_failed[show_cols].head(max_console_fail_preview).copy()
    print(preview_df.to_string(index=False))

    remain_count = len(df_failed) - len(preview_df)
    if remain_count > 0:
        print(f"\n... 尚有 {remain_count} 筆 FAIL 未顯示，請直接查看 CSV / Excel。")

    failed_real_summary = (
        df_failed.loc[failed_real_mask]
        .groupby("ticker", dropna=False)
        .agg(failed_checks=("passed", "size"))
        .reset_index()
        .sort_values(by=["failed_checks", "ticker"], ascending=[False, True])
        .head(max_console_fail_preview)
        if not df_failed.empty else pd.DataFrame()
    )

    if not failed_real_summary.empty:
        print("\n失敗真實股票前覽：")
        print(failed_real_summary.to_string(index=False))

    failed_non_real_summary = (
        df_failed.loc[~failed_real_mask]
        .groupby("ticker", dropna=False)
        .agg(failed_checks=("passed", "size"))
        .reset_index()
        .sort_values(by=["failed_checks", "ticker"], ascending=[False, True])
        .head(max_console_fail_preview)
        if not df_failed.empty else pd.DataFrame()
    )

    if not failed_non_real_summary.empty:
        print("\n失敗 synthetic/system 前覽：")
        print(failed_non_real_summary.to_string(index=False))
